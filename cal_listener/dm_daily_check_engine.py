"""
Delivery Master daily check.

Loops through the In Progress, Katie, Steven, Kyle, and Jamie C filter views.
For each view, reads every row (handling Telerik's row virtualization with
page-down scrolling), then flags any row that meets ANY of:
  1. Delivery date is in the past AND status is not "POD"
  2. Cust. Ref is empty
  3. Cust. Ref is "TBC" (case-insensitive)
  4. Cust. Ref contains "chased" (case-insensitive)

Outputs:
  - dm_daily_report.html  (open in browser, copy into Outlook)
  - dm_daily_report.eml   (double-click to open an Outlook draft pre-filled
                           with subject, To, and body — just press Send)
"""
import sys
import os
import subprocess
import csv
import gc
import re
import time
import email.mime.multipart
import email.mime.text
import faulthandler
from contextlib import contextmanager
from pathlib import Path
from collections import defaultdict
from datetime import datetime, date


# Disable Python's cyclic garbage collector for the entire process lifetime.
#
# Background: pywinauto wraps each UIA element in a Python object whose
# parent/child references form REFERENCE CYCLES (a wrapped DataItem holds its
# wrapped cells, each cell holds a back-reference to its DataItem). Plain
# refcounting can't free cycles, so those wrappers stick around until the
# cyclic garbage collector runs. By that time, DM's RadGridView virtualization
# has often already recycled the underlying row visuals — the COM pointer
# inside each wrapper now points at freed memory. When GC then calls __del__
# on the cycle, comtypes calls IUnknown::Release on that stale pointer, which
# segfaults the entire process with ACCESS_VIOLATION (Windows code 3221225477).
#
# Owen's PC happens to ship a pywinauto/comtypes/Telerik combo where this race
# rarely lines up. A colleague's miniconda Python ships a combo where it
# ALWAYS lines up. The robust fix is to never let cyclic GC run inside this
# script — refcount cleanup still happens for everything that isn't part of a
# cycle, the cycles leak (harmless in a short-lived worker), and at the end
# of the worker we use os._exit(0) so Python's finalization GC sweep is
# bypassed entirely. The OS just reaps the process.
try:
    gc.disable()
except Exception:
    pass


@contextmanager
def gc_paused():
    """Historical marker that this section touches UIA. Body is now a no-op
    because cyclic GC is disabled module-wide (see the gc.disable() call
    above). We keep the context manager so the call sites still read as
    'this is a UIA hot path that must not be interrupted by GC', and so that
    if GC ever has to be re-enabled in this script the call sites are already
    there to defend the hot paths."""
    yield

# Enable faulthandler IMMEDIATELY. When pywinauto / UIAutomationCore native-
# crashes the process with an access violation (Windows code 3221225477), we
# get nothing in the console — Python exits silently and we never learn which
# line triggered the crash. With faulthandler enabled, Python intercepts the
# signal and prints a Python-level stack trace before exiting, telling us
# exactly which UIA call was on the stack at the moment of the crash. This is
# essential for diagnosing crashes on machines where pywinauto is fragile.
try:
    faulthandler.enable(all_threads=True)
except Exception:
    pass

# Mark the Python process as DPI-aware BEFORE we do anything that interacts
# with the screen. Without this, ImageGrab.grab() can sample physical-pixel
# coordinates while UIA gives us logical-pixel rects, so the header screenshot
# captures the wrong region (or a blank one) on PCs running display scaling
# above 100%. This is the actual cause of "OCR recognised 0 words" on those
# machines.
try:
    import ctypes
    try:
        # Per-monitor v2 (best on Windows 10+)
        ctypes.windll.user32.SetProcessDpiAwarenessContext(-4)
    except Exception:
        try:
            # Process system DPI aware (Win 8.1+)
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                # Plain DPI-aware (Vista+)
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass
except Exception:
    pass

try:
    from pywinauto import Desktop  # type: ignore
    from pywinauto.controls.uiawrapper import UIAWrapper  # type: ignore
    from pywinauto.keyboard import send_keys  # type: ignore
except ImportError:
    # In a frozen .exe `sys.executable` is the listener itself, so a
    # subprocess pip install would just launch another daemon and hit
    # the singleton mutex. Skip the fallback when frozen — pywinauto
    # is bundled via --collect-all, so this only fires if something
    # broke at build time.
    if getattr(sys, "frozen", False):
        raise
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pywinauto"])
    from pywinauto import Desktop  # type: ignore
    from pywinauto.controls.uiawrapper import UIAWrapper  # type: ignore
    from pywinauto.keyboard import send_keys  # type: ignore

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except ImportError:
    if getattr(sys, "frozen", False):
        raise
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook, load_workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

# PIL for screenshotting the header strip (used by OCR-based column detection).
try:
    from PIL import ImageGrab  # type: ignore
except ImportError:
    if getattr(sys, "frozen", False):
        # OCR is optional — the v2 dm_columns resolver doesn't need it.
        # Just leave ImageGrab as None; no pip install in frozen mode
        # (it would re-launch CalListener.exe and hit the singleton).
        ImageGrab = None  # type: ignore
    else:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "Pillow"])
            from PIL import ImageGrab  # type: ignore
        except Exception:
            ImageGrab = None  # type: ignore

# v2 column resolver (content-first, screen-independent). Demotes the fragile
# OCR/heuristic path to a fallback. If the module is missing for any reason we
# leave the helpers as None and fall back to the legacy detect_columns path.
try:
    from dm_columns import (  # type: ignore
        resolve_columns as _resolve_columns_v2,
        header_map_from_labels as _header_map_from_labels,
        canonical_from_header as _canonical_from_header_v2,
    )
except Exception:
    try:
        import importlib.util as _ilu
        _spec = _ilu.spec_from_file_location(
            "dm_columns", str(Path(__file__).parent / "dm_columns.py"))
        _mod = _ilu.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)  # type: ignore
        # CRITICAL: register the loaded module in sys.modules so later
        # `from dm_columns import X` calls (e.g. in
        # _load_tms_customer_names's sidecar path) succeed. Without this
        # the module exists as a local but the import machinery can't
        # find it and we fall back to content heuristics — which then
        # gets Customer / Cust. Ref backwards on tricky views.
        sys.modules["dm_columns"] = _mod
        _resolve_columns_v2 = _mod.resolve_columns
        _header_map_from_labels = _mod.header_map_from_labels
        _canonical_from_header_v2 = _mod.canonical_from_header
    except Exception as _ce:
        print(f"[dm_columns] v2 resolver unavailable ({_ce}); using legacy detection")
        _resolve_columns_v2 = None
        _header_map_from_labels = None
        _canonical_from_header_v2 = None

# Shared-data client (records run history + result files to Supabase). Optional:
# if it can't be imported the checker still runs, just without cloud recording.
try:
    _ROOT = Path(__file__).resolve().parent.parent.parent
    if str(_ROOT) not in sys.path:
        sys.path.insert(0, str(_ROOT))
    import cloud_sync as _cloud  # type: ignore
except Exception as _cse:
    print(f"[cloud_sync] unavailable ({_cse}); run history will not be recorded")
    _cloud = None


SCRIPT_DIR = Path(__file__).parent
# When running inside a PyInstaller --onefile bundle, __file__ lives in a
# random temp extract that vanishes when the process exits AND isn't shared
# between sibling subprocesses. So redirect SCRIPT_DIR/HERE to a stable
# %APPDATA%\CalListener\dm_workdir folder so:
#   - the orchestrator and its per-view subprocesses see the same
#     view_results directory,
#   - the listener daemon that spawned the orchestrator can read those
#     per-view JSONs after the engine exits.
# (The temp extract is still used for loading dm_columns.py — see the
#  Path(__file__).parent fallback in the v2 column resolver import above.)
if getattr(sys, "frozen", False):
    _appdata = Path(os.environ.get("APPDATA", str(Path.home())))
    SCRIPT_DIR = _appdata / "CalListener" / "dm_workdir"
    SCRIPT_DIR.mkdir(parents=True, exist_ok=True)
# Outputs go to the parent folder when we're inside a 'scripts' subfolder;
# otherwise we write next to ourselves (single-folder layout).
HERE = (
    SCRIPT_DIR.parent
    if SCRIPT_DIR.name.lower() in ("scripts", "_internal")
    else SCRIPT_DIR
)
VIEWS = ["In Progress", "Katie", "Steven", "Kyle", "Jamie C", "Complete"]
# Statuses that count as "delivered" — these rows are NOT flagged as overdue
# (cust-ref issues are still flagged separately).
DELIVERED_STATUSES = {"POD", "Complete"}
EMAIL_TO = "lauren@cal.delivery"
RULES_PATH = SCRIPT_DIR / "dm_rules.xlsx"

# Regex & known values.
# BT_RE accepts both the displayed form ("BT62882") and the raw clipboard form
# ("62882") so the same detection logic works against either data source.
INDEX_RE = re.compile(r"Column Display Index:\s*(\d+)")
BT_RE = re.compile(r"^(?:BT)?\d{4,8}$")
ON_CX_RE = re.compile(r"^On CX\b")
# Dates: displayed form is "DD-MM-YY HH:MM"; clipboard form is "DD-MM-YYYY HH:MM:SS".
DATE_RE = re.compile(r"^\d{2}-\d{2}-\d{2,4}(\s+\d{2}:\d{2}(:\d{2})?)?")
DEC_RE = re.compile(r"^-?\d+(\.\d+)?$")
POSTCODE_RE = re.compile(r"\b[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}\b")
STATUS_VALUES = {
    "Waiting", "Allocated", "POB", "POB Air", "POB On Way",
    "Part POD", "POD", "Complete", "Quoted",
}


def _load_tms_customer_names(company=None):
    """Load known TMS customer names for a company ('north'/'south').

    Two sources, tried in order:

      1. The desktop's local `invoice_store` module (present when the
         engine runs as part of the desktop CalToolkit).
      2. `tms_customers_<company>.json` in SCRIPT_DIR — written by the
         cal_listener handler before each scrape, populated from the
         Supabase `customer_profiles` dataset. This is the listener's
         equivalent and the path that fires for the bundled .exe.

    Returns (company, normalised-set). Empty set on any failure - the
    resolver then falls back to content heuristics, so this never
    blocks a run, but column disambiguation will be weaker (we saw
    Customer/Cust.Ref swap on the Steven view without this list).
    """
    co = company or os.environ.get("DM_COMPANY") or ""

    # Source 1: desktop's invoice_store.
    try:
        ir_dir = SCRIPT_DIR.parent / "invoicing_rules"
        if str(ir_dir) not in sys.path:
            sys.path.insert(0, str(ir_dir))
        import invoice_store  # type: ignore
        if not co:
            co = invoice_store.get_active_company()
        rows = invoice_store.list_customers(co)
        names = [r.get("name", "") for r in rows if r.get("name")]
        if _resolve_columns_v2 is not None:
            from dm_columns import normalise_customer_names as _ncn  # type: ignore
            return co, _ncn(names)
        return co, set()
    except Exception as e:
        # Fall through to source 2 — don't print a scary failure line
        # yet, the JSON sidecar might cover us.
        invoice_store_err = e

    # Source 2: listener-side JSON sidecar.
    try:
        json_co = co or "north"
        sidecar = SCRIPT_DIR / f"tms_customers_{json_co}.json"
        if sidecar.exists():
            import json as _json
            with sidecar.open(encoding="utf-8") as f:
                blob = _json.load(f)
            names = blob.get("names") or []
            if _resolve_columns_v2 is not None:
                from dm_columns import normalise_customer_names as _ncn  # type: ignore
                normed = _ncn(names)
            else:
                normed = set()
            print(f"  TMS customer-list loaded from {sidecar.name}: "
                  f"{len(names)} names", flush=True)
            return json_co, normed
    except Exception as je:
        print(f"  (TMS sidecar load also failed: {je})", flush=True)

    print(f"  (TMS customer-list load failed: {invoice_store_err}; "
          "using content heuristics)", flush=True)
    return (co or "north"), set()


# ---------- UIA helpers ----------

def find_dm():
    for w in Desktop(backend="uia").windows():
        try:
            t = w.window_text() or ""
            if "Cal (" in t or "Delivery Master" in t:
                return w
        except Exception:
            pass
    return None


def find_first(elem, predicate, max_depth=14):
    stack = [(elem, 0)]
    while stack:
        e, d = stack.pop()
        try:
            if predicate(e.element_info):
                return e
            if d < max_depth:
                for c in reversed(e.children()):
                    stack.append((c, d + 1))
        except Exception:
            pass
    return None


def find_button(dm, name):
    return find_first(dm, lambda i:
        (i.control_type or "") in ("Button", "ToggleButton", "RadioButton", "ListItem", "MenuItem")
        and (i.name or "").strip() == name
    )


def find_grid_and_panel(dm):
    grid = find_first(dm, lambda i: (i.class_name or "") == "RadGridView")
    if grid is None:
        return None, None
    panel = None
    for c in grid.children():
        try:
            if (c.element_info.class_name or "") == "GridViewVirtualizingPanel":
                panel = c
                break
        except Exception:
            pass
    return grid, panel


# ---------- Clipboard fast path ----------

import ctypes  # noqa: E402  (late import so the file works without it on first-import errors)
from ctypes import wintypes  # noqa: E402


def _read_clipboard_text():
    """Return Unicode text from the Windows clipboard, or None."""
    CF_UNICODETEXT = 13
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    user32.OpenClipboard.argtypes = [wintypes.HWND]
    user32.OpenClipboard.restype = wintypes.BOOL
    user32.GetClipboardData.argtypes = [wintypes.UINT]
    user32.GetClipboardData.restype = wintypes.HANDLE
    user32.CloseClipboard.restype = wintypes.BOOL
    user32.IsClipboardFormatAvailable.argtypes = [wintypes.UINT]
    kernel32.GlobalLock.argtypes = [wintypes.HANDLE]
    kernel32.GlobalLock.restype = ctypes.c_void_p
    kernel32.GlobalUnlock.argtypes = [wintypes.HANDLE]
    if not user32.IsClipboardFormatAvailable(CF_UNICODETEXT):
        return None
    if not user32.OpenClipboard(0):
        return None
    try:
        handle = user32.GetClipboardData(CF_UNICODETEXT)
        if not handle:
            return None
        pointer = kernel32.GlobalLock(handle)
        if not pointer:
            return None
        try:
            return ctypes.c_wchar_p(pointer).value
        finally:
            kernel32.GlobalUnlock(handle)
    finally:
        user32.CloseClipboard()


def _empty_clipboard():
    user32 = ctypes.windll.user32
    user32.OpenClipboard.argtypes = [wintypes.HWND]
    user32.OpenClipboard.restype = wintypes.BOOL
    user32.EmptyClipboard.restype = wintypes.BOOL
    user32.CloseClipboard.restype = wintypes.BOOL
    if not user32.OpenClipboard(0):
        return False
    try:
        return bool(user32.EmptyClipboard())
    finally:
        user32.CloseClipboard()


def _foreground_window_title():
    try:
        user32 = ctypes.windll.user32
        user32.GetForegroundWindow.restype = wintypes.HWND
        user32.GetWindowTextLengthW.argtypes = [wintypes.HWND]
        user32.GetWindowTextLengthW.restype = ctypes.c_int
        user32.GetWindowTextW.argtypes = [wintypes.HWND, ctypes.c_wchar_p, ctypes.c_int]
        hwnd = user32.GetForegroundWindow()
        length = user32.GetWindowTextLengthW(hwnd) + 1
        if length <= 1:
            return ""
        buf = ctypes.create_unicode_buffer(length)
        user32.GetWindowTextW(hwnd, buf, length)
        return buf.value
    except Exception:
        return "<unknown>"


def _dm_is_foreground():
    """True if Delivery Master currently owns the foreground window. DM's
    title contains 'Cal (' or 'Delivery Master' (see find_dm)."""
    t = _foreground_window_title() or ""
    return ("Cal (" in t) or ("Delivery Master" in t)


def _ensure_dm_foreground(dm, grid, cell=None, attempts=5, verbose=True):
    """Bring DM to the foreground and VERIFY it actually came forward before we
    send keystrokes. This is the defence against the focus-steal race (another
    app — we observed Excel doing exactly this — grabbing focus between the
    click and the copy, which leaves the clipboard empty and forces the slow
    scroll fallback). Returns True if DM is foreground when we give up trying.
    """
    for i in range(attempts):
        try:
            if dm is not None:
                UIAWrapper(dm.element_info).set_focus()
                time.sleep(0.25)
            if grid is not None:
                UIAWrapper(grid.element_info).set_focus()
                time.sleep(0.2)
            if cell is not None and i == 0:
                # A real click lands keyboard focus on a row; only needed once.
                UIAWrapper(cell.element_info).click_input()
                time.sleep(0.3)
        except Exception as e:
            if verbose:
                print(f"      focus attempt {i+1} warning: {e}", flush=True)
        if _dm_is_foreground():
            return True
        if verbose:
            print(f"      DM not foreground (saw {_foreground_window_title()!r}); "
                  f"re-asserting [{i+1}/{attempts}]", flush=True)
        time.sleep(0.4 * (i + 1))  # backoff
    return _dm_is_foreground()


def _first_data_cell(panel):
    """Find a clickable cell inside the first visible data row.

    Wrapped in gc_paused() because the descendants() walk here was one of the
    crash sites in v35 (the In Progress clipboard-mode crash).
    """
    with gc_paused():
        for di in panel.children():
            try:
                if di.element_info.control_type != "DataItem":
                    continue
            except Exception:
                continue
            for child in di.descendants():
                try:
                    if (child.element_info.class_name or "") == "GridViewCell":
                        return child
                except Exception:
                    pass
            return di
    return None


def _parse_clipboard_tsv(text):
    """Parse Telerik's clipboard output (tab-separated rows) into list of {idx: value}."""
    rows = []
    for line in text.splitlines():
        if not line.strip():
            continue
        cells = line.split("\t")
        rows.append({idx: cell.strip() for idx, cell in enumerate(cells)})
    return rows


def _looks_like_grid_text(text: str) -> bool:
    """Heuristic: does this clipboard text look like a Telerik grid copy?

    Telerik exports rows as tab-separated values with newlines between rows.
    A single line of arbitrary text (e.g. stale clipboard from Excel or a
    chat app) wouldn't have tabs OR would only be one line.

    The check is intentionally loose: any text with at least one tab AND
    at least two lines passes. A single-row view would fail this, but in
    practice DM views always have headers OR multiple rows."""
    if not text:
        return False
    if "\t" not in text:
        return False
    lines = [l for l in text.splitlines() if l.strip()]
    return len(lines) >= 2


def read_view_via_clipboard(grid, panel, dm=None, verbose=True):
    """Focus the grid, send Ctrl+A then Ctrl+C, parse the clipboard.
    Returns a list of row dicts on success, or None on failure.

    Three attempts with escalating waits and re-focus between each. The
    text must look like a TSV (tab-separated, multi-line) to count - this
    rejects stale clipboard data left over from another app.

    The whole UIA-touching section runs under gc_paused() so the cyclic
    garbage collector can't fire mid-call and free a stale COM pointer.
    See the gc_paused() docstring for full background.
    """
    # Per attempt: (after-Ctrl+A wait, after-Ctrl+C MAX wait). The
    # Ctrl+C wait is now a CEILING - we poll the clipboard every 0.3s
    # and break out as soon as it has grid-shaped content. This is the
    # real fix for the "timer icon" failure: Telerik shows a busy
    # cursor while it serializes the selection, and the clipboard
    # doesn't get populated until that finishes. A fixed sleep that
    # runs out before serialization completes always returns empty.
    BASE_ATTEMPTS = (
        (0.6, 6.0),    # attempt 1
        (1.2, 10.0),   # attempt 2 - longer ceilings for slow Telerik runs
        (2.0, 15.0),   # attempt 3 - generous for Complete (~1000 rows)
    )
    # Scale every wait by the per-machine timing multiplier set by the
    # calibration probe. Slow machines get longer waits without any
    # plugin code change.
    try:
        import sys as _sys
        _root = Path(__file__).resolve().parent.parent.parent
        if str(_root) not in _sys.path:
            _sys.path.insert(0, str(_root))
        import calibration as _cal_mod  # type: ignore
        _mult = _cal_mod.load().timing_multiplier
    except Exception:
        _mult = 1.0
    ATTEMPTS = tuple(
        (round(a * _mult, 2), round(c * _mult, 2))
        for a, c in BASE_ATTEMPTS)
    POLL_INTERVAL = 0.3
    with gc_paused():
        cell = _first_data_cell(panel)
        if cell is None:
            if verbose:
                print("      no cell found")
            return None

        text = ""
        for attempt_idx, (wait_a, wait_c) in enumerate(ATTEMPTS, start=1):
            # Pull DM forward and verify it stuck. Without this, a window
            # that steals focus between here and the copy (we watched Excel
            # do exactly that) leaves the clipboard empty and forces the
            # slow scroll fallback.
            if not _ensure_dm_foreground(dm, grid, cell=cell, verbose=verbose):
                if verbose:
                    print(f"      attempt {attempt_idx}: DM still not "
                          f"foreground (saw "
                          f"{_foreground_window_title()!r}); proceeding",
                          flush=True)
            time.sleep(0.4 if attempt_idx == 1 else 0.6)

            if verbose:
                print(f"      attempt {attempt_idx}: foreground window "
                      f"is {_foreground_window_title()!r}")

            _empty_clipboard()
            send_keys("^a")
            time.sleep(wait_a)
            send_keys("^c")
            # Poll the clipboard up to wait_c seconds. Break out as soon
            # as we see grid-shaped content. This is the right approach
            # for Telerik: it shows a busy "timer" cursor while it
            # serializes the selection to clipboard, and the content
            # appears only after that finishes. A fixed sleep that runs
            # out before serialization completes always returns empty -
            # which was the user-visible failure ("highlights the page
            # but clipboard says nothing there").
            deadline = time.time() + wait_c
            text = ""
            poll_count = 0
            while time.time() < deadline:
                time.sleep(POLL_INTERVAL)
                poll_count += 1
                text = _read_clipboard_text()
                if _looks_like_grid_text(text):
                    if verbose:
                        n_lines = len([
                            ln for ln in text.splitlines() if ln.strip()])
                        elapsed = wait_c - (deadline - time.time())
                        print(f"      attempt {attempt_idx}: clipboard "
                              f"populated after {elapsed:.1f}s "
                              f"({poll_count} polls), {len(text)} chars, "
                              f"{n_lines} non-empty lines")
                    break
            else:
                # Loop completed without break - timed out. Fall through
                # to retry logging below.
                pass

            if _looks_like_grid_text(text):
                break
            # If we got some text but it doesn't look like TSV, that's
            # almost certainly a stale clipboard from another app - log
            # what we saw so the run log is informative.
            if text and verbose:
                preview = text[:60].replace("\n", " | ").replace("\t", "->")
                print(f"      attempt {attempt_idx}: clipboard had "
                      f"{len(text)} chars but doesn't look like a grid "
                      f"copy (preview: {preview!r}); retrying")
            elif verbose:
                print(f"      attempt {attempt_idx}: clipboard still "
                      f"empty after {wait_c}s of polling "
                      f"(foreground was {_foreground_window_title()!r}); "
                      f"retrying")

        # Click the first cell again to clear the 'all rows selected'
        # state. If that state is left active the next view-button click
        # can crash the pywinauto/comtypes stack. We do this even on
        # failure - the state was set when we sent Ctrl+A.
        try:
            UIAWrapper(cell.element_info).click_input()
            time.sleep(0.2)
        except Exception:
            pass

        if not _looks_like_grid_text(text):
            if verbose:
                print(f"      clipboard failed after "
                      f"{len(ATTEMPTS)} attempts; caller will fall back "
                      f"to scroll mode")
            return None

    return _parse_clipboard_tsv(text)


def _check_stop_file():
    """True if the user dropped a STOP.txt in the outputs folder."""
    return (HERE / "STOP.txt").exists()


def _check_ocr_backend():
    """Probe whether either OCR backend is usable. Returns the name of the
    available backend, or None. Tries to auto-install winsdk if missing."""
    # WinRT (built into Windows 10/11 — no separate install needed if the
    # winsdk Python package is available).
    try:
        import winsdk.windows.media.ocr as _ocr  # noqa: F401
        return "WinRT (winsdk)"
    except ImportError:
        pass
    try:
        import winrt.windows.media.ocr as _ocr  # noqa: F401
        return "WinRT (winrt)"
    except ImportError:
        pass
    # Try a one-off pip install of winsdk; on Python <= 3.11 this works
    # cleanly, on 3.12+ there's no wheel and it'll fail — that's fine.
    # Suppress all output so the user doesn't see the wall of CMake errors.
    try:
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet", "winsdk"],
            timeout=60,
            capture_output=True,
            text=True,
        )
        if result.returncode == 0:
            try:
                import winsdk.windows.media.ocr as _ocr  # noqa: F401
                return "WinRT (winsdk, just installed)"
            except ImportError:
                pass
    except Exception:
        pass
    # Tesseract
    try:
        import pytesseract  # type: ignore  # noqa: F401
        for p in (r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                  r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"):
            if Path(p).exists():
                return "Tesseract"
        result = subprocess.run(
            ["tesseract", "--version"], capture_output=True, timeout=5
        )
        if result.returncode == 0:
            return "Tesseract (PATH)"
    except Exception:
        pass
    return None


# Highest "total" value the script saw in the pagination indicator
# during the most recent read. Set by read_all_rows_via_clipboard /
# read_all_rows_paged, consumed by process_view for data-validation.
# A separate value per worker subprocess - we run one view per process.
_LAST_RUN_MAX_TOTAL: int | None = None


def read_all_rows_via_clipboard(dm, grid, panel, max_server_pages=15, on_page=None):
    """Read all rows across all server pages using the Ctrl+A+Ctrl+C fast path.
    Returns a list of row dicts, or None if the clipboard approach failed at
    the start (caller can then fall back to the scroll-based reader).

    If `on_page` is supplied, it's called as on_page(page_number, rows_so_far)
    AFTER each successful page read but BEFORE the next-page click. This lets
    the caller checkpoint to disk so a subsequent native crash on page-N+1
    navigation doesn't lose page-N's data.

    Side effect: stores the highest 'total' value read from the on-screen
    pagination indicator in the module-level _LAST_RUN_MAX_TOTAL, so
    process_view can use it as ground-truth for data-validation even if
    the run aborts on a later page (the indicator may go to 0 after a
    failed clipboard read).
    """
    global _LAST_RUN_MAX_TOTAL
    _LAST_RUN_MAX_TOTAL = None
    all_rows = {}
    last_pagination = None
    last_total = None
    # Defensive counters - DM occasionally bugs its own pagination indicator
    # (e.g. on the Complete view it has shown "1-52 of 352" on page 2 when it
    # should have been "301-352 of 352"). Without these guards the loop spins
    # until max_server_pages, which can take minutes.
    last_start = None         # for backward-progress detection
    zero_added_streak = 0     # for stuck-on-same-rows detection
    for server_page in range(max_server_pages):
        if _check_stop_file():
            print("    STOP.txt detected — stopping pagination loop.", flush=True)
            break
        if server_page > 0:
            print(f"  -- Server page {server_page + 1} --")
            # Re-find the panel after page change — the underlying DataItems were rebuilt
            new_panel = None
            for c in grid.children():
                try:
                    if (c.element_info.class_name or "") == "GridViewVirtualizingPanel":
                        new_panel = c
                        break
                except Exception:
                    pass
            if new_panel is not None:
                panel = new_panel

        page_rows = read_view_via_clipboard(grid, panel, dm=dm)
        # If the clipboard read failed on a NON-FIRST page, the page has
        # navigated but the grid isn't ready yet (the Complete view's
        # ~300-row pages take noticeably longer to render). Wait, refresh
        # the grid + panel references, and retry up to twice before
        # falling back to scroll-mode for this single page. Without this
        # the run gives up at the first stall and discards 2+ pages of
        # real data (Owen's Complete view: 300/906 instead of 906/906).
        if page_rows is None and server_page > 0:
            print(
                f"    page-{server_page + 1} clipboard read returned "
                "nothing; rendering may not be complete — retrying.",
                flush=True)
            for retry in (1, 2):
                time.sleep(2.5 + retry * 1.5)
                # Re-find grid first (DM rebuilds the whole grid on
                # some page changes, not just the panel).
                try:
                    for c in dm.descendants():
                        if (c.element_info.class_name or "") == "GridViewDataControl":
                            grid = c
                            break
                except Exception:
                    pass
                for c in grid.children():
                    try:
                        if (c.element_info.class_name or "") == "GridViewVirtualizingPanel":
                            panel = c
                            break
                    except Exception:
                        pass
                page_rows = read_view_via_clipboard(
                    grid, panel, dm=dm)
                if page_rows is not None:
                    print(
                        f"    retry {retry} succeeded - clipboard "
                        "produced rows.", flush=True)
                    break
            if page_rows is None:
                # Last resort: read this page via the scroll/UIA path so
                # we don't lose it entirely. Slower but reliable.
                print(
                    f"    clipboard still failing - falling back to "
                    "scroll-mode for this single page.", flush=True)
                try:
                    page_rows = read_all_rows(dm, grid, panel)
                except Exception as e:
                    print(
                        f"    scroll-mode fallback also errored: {e}",
                        flush=True)
                    page_rows = None
        if page_rows is None:
            if server_page == 0:
                return None  # never got anything; let caller fall back
            print(f"    clipboard returned nothing on server page {server_page + 1}; stopping")
            break

        added = 0
        for row in page_rows:
            if not any((v or "").strip() for v in row.values()):
                continue
            key = _row_dedup_key(row)
            if key not in all_rows:
                all_rows[key] = row
                added += 1
        print(f"    server page {server_page + 1}: parsed {len(page_rows)} rows, "
              f"added {added}, cumulative {len(all_rows)}")

        # Checkpoint AFTER read but BEFORE the next-page navigation, so a
        # crash on btnNext doesn't lose what we've collected.
        if on_page is not None:
            try:
                on_page(server_page + 1, list(all_rows.values()))
            except Exception as cb_err:
                print(f"    on_page callback errored: {cb_err}", flush=True)

        status = find_pagination_status(dm)
        if status is None:
            break
        start, end, total = status
        print(f"    pagination: {start}-{end} of {total}")
        # Remember the highest 'total' we've seen so data-validation
        # has the ground-truth row count even if a later page fails.
        if total > 0 and (
                _LAST_RUN_MAX_TOTAL is None or total > _LAST_RUN_MAX_TOTAL):
            globals()['_LAST_RUN_MAX_TOTAL'] = total

        # ---- Robustness layer 1: collected enough rows ----------------------
        # If we've already accumulated `total` or more deduplicated rows, the
        # job is done regardless of what the indicator now claims. This is
        # the strongest possible "we're finished" signal and beats every
        # display bug DM can throw at us.
        if total > 0 and len(all_rows) >= total:
            print(f"    collected {len(all_rows)} >= total {total}, stopping",
                  flush=True)
            break

        # ---- Normal end-of-dataset check ------------------------------------
        if end >= total:
            print(f"    reached end of dataset")
            break

        # ---- Robustness layer 2: bogus indicator ----------------------------
        # If start/end/total don't form a sane range (e.g. DM returned -1, or
        # end < start, or total == 0) bail out. Continuing would just guess.
        if start <= 0 or end < start or total <= 0:
            print(f"    pagination indicator looks bogus "
                  f"({start}-{end} of {total}), stopping.", flush=True)
            break

        # ---- Robustness layer 3: forward-progress detection -----------------
        # The DM "reset" bug shows up as END decreasing (e.g. page 1 says
        # "1-300 of 352" and page 2 says "1-52 of 352"). That's a sure sign
        # the indicator has bugged and the loop will otherwise spin forever.
        #
        # An earlier version of this guard checked `start` for forward
        # progress, but that turned out to be wrong: on the Complete view
        # DM keeps start=1 and grows end cumulatively (1-300, then 1-600,
        # then 1-900...), and treating that as "no progress" stopped after
        # page 2 with two-thirds of the rows missing. The right test is
        # END must not decrease - start can stay still as long as end is
        # moving forward.
        if last_pagination is not None:
            last_s, last_e, _ = last_pagination
            if end < last_e:
                print(f"    pagination end went BACKWARDS "
                      f"({last_e} -> {end}); DM display bug suspected, "
                      f"stopping at {len(all_rows)} cumulative.", flush=True)
                break
            if start < last_s:
                print(f"    pagination start went BACKWARDS "
                      f"({last_s} -> {start}); DM display bug suspected, "
                      f"stopping at {len(all_rows)} cumulative.", flush=True)
                break
        last_start = start

        # ---- Stale-pagination guard (existing) ------------------------------
        # If the indicator hasn't moved since the previous page, the
        # next-page click didn't actually advance the grid. Without this
        # check the script spins forever (Telerik's clipboard returns
        # slightly different cached data each time, so the dedup keeps
        # "making progress" on the same page).
        if last_pagination is not None and (start, end, total) == last_pagination:
            print(f"    pagination didn't advance - assuming end of usable "
                  f"data, stopping.", flush=True)
            break
        last_pagination = (start, end, total)

        # ---- Robustness layer 4: zero-new-rows streak -----------------------
        # If two pages in a row added 0 new rows, the grid isn't really
        # paginating - we keep getting the same rows. Stop instead of
        # spinning. (One zero-page can be a transient clipboard glitch; two
        # in a row is a real problem.)
        if added == 0:
            zero_added_streak += 1
            if zero_added_streak >= 2:
                print(f"    no new rows for {zero_added_streak} pages in a "
                      f"row, stopping at {len(all_rows)} cumulative.",
                      flush=True)
                break
        else:
            zero_added_streak = 0

        nb_info = find_next_page_button(dm)
        if nb_info is None:
            print(f"    no next-page button found; stopping at {len(all_rows)} of {total}")
            break
        try:
            UIAWrapper(nb_info).invoke()
        except Exception:
            try:
                UIAWrapper(nb_info).click_input()
            except Exception as e:
                print(f"    next-page click failed: {e}")
                break
        # Bumped from 2.5s -> 4.0s. The Complete view's ~300-row server
        # pages take noticeably longer to re-render than the short
        # views; the old wait was leaving the grid half-built when the
        # next read fired, hence the "no cell found" failure.
        # _LAST_RUN_TIMING_MULT is applied if calibration set one.
        try:
            import calibration as _cal
            _mult = _cal.load().timing_multiplier
        except Exception:
            _mult = 1.0
        time.sleep(4.0 * _mult)

    return list(all_rows.values())


# ---------- Row reading ----------

def _find_ref_in_row(row):
    """Find an Our Ref value anywhere in the row.
    Accepts the displayed form 'BT62882' and the clipboard raw form '62882';
    always returns the displayed form ('BT'-prefixed) for stable identity."""
    for idx in sorted(row.keys()):
        val = (row[idx] or "").strip()
        if val and BT_RE.match(val):
            if val.isdigit() and len(val) <= 6:
                return "BT" + val
            return val
    return ""


def _row_dedup_key(row):
    """Stable identifier for a row. Prefers a BT-prefixed Our Ref; falls back to a
    content signature so we still dedup rows even when the Our Ref column is hidden
    on a particular view."""
    ref = _find_ref_in_row(row)
    if ref:
        return ("ref", ref)
    sig = tuple(sorted((idx, (v or "").strip()) for idx, v in row.items()))
    return ("sig", sig)


def read_visible_rows(panel):
    """Read all currently-rendered rows. Each row is {display_index: cell_text}.

    The full body runs under gc_paused() — every native crash trace we have
    shows cyclic GC firing inside one of these descendants() calls and
    releasing a stale UIA COM pointer. We let refcount-driven cleanup happen
    inside the block (that's harmless) but block the periodic sweep.
    """
    with gc_paused():
        data_items = []
        for c in panel.children():
            try:
                if c.element_info.control_type == "DataItem":
                    data_items.append(c)
            except Exception:
                pass

        rows = []
        for di in data_items:
            row = {}
            try:
                for cell in di.descendants():
                    try:
                        ci = cell.element_info
                        m = INDEX_RE.search(ci.name or "")
                        if not m:
                            continue
                        idx = int(m.group(1))
                        if idx in row and row[idx]:
                            continue
                        val = ""
                        for tb in cell.descendants():
                            try:
                                tbi = tb.element_info
                                if tbi.control_type == "Text" and tbi.name:
                                    val = tbi.name.strip()
                                    break
                            except Exception:
                                pass
                        row[idx] = val
                    except Exception:
                        pass
            except Exception:
                pass
            rows.append(row)
        return rows


def scroll_grid_one_page(grid, panel):
    """Scroll the grid down by one page. Returns True if it moved."""
    with gc_paused():
        # Try UIA Scroll pattern first
        try:
            wrapper = UIAWrapper(grid.element_info)
            try:
                wrapper.scroll("down", "page")
                return True
            except Exception:
                pass
        except Exception:
            pass
        # Fallback: focus and PageDown
        try:
            wrapper = UIAWrapper(panel.element_info)
            wrapper.set_focus()
            time.sleep(0.1)
            send_keys("{PGDN}")
            return True
        except Exception:
            pass
    return False


# ---- Server-side pagination (DM splits long views at 300 records / page) ----

PAGINATION_RE = re.compile(r"(\d+)\s*-\s*(\d+)\s*of\s*(\d+)", re.IGNORECASE)

# Substrings we'll match against a Button's automation name (case-insensitive)
NEXT_PAGE_NAME_HINTS = (
    "next page",
    "move to next page",
    "page next",
)


def _raw_descendants(start_elem, max_visit=15000):
    """Yield raw IUIAutomationElement pointers under start_elem (raw walker)."""
    from pywinauto.uia_defines import IUIA  # type: ignore
    raw_walker = IUIA().iuia.RawViewWalker
    visited = [0]

    def gen(ptr):
        if visited[0] > max_visit:
            return
        visited[0] += 1
        yield ptr
        try:
            c = raw_walker.GetFirstChildElement(ptr)
            while c:
                yield from gen(c)
                try:
                    c = raw_walker.GetNextSiblingElement(c)
                except Exception:
                    break
        except Exception:
            return

    try:
        yield from gen(start_elem.element_info.element)
    except Exception:
        return


def find_pagination_status(start_elem):
    """Read pagination from DM's labelled status elements (probed via AutomationId)."""
    max_records = None
    total_records = None
    for ptr in _raw_descendants(start_elem):
        try:
            auto_id = ptr.CurrentAutomationId or ""
        except Exception:
            continue
        if not auto_id:
            continue
        if auto_id == "lblMaxRecords":
            try:
                max_records = int((ptr.CurrentName or "").strip())
            except (ValueError, TypeError, AttributeError):
                pass
        elif auto_id == "lblTotalRecords":
            try:
                total_records = int((ptr.CurrentName or "").strip())
            except (ValueError, TypeError, AttributeError):
                pass
        if max_records is not None and total_records is not None:
            break
    if max_records is None or total_records is None:
        return None
    return (1, max_records, total_records)


def find_next_page_button(start_elem):
    """Find the pager 'Next' button (DM exposes it with AutomationId 'btnNext')."""
    from pywinauto.uia_element_info import UIAElementInfo  # type: ignore
    for ptr in _raw_descendants(start_elem):
        try:
            if (ptr.CurrentAutomationId or "") == "btnNext":
                return UIAElementInfo(ptr)
        except Exception:
            continue
    return None


def _get_value_pattern_value(ptr):
    """Read CurrentValue via UIA ValuePattern, returns '' if unavailable."""
    try:
        from comtypes import cast  # type: ignore
        from ctypes import POINTER
        from comtypes.gen.UIAutomationClient import IUIAutomationValuePattern  # type: ignore
        UIA_ValuePatternId = 10002
        vp = ptr.GetCurrentPattern(UIA_ValuePatternId)
        if not vp:
            return ""
        vpc = cast(vp, POINTER(IUIAutomationValuePattern))
        return vpc.CurrentValue or ""
    except Exception:
        return ""


def _dump_pager_candidates(dm):
    """Dump every named/interactive raw-walker element in the bottom region."""
    print("    --- pager-area diagnostic (raw walker) ---")
    try:
        dm_rect = dm.element_info.rectangle
        top_threshold = dm_rect.top + int((dm_rect.bottom - dm_rect.top) * 0.6)
        count = 0
        for ptr in _raw_descendants(dm, max_visit=30000):
            try:
                r = ptr.CurrentBoundingRectangle
                if r.top < top_threshold:
                    continue
                name = ptr.CurrentName or ""
                ctype = ptr.CurrentControlType
                val = _get_value_pattern_value(ptr)
                # Show anything with content OR any interactive control
                if not name and not val and ctype not in (50000, 50028, 50020, 50030):
                    continue
                cls = ptr.CurrentClassName or ""
                print(f"      ctype={ctype} cls={cls!r} name={name!r} value={val!r} rect=({r.left},{r.top},{r.right},{r.bottom})")
                count += 1
                if count > 80:
                    print("      ...(truncated)")
                    break
            except Exception:
                pass
        print(f"    --- end diagnostic ({count} entries) ---")
    except Exception as e:
        print(f"    diagnostic failed: {e}")


def read_all_rows_paged(dm, grid, panel, max_server_pages=15, on_page=None):
    """Read all rows across all server pages (DM caps the grid at ~300 rows / page).
    Within each server page, uses read_all_rows() to handle row virtualization scrolling.

    If `on_page` is supplied, it's called as on_page(page_number, rows_so_far)
    after each scroll-page within a server page. Same contract as the clipboard
    path — guarantees we save to disk after every chunk so a Phase 2 native
    crash doesn't throw away earlier work.

    Side effect: same as read_all_rows_via_clipboard - stashes the
    highest 'total' seen in _LAST_RUN_MAX_TOTAL for data-validation.
    """
    global _LAST_RUN_MAX_TOTAL
    # Don't reset if the clipboard path already set it - that read may
    # have seen the real total before we fell back here. Only reset if
    # nothing has captured one yet.
    if _LAST_RUN_MAX_TOTAL is None:
        _LAST_RUN_MAX_TOTAL = None
    all_rows = {}
    # Track total pages across server pages for consistent on_page numbering.
    total_pages_done = [0]

    def _wrap_on_page(scroll_page_idx, page_rows):
        total_pages_done[0] += 1
        # Merge this scroll-page's rows into all_rows for the checkpoint.
        for row in page_rows:
            if not any((v or "").strip() for v in row.values()):
                continue
            key = _row_dedup_key(row)
            if key not in all_rows:
                all_rows[key] = row
        if on_page is not None:
            try:
                on_page(total_pages_done[0], list(all_rows.values()))
            except Exception as cb_err:
                print(f"    on_page callback errored: {cb_err}", flush=True)

    for server_page in range(max_server_pages):
        if server_page > 0:
            print(f"  -- Server page {server_page + 1} --")

        page_rows = read_all_rows(dm, grid, panel, on_page=_wrap_on_page)
        added = 0
        for row in page_rows:
            if not any((v or "").strip() for v in row.values()):
                continue
            key = _row_dedup_key(row)
            if key not in all_rows:
                all_rows[key] = row
                added += 1
        print(f"    server page {server_page + 1}: added {added}, cumulative {len(all_rows)}")

        # Decide whether to move to the next server page
        status = find_pagination_status(dm)
        if status is None:
            if server_page == 0 and len(all_rows) >= 290:
                print(f"    NOTE: read {len(all_rows)} rows (close to 300 cap). Dumping pager area:")
                _dump_pager_candidates(dm)
            print(f"    no pagination indicator -> single-page view, stopping")
            break
        start, end, total = status
        print(f"    pagination: {start}-{end} of {total}")
        if total > 0 and (
                _LAST_RUN_MAX_TOTAL is None or total > _LAST_RUN_MAX_TOTAL):
            globals()['_LAST_RUN_MAX_TOTAL'] = total
        if end >= total:
            print(f"    reached end of dataset")
            break

        # Click next page
        nb_info = find_next_page_button(dm)
        if nb_info is None:
            print(f"    no next-page button found; stopping at {len(all_rows)} of {total}")
            _dump_pager_candidates(dm)
            break
        try:
            UIAWrapper(nb_info).invoke()
        except Exception as e:
            try:
                UIAWrapper(nb_info).click_input()
            except Exception as e2:
                print(f"    next-page click failed: {e} / {e2}")
                break
        time.sleep(2.5)  # let the new page load

        # Refresh panel reference (DataItems are rebuilt when the page changes)
        for c in grid.children():
            try:
                if (c.element_info.class_name or "") == "GridViewVirtualizingPanel":
                    panel = c
                    break
            except Exception:
                pass

    return list(all_rows.values())


def read_all_rows(dm, grid, panel, max_pages=25, on_page=None):
    """Scroll through the grid and collect all unique rows by Our Ref.

    If `on_page` is given, it's called as on_page(page_index, page_rows) after
    each scroll page is read — used by callers to checkpoint to disk so a
    native crash on the next scroll doesn't lose what we already have."""
    # Try to scroll to the top first. set_focus calls into UIAutomationCore
    # and is one of the calls we've seen native-crash mid-call when cyclic GC
    # fires inside it — gc_paused() prevents that.
    with gc_paused():
        try:
            UIAWrapper(grid.element_info).set_focus()
            time.sleep(0.2)
            send_keys("^{HOME}")
            time.sleep(0.4)
        except Exception:
            pass

    all_rows = {}  # our_ref -> row dict
    pages_without_progress = 0
    for page in range(max_pages):
        # Re-find panel each iteration in case it gets recreated. grid.children()
        # is another UIA walk that must be GC-guarded.
        panel = None
        with gc_paused():
            for c in grid.children():
                try:
                    if (c.element_info.class_name or "") == "GridViewVirtualizingPanel":
                        panel = c
                        break
                except Exception:
                    pass
        if panel is None:
            break

        visible = read_visible_rows(panel)
        added = 0
        for row in visible:
            # Skip completely empty rows
            if not any((v or "").strip() for v in row.values()):
                continue
            key = _row_dedup_key(row)
            if key not in all_rows:
                all_rows[key] = row
                added += 1

        print(f"    page {page + 1}: rows_visible={len(visible)} added={added} total={len(all_rows)}")

        # Checkpoint AFTER each scroll page, BEFORE the next scroll attempt.
        # If `scroll_grid_one_page` native-crashes us, this page's rows are
        # safely on disk via the orchestrator.
        if on_page is not None:
            try:
                on_page(page + 1, list(all_rows.values()))
            except Exception as cb_err:
                print(f"    on_page callback errored: {cb_err}", flush=True)

        if added == 0:
            pages_without_progress += 1
            if pages_without_progress >= 2:
                break
        else:
            pages_without_progress = 0

        if not scroll_grid_one_page(grid, panel):
            break
        time.sleep(0.4)

    return list(all_rows.values())


# ---------- Header-based column detection via OCR ----------

def _ocr_winrt(pil_image):
    """OCR via Windows.Media.Ocr (WinRT). Returns list of {text, left, right} in
    image coords, or None if WinRT isn't available."""
    try:
        import asyncio
        import io
        # Try winsdk first, then winrt
        modules = None
        try:
            import winsdk.windows.graphics.imaging as imaging
            import winsdk.windows.media.ocr as ocr_mod
            from winsdk.windows.storage.streams import (
                InMemoryRandomAccessStream, DataWriter,
            )
            modules = (imaging, ocr_mod, InMemoryRandomAccessStream, DataWriter)
        except Exception:
            try:
                import winrt.windows.graphics.imaging as imaging
                import winrt.windows.media.ocr as ocr_mod
                from winrt.windows.storage.streams import (
                    InMemoryRandomAccessStream, DataWriter,
                )
                modules = (imaging, ocr_mod, InMemoryRandomAccessStream, DataWriter)
            except Exception:
                return None
        imaging, ocr_mod, InMemoryRandomAccessStream, DataWriter = modules

        async def _do_ocr():
            buf = io.BytesIO()
            pil_image.save(buf, format="PNG")
            data = buf.getvalue()
            stream = InMemoryRandomAccessStream()
            writer = DataWriter(stream)
            writer.write_bytes(data)
            await writer.store_async()
            await writer.flush_async()
            stream.seek(0)
            decoder = await imaging.BitmapDecoder.create_async(stream)
            bitmap = await decoder.get_software_bitmap_async()
            engine = ocr_mod.OcrEngine.try_create_from_user_profile_languages()
            if engine is None:
                return None
            return await engine.recognize_async(bitmap)

        result = asyncio.run(_do_ocr())
        if result is None:
            return None
        words = []
        for line in result.lines:
            for word in line.words:
                b = word.bounding_rect
                words.append({
                    "text": word.text,
                    "left": int(b.x),
                    "right": int(b.x + b.width),
                })
        return words
    except Exception as e:
        print(f"  WinRT OCR errored: {e}", flush=True)
        return None


def _ocr_tesseract(pil_image):
    """OCR via Tesseract. Tries multiple configurations and returns whichever
    finds the most words. Always prints a diagnostic so we can see what happened."""
    try:
        try:
            import pytesseract  # type: ignore
        except ImportError:
            try:
                subprocess.run(
                    [sys.executable, "-m", "pip", "install", "--quiet", "pytesseract"],
                    capture_output=True, check=True, timeout=60,
                )
                import pytesseract  # type: ignore
            except Exception as install_err:
                print(f"  Tesseract OCR: could not install pytesseract ({install_err})", flush=True)
                return None
        # Auto-detect Tesseract path
        for p in (r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                  r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"):
            if Path(p).exists():
                pytesseract.pytesseract.tesseract_cmd = p
                break

        def _run(image, config, scale):
            data = pytesseract.image_to_data(
                image, output_type=pytesseract.Output.DICT, config=config,
            )
            ws = []
            for i in range(len(data["text"])):
                text = (data["text"][i] or "").strip()
                if not text:
                    continue
                ws.append({
                    "text": text,
                    "left": int(data["left"][i] / scale),
                    "right": int((data["left"][i] + data["width"][i]) / scale),
                })
            return ws

        attempts = [("default", pil_image, "", 1)]
        # Add a 2x-upscaled attempt as a fallback for very small text.
        try:
            from PIL import Image  # type: ignore
            up2 = pil_image.resize((pil_image.width * 2, pil_image.height * 2), Image.LANCZOS)
            attempts.append(("2x upscale", up2, "", 2))
        except Exception:
            pass

        best = []
        best_label = ""
        for label, img, config, scale in attempts:
            try:
                words = _run(img, config, scale)
                if len(words) > len(best):
                    best = words
                    best_label = label
            except Exception as ex:
                print(f"  Tesseract attempt {label!r} errored: {ex}", flush=True)
                continue

        if best:
            print(f"  Tesseract: {len(best)} words from {best_label!r} attempt", flush=True)
        else:
            print(f"  Tesseract: 0 words recognised "
                  f"(image size {pil_image.size}, attempts: "
                  f"{[a[0] for a in attempts]!r})", flush=True)
        return best
    except Exception as e:
        print(f"  Tesseract OCR errored: {e}", flush=True)
        return None


def _group_ocr_words(words, max_gap=30):
    """Cluster nearby words into header phrases (e.g. 'Cust' + '.' + 'Ref' -> one)."""
    if not words:
        return []
    words.sort(key=lambda w: w["left"])
    groups = []
    current = [words[0]]
    for w in words[1:]:
        prev = current[-1]
        if w["left"] - prev["right"] <= max_gap:
            current.append(w)
        else:
            groups.append({
                "text": " ".join(g["text"] for g in current),
                "left": current[0]["left"],
                "right": current[-1]["right"],
            })
            current = [w]
    groups.append({
        "text": " ".join(g["text"] for g in current),
        "left": current[0]["left"],
        "right": current[-1]["right"],
    })
    return groups


def detect_columns_via_ocr(grid, panel, dm=None):
    """OCR the grid header strip and map the recognised header text to display
    indices using the cell rectangles from the first data row.
    Returns {canonical_name: display_index} or {}."""
    if ImageGrab is None:
        print("  OCR skipped: PIL.ImageGrab not available", flush=True)
        return {}
    # Find the Header element of the grid
    header_elem = None
    try:
        for c in grid.children():
            try:
                if c.element_info.control_type == "Header":
                    header_elem = c
                    break
            except Exception:
                pass
    except Exception:
        return {}
    if header_elem is None:
        return {}
    hr = header_elem.element_info.rectangle
    bbox = (hr.left - 2, hr.top - 2, hr.right + 2, hr.bottom + 4)

    # Bring DM to the foreground BEFORE the screenshot — otherwise a console
    # window or other app sitting on top of the header strip will be OCR'd
    # instead of DM's headers. (Observed in colleague's run: "Press Ctrl+C..."
    # console text bled into the OCR results.)
    try:
        if dm is not None:
            UIAWrapper(dm.element_info).set_focus()
        else:
            UIAWrapper(grid.element_info).set_focus()
        time.sleep(0.3)
    except Exception:
        pass

    try:
        img = ImageGrab.grab(bbox=bbox, all_screens=True)
    except Exception as e:
        print(f"  OCR screenshot failed: {e}", flush=True)
        return {}

    # Try OCR backends in order, falling through if a backend runs but finds
    # nothing — some machines have winsdk installed but the WinRT OCR engine
    # returns empty results (e.g. DPI / language-pack issues), and Tesseract
    # handles those cases fine.
    words = _ocr_winrt(img)
    backend = "WinRT"
    if not words:  # None (not installed) OR empty (ran but found nothing)
        tess_words = _ocr_tesseract(img)
        if tess_words:
            words = tess_words
            backend = "Tesseract"
        elif words is None and tess_words is None:
            print("  OCR: no backend available — install winsdk (pip) or Tesseract for header detection", flush=True)
            return {}
    if not words:
        try:
            debug_path = HERE / "ocr_failed_header.png"
            img.save(debug_path)
            print(f"  OCR recognised 0 words on both backends. Image saved to {debug_path.name}", flush=True)
        except Exception:
            print(f"  OCR recognised 0 words on both backends.", flush=True)
        return {}

    # Translate image coords back to screen coords
    for w in words:
        w["left"] += bbox[0]
        w["right"] += bbox[0]

    groups = _group_ocr_words(words, max_gap=25)
    print(f"  OCR ({backend}): {len(groups)} header phrase(s) recognised", flush=True)

    # Get cell display indices + x ranges from first data row
    first_row = None
    for c in panel.children():
        try:
            if c.element_info.control_type == "DataItem":
                first_row = c
                break
        except Exception:
            pass
    if first_row is None:
        return {}
    cell_ranges = {}
    for cell in first_row.descendants():
        try:
            ci = cell.element_info
            m = INDEX_RE.search(ci.name or "")
            if m:
                idx = int(m.group(1))
                r = ci.rectangle
                cell_ranges[idx] = (r.left, r.right)
        except Exception:
            pass
    if not cell_ranges:
        return {}

    # Match each header group to a cell (by x overlap) and to a canonical name.
    name_to_idx = {}
    unmatched = []
    for g in groups:
        center = (g["left"] + g["right"]) / 2
        matched_idx = None
        for idx, (cl, cr) in cell_ranges.items():
            if cl <= center <= cr:
                matched_idx = idx
                break
        if matched_idx is None:
            # Fallback: nearest cell center
            best_d = None
            for idx, (cl, cr) in cell_ranges.items():
                d = abs(center - (cl + cr) / 2)
                if best_d is None or d < best_d:
                    best_d = d
                    matched_idx = idx
        if matched_idx is None:
            continue
        canonical = _canonical_header(g["text"])
        if canonical and canonical not in name_to_idx:
            # Reject the mapping if this column index is already claimed by
            # a DIFFERENT canonical header. Otherwise multiple OCR phrases all
            # snap to the same nearest column (e.g. Customer + Cust. Ref both
            # ending up at col 1), which silently corrupts every downstream
            # check.
            if matched_idx in name_to_idx.values():
                clash = next((n for n, i in name_to_idx.items() if i == matched_idx), "?")
                print(f"    OCR skipped {g['text']!r} -> col {matched_idx} (already claimed by {clash!r})", flush=True)
                continue
            name_to_idx[canonical] = matched_idx
            print(f"    OCR mapped {g['text']!r} -> {canonical} = {matched_idx}", flush=True)
        elif not canonical:
            unmatched.append((g["text"], matched_idx))
    if unmatched:
        # Visibility into what OCR saw but we couldn't map — useful for spotting
        # things like "Cust. Ref Req." where the rule deliberately rejects it.
        print(f"    OCR phrases not mapped: " +
              ", ".join(f"{t!r}@col{i}" for t, i in unmatched), flush=True)

    # Helpful warning: if OCR found 'Cust. Ref Req.' but no real 'Cust. Ref',
    # the actual reference column probably isn't visible in this DM view.
    if "Cust. Ref" not in name_to_idx:
        saw_req_only = any(
            "req" in t.lower() and "ref" in t.lower()
            for t, _i in unmatched
        )
        if saw_req_only:
            print("    !!! NOTE: 'Cust. Ref Req.' found but the actual 'Cust. Ref'", flush=True)
            print("        column isn't visible in this view. Add it via right-click", flush=True)
            print("        → Column Chooser → drag 'Cust. Ref' onto the grid.", flush=True)
            print("        Falling back to content-based detection (less reliable).", flush=True)

    return name_to_idx


# ---------- Header-based column detection via UIA TableItem pattern ----------

HEADER_PATTERNS = [
    # (canonical name, predicate on normalised header text)
    # Exact-match style — picky on purpose so e.g. "Cust. Ref Req." (a separate
    # required-flag column DM has) doesn't get matched as the actual "Cust. Ref".
    # Order matters slightly: Cust. Ref before Customer so the more specific
    # phrase wins if there's any overlap.
    ("Our Ref",       lambda t: t == "our ref"),
    ("Cust. Ref",     lambda t: t in ("cust ref", "cust. ref", "customer ref", "customer reference")),
    ("Customer",      lambda t: t == "customer"),
    ("Status",        lambda t: t == "status"),
    ("Del Date Time", lambda t: t in ("del date time", "del. date time", "del date", "delivery date", "delivery date time")),
    ("Col Date Time", lambda t: t in ("col date time", "col. date time", "col date", "collection date", "collection date time")),
]


def _normalize_header(s):
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = " ".join(s.split())
    return s


# Fuzzy-match candidates used after strict matching fails. The "flat" form
# (lowercased, no spaces or punctuation) is what we compare against — OCR
# often loses the space between words ("custret" instead of "cust ref") so
# matching with spaces stripped is more robust. Order doesn't matter; we
# pick the best ratio across all candidates.
_FUZZY_CANONICAL_CANDIDATES = (
    ("Our Ref", "ourref"),
    ("Customer", "customer"),
    ("Cust. Ref", "custref"),
    ("Status", "status"),
    ("Del Date Time", "deldatetime"),
    ("Col Date Time", "coldatetime"),
)


def _fuzzy_canonical_match(norm_text, threshold=0.7, length_ratio_max=1.5):
    """Sequence-similarity match against the canonical column names. Used as
    a fallback when strict regex/equality matching fails — OCR often mangles
    short headers (e.g. 'ourres' for 'Our Ref', 'custret' for 'Cust. Ref',
    'tate Time' for 'Del Date Time'). The fuzzy matcher catches those
    without false-matching real non-canonical headers (which score below the
    threshold).

    Length sanity check: if the OCR phrase's flat form is significantly
    longer than the canonical's flat form (default 1.5×), reject the match.
    OCR sometimes merges TWO adjacent headers into a single phrase (e.g.
    'ust ref | Del DateTime' → 'ustrefdeldatetime') and the longer match
    (Del Date Time) would otherwise dominate. The length filter catches that.

    `norm_text` should already be lowercased and have non-alphanumerics
    collapsed to spaces (which _normalize_header does). Returns the
    canonical name or None.
    """
    if not norm_text:
        return None
    try:
        from difflib import SequenceMatcher
    except Exception:
        return None
    flat = norm_text.replace(" ", "")
    if not flat:
        return None
    best_canonical = None
    best_ratio = 0.0
    for canonical, canonical_flat in _FUZZY_CANONICAL_CANDIDATES:
        # Reject when the OCR phrase is much longer than the canonical —
        # likely two headers smashed together.
        if len(flat) > length_ratio_max * len(canonical_flat):
            continue
        ratio = SequenceMatcher(None, flat, canonical_flat).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_canonical = canonical
    return best_canonical if best_ratio >= threshold else None


def _canonical_header(ocr_or_uia_text):
    """Map an OCR'd / UIA-derived header phrase to one of the canonical
    column names we care about. Strict pattern matching first (preserves
    the existing reliable cases), then fuzzy fallback for mangled OCR.
    """
    t = _normalize_header(ocr_or_uia_text)
    if not t:
        return None
    for canonical, pred in HEADER_PATTERNS:
        try:
            if pred(t):
                return canonical
        except Exception:
            pass
    # Strict matching failed. Try fuzzy.
    return _fuzzy_canonical_match(t)


def detect_columns_via_header_row(grid):
    """Walk the Telerik grid's header row directly via UIA.

    Telerik RadGridView exposes the column headers as a row of HeaderItem-typed
    children sitting under the grid (typically inside a GridViewHeaderRow
    container). Each header cell's UIA Name is the displayed header text, and
    its rectangle's left edge tells us the column's display order.

    Returns (canonical_map, raw_headers) where:
      - canonical_map: {canonical_name: display_index}   (Our Ref, Customer, ...)
      - raw_headers: {display_index: raw_text}           (every header we saw)

    This sidesteps OCR (incomplete on PCs where the header strip is cut off)
    and content heuristics (fooled by sparsely-populated columns). The display
    indices match clipboard tab positions because Telerik exports columns in
    display order.

    Empty dicts if we can't find headers (caller falls back to OCR + content).
    """
    canonical_map = {}
    raw_headers = {}
    candidates = []  # (left_px, raw_text)
    with gc_paused():
        try:
            # Walk every descendant of the grid, looking for header-cell elements.
            # The header strip is usually shallow (immediate child of the grid)
            # but we walk descendants() to be robust to layout variations.
            for el in grid.descendants():
                try:
                    ei = el.element_info
                    ct = (ei.control_type or "")
                    cn = (ei.class_name or "")
                    name = (ei.name or "").strip()
                except Exception:
                    continue
                # Skip anything that doesn't look like a header. Telerik's WPF
                # header cells are control_type "Header" or class_name
                # containing "HeaderCell" / "GridViewHeaderCell".
                # Accept the usual Telerik header cells AND — for robustness
                # across Telerik builds/themes where the header control type
                # differs — any element whose NAME is itself a known header
                # label. Data cells are named "Item: …Column Display Index: N",
                # so they can't be mistaken for a header by this test.
                label_is_header = (
                    _canonical_from_header_v2 is not None
                    and bool(name)
                    and _canonical_from_header_v2(name) is not None
                )
                looks_like_header = (
                    ct in ("Header", "HeaderItem")
                    or "HeaderCell" in cn
                    or "GridViewHeaderCell" in cn
                    or label_is_header
                )
                if not looks_like_header:
                    continue
                if not name:
                    continue
                # Skip the "select all" / row-header corner cell that has no
                # meaningful text or has duplicate names like "checkbox".
                if name.lower() in ("checkbox", "select", "select all"):
                    continue
                try:
                    rect = ei.rectangle
                    left = int(getattr(rect, "left", 0))
                except Exception:
                    continue
                candidates.append((left, name))
        except Exception as e:
            print(f"    header-row walk errored: {e}", flush=True)
            return {}, {}

    if not candidates:
        return {}, {}

    # Sort by left edge → display order. De-dupe by name (same header sometimes
    # appears in both a sort-arrow inner element and the outer cell).
    candidates.sort(key=lambda x: x[0])
    seen_names = set()
    ordered = []
    for left, name in candidates:
        # Strip any trailing sort-direction suffix Telerik appends.
        clean = re.sub(r"\s+(ascending|descending)$", "", name, flags=re.IGNORECASE).strip()
        if clean in seen_names:
            continue
        seen_names.add(clean)
        ordered.append(clean)

    for idx, header_text in enumerate(ordered):
        raw_headers[idx] = header_text
        canonical = _canonical_header(header_text)
        if canonical and canonical not in canonical_map:
            canonical_map[canonical] = idx

    if canonical_map:
        print(
            f"    UIA header row mapped {len(canonical_map)} canonical / "
            f"{len(ordered)} total headers: "
            + ", ".join(f"{k}={v}" for k, v in sorted(canonical_map.items(), key=lambda x: x[1])),
            flush=True,
        )
        # Show all headers in order — useful for debugging which raw text
        # appeared at each index.
        print(
            "      ordered headers: "
            + ", ".join(f"[{i}]={t!r}" for i, t in enumerate(ordered)),
            flush=True,
        )
    return canonical_map, raw_headers


def detect_columns_via_uia(panel):
    """For each cell in the first data row, try UIA TableItemPattern to get the
    column header element, then read its name. Returns {canonical_name: display_index}
    or {} if nothing found."""
    from ctypes import POINTER
    try:
        from comtypes import cast  # type: ignore
        from comtypes.gen.UIAutomationClient import (  # type: ignore
            IUIAutomationTableItemPattern,
        )
    except Exception:
        return {}
    UIA_TableItemPatternId = 10013

    # Find the first DataItem (row)
    first_row = None
    for c in panel.children():
        try:
            if c.element_info.control_type == "DataItem":
                first_row = c
                break
        except Exception:
            pass
    if first_row is None:
        return {}

    mapping = {}  # canonical -> display_index
    raw_seen = {}  # canonical -> raw header text (for diagnostics)
    for cell in first_row.descendants():
        try:
            ci = cell.element_info
            m = INDEX_RE.search(ci.name or "")
            if not m:
                continue
            idx = int(m.group(1))
        except Exception:
            continue
        try:
            iuia_cell = cell.element_info.element
            ptr = iuia_cell.GetCurrentPattern(UIA_TableItemPatternId)
            if not ptr:
                continue
            ti = cast(ptr, POINTER(IUIAutomationTableItemPattern))
            if not ti:
                continue
            headers_arr = ti.GetCurrentColumnHeaderItems()
            if not headers_arr:
                continue
            n = headers_arr.Length
            if n <= 0:
                continue
            header_elem = headers_arr.GetElement(0)
            header_name = ""
            try:
                header_name = header_elem.CurrentName or ""
            except Exception:
                pass
            if not header_name:
                continue
            canonical = _canonical_header(header_name)
            if canonical and canonical not in mapping:
                mapping[canonical] = idx
                raw_seen[canonical] = header_name
        except Exception:
            continue
    if mapping:
        print(f"    UIA-detected columns ({len(mapping)}): "
              + ", ".join(f"{k}({raw_seen.get(k, '?')!r})={v}" for k, v in mapping.items()))
    return mapping


# ---------- Column detection ----------

def detect_columns(rows, uia_seed=None):
    """Return (detected, diagnostics) where detected maps canonical name -> display index.

    If `uia_seed` is supplied (from detect_columns_via_uia), those mappings take
    priority. Anything not provided by UIA is filled in from content-based detection.
    """
    by_col = defaultdict(list)
    for row in rows:
        for idx, val in row.items():
            by_col[idx].append(val)

    detected = dict(uia_seed) if uia_seed else {}
    diagnostics = {"uia_seed": dict(uia_seed) if uia_seed else {}}

    # Our Ref — only if not already provided by UIA
    if "Our Ref" not in detected:
        s = {idx: sum(1 for v in vals if BT_RE.match(v)) for idx, vals in by_col.items()}
        if s and max(s.values()) > 0:
            detected["Our Ref"] = max(s, key=s.get)
    # Status
    if "Status" not in detected:
        s = {idx: sum(1 for v in vals if v in STATUS_VALUES or ON_CX_RE.match(v))
             for idx, vals in by_col.items()}
        if s and max(s.values()) > 0:
            detected["Status"] = max(s, key=s.get)
    # Dates
    ds = {idx: sum(1 for v in vals if DATE_RE.match(v)) for idx, vals in by_col.items()}
    threshold = max(2, int(len(rows) * 0.5))
    by_count = sorted(
        [(idx, count) for idx, count in ds.items() if count >= threshold],
        key=lambda kv: (-kv[1], kv[0]),
    )
    top_two_idx = sorted(idx for idx, _ in by_count[:2])
    if "Col Date Time" not in detected and "Del Date Time" not in detected:
        if len(top_two_idx) >= 2:
            detected["Col Date Time"] = top_two_idx[0]
            detected["Del Date Time"] = top_two_idx[1]
        elif len(top_two_idx) == 1:
            detected["Del Date Time"] = top_two_idx[0]
    elif "Del Date Time" not in detected and len(top_two_idx) >= 2:
        # Col already set, pick the other one
        col_idx = detected.get("Col Date Time")
        candidates = [i for i in top_two_idx if i != col_idx]
        if candidates:
            detected["Del Date Time"] = candidates[0]
    elif "Col Date Time" not in detected and len(top_two_idx) >= 2:
        del_idx = detected.get("Del Date Time")
        candidates = [i for i in top_two_idx if i != del_idx]
        if candidates:
            detected["Col Date Time"] = candidates[0]
    diagnostics["date_match_counts"] = dict(sorted(ds.items()))
    used = set(detected.values())

    # Customer — distinguish the real customer column from columns containing
    # driver/contact names. The actual customer column tends to have:
    #   * company-name suffixes (LTD, LIMITED, LLC, LLP, INC, PLC, GMBH...)
    #   * ALL-CAPS company names
    #   * the SAME customer repeated across many rows (one customer, many jobs)
    # Whereas a driver column has individual person names — usually unique
    # per row, mixed-case, no company suffix.
    if "Customer" not in detected:
        # Tokens that strongly indicate a company name. Case-insensitive whole
        # word / suffix match.
        company_suffix_re = re.compile(
            r"\b(LTD|LIMITED|LLC|LLP|INC|PLC|GMBH|CO|COMPANY|GROUP|"
            r"SERVICES|LOGISTICS|TRANSPORT|COURIER|FREIGHT|HAULAGE|"
            r"INTERNATIONAL|UK|EUROPE|SUPPLY|SOLUTIONS|HOLDINGS)\b\.?",
            re.IGNORECASE,
        )
        total_rows = len(rows) if rows else 0
        ts = {}
        for idx, vals in by_col.items():
            if idx in used:
                continue
            non_empty = [v.strip() for v in vals if v and v != "---"]
            if not non_empty:
                continue
            n = len(non_empty)
            is_other = sum(1 for v in non_empty if DATE_RE.match(v) or DEC_RE.match(v) or POSTCODE_RE.search(v))
            if is_other > n * 0.5:
                continue
            distinct = len(set(non_empty))
            uniqueness = distinct / n
            # REPETITION is a positive signal (one customer = many jobs). The
            # driver column is typically near-unique. We score repetition as
            # 1 - uniqueness so 0.0 = every row distinct, 1.0 = all the same.
            repetition = 1.0 - uniqueness
            # Fraction of rows containing a company suffix word.
            company_hits = sum(1 for v in non_empty if company_suffix_re.search(v))
            company_ratio = company_hits / n
            # Fraction of ALL-CAPS values (companies often shouted, drivers not).
            allcaps = sum(1 for v in non_empty if v.upper() == v and any(c.isalpha() for c in v))
            allcaps_ratio = allcaps / n
            avg_len = sum(len(v) for v in non_empty) / n
            # COVERAGE = how often this column is populated at all. A driver
            # column populated only on allocated rows might score well on
            # ratios over its few non-empty samples while being mostly empty.
            # Multiply by coverage so we prefer columns populated across the
            # whole view.
            coverage = n / total_rows if total_rows else 0
            # Combined score — company-suffix matches dominate, then ALL-CAPS,
            # then repetition, with length as a tie-breaker. Coverage gates
            # everything: a sparsely-populated column can't win regardless of
            # how strongly its few values look like company names.
            raw_score = (
                company_ratio * 2.0
                + allcaps_ratio * 0.6
                + repetition * 0.4
                + min(avg_len, 30) / 30 * 0.1
            )
            ts[idx] = raw_score * coverage
        if ts:
            detected["Customer"] = max(ts, key=ts.get)
            used.add(detected["Customer"])
    # Cust. Ref — pick by VALUE UNIQUENESS so it isn't fooled by columns like
    # "A/C Type" (where every row is "Invoice") or CX Load ID (pure numeric).
    if "Cust. Ref" in detected:
        return detected, diagnostics
    cs = {}
    for idx, vals in by_col.items():
        if idx in used:
            continue
        non_empty = [v.strip() for v in vals if v and v != "---"]
        if not non_empty:
            continue
        n = len(non_empty)
        # Reject mostly-date or mostly-postcode columns
        date_or_pc = sum(1 for v in non_empty if DATE_RE.match(v) or POSTCODE_RE.search(v))
        if date_or_pc > n * 0.3:
            continue
        distinct = len(set(non_empty))
        uniqueness = distinct / n
        pure_numeric = all(DEC_RE.match(v) for v in non_empty)
        kw_hits = sum(
            1 for v in non_empty
            if any(kw in v.upper() for kw in ["CHAS", "TBC", "NEED", "QUOTE", "CHECK", "REPORT"])
        )
        # Primary signal: how varied the column is. Bonus for placeholder keywords.
        score = uniqueness + (kw_hits / n) * 2
        # Heavy penalty for pure-numeric high-uniqueness columns (typically CX Load ID).
        if pure_numeric and uniqueness > 0.9:
            score *= 0.2
        cs[idx] = score
    if cs and max(cs.values()) > 0:
        ranked = sorted(cs.items(), key=lambda kv: (-kv[1], kv[0]))
        detected["Cust. Ref"] = ranked[0][0]
    return detected, diagnostics


# ---------- Criteria ----------

def parse_dm_date(s):
    """Parse either the display-formatted 'DD-MM-YY HH:MM' or the clipboard-formatted
    'DD-MM-YYYY HH:MM:SS' style date that Delivery Master uses."""
    s = (s or "").strip()
    # 4-digit-year form first (clipboard): "18-05-2026 06:00:15"
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})(?:\s+(\d{2}):(\d{2})(?::(\d{2}))?)?", s)
    if m:
        d, mo, y, h, mi, sec = m.groups()
        try:
            return datetime(int(y), int(mo), int(d), int(h or 0), int(mi or 0), int(sec or 0))
        except ValueError:
            return None
    # 2-digit-year form (display): "13-05-26 22:26"
    m = re.match(r"^(\d{2})-(\d{2})-(\d{2})(?:\s+(\d{2}):(\d{2}))?", s)
    if m:
        d, mo, y, h, mi = m.groups()
        try:
            return datetime(2000 + int(y), int(mo), int(d), int(h or 0), int(mi or 0))
        except ValueError:
            return None
    return None


# ---------- Clipboard ↔ display-index rosetta stone ----------
#
# The clipboard and the UIA "Column Display Index" property use DIFFERENT
# coordinate systems for the same grid:
#
#   - UIA Column Display Index follows the user's VISIBLE column order. If
#     the user drags Customer to the left of Cust. Ref in the grid, the
#     display index for Customer becomes lower.
#   - Telerik's clipboard exports columns in the underlying DATA-MODEL order
#     (the order columns were defined in DM's source). The clipboard order
#     does NOT change when the user drags columns around the grid.
#
# OCR-based header detection lives in display-index space (it's reading the
# rendered headers on screen, in visible order). Categorisation indexes rows
# in the same space (cols = {canonical_header: display_idx}). So we MUST
# translate every clipboard row from tab-position space to display-index
# space before the rest of the pipeline can use it correctly. Otherwise
# Customer ends up showing the Cust. Ref text, Cust. Ref ends up showing a
# cost value, and so on.
#
# The "rosetta stone" is a single matching row: we walk UIA for the first
# visible row (already done in Phase 1) to get {display_idx -> value}, find
# the same row in the clipboard (via Our Ref), and for each value match
# work out which tab position contains it. That gives us {display_idx ->
# tab_pos} and we can translate every other clipboard row using it.


def _norm_for_match(s):
    """Normalise a cell value for content-matching across UIA and clipboard.
    Both sides may add/strip whitespace, BT prefixes, and slightly different
    date formats."""
    return (s or "").strip()


def _value_matches(uia_val, clip_val):
    """True if these two values are 'the same' across UIA and clipboard
    representations. Handles BT-prefix and date-format differences."""
    a = _norm_for_match(uia_val)
    b = _norm_for_match(clip_val)
    if not a or not b:
        return False
    if a == b:
        return True
    # BT prefix: UIA shows "BT62946", clipboard shows "62946"
    if a.startswith("BT") and a[2:] == b:
        return True
    if b.startswith("BT") and b[2:] == a:
        return True
    # Dates: UIA gives "18-05-26 08:30", clipboard gives "18-05-2026 08:30:54"
    if DATE_RE.match(a) and DATE_RE.match(b):
        da = parse_dm_date(a)
        db = parse_dm_date(b)
        if da and db:
            # Same minute is good enough — clipboard may have seconds, UIA
            # may have trimmed them.
            return da.replace(second=0) == db.replace(second=0)
    return False


def find_clipboard_anchor(uia_row, clip_rows):
    """Find a clipboard row that corresponds to the supplied UIA row.

    Matches by Our-Ref-shaped values first (most unique), then by any other
    value as a fallback. Returns the matching clipboard row dict, or None.
    """
    if not uia_row or not clip_rows:
        return None
    # Pull all Our-Ref-shaped values from the UIA row
    ref_candidates = [
        _norm_for_match(v) for v in uia_row.values()
        if _norm_for_match(v) and BT_RE.match(_norm_for_match(v))
    ]
    if ref_candidates:
        for clip_row in clip_rows:
            for v in clip_row.values():
                vv = _norm_for_match(v)
                if not vv:
                    continue
                for ref in ref_candidates:
                    if _value_matches(ref, vv):
                        return clip_row
    # Fallback: count value-match across rows, pick the best
    best = (0, None)
    for clip_row in clip_rows:
        matches = 0
        for u in uia_row.values():
            for c in clip_row.values():
                if _value_matches(u, c):
                    matches += 1
                    break
        if matches > best[0]:
            best = (matches, clip_row)
    return best[1]


def build_display_to_tab(uia_row, clip_row):
    """Given one matching pair, build {display_idx -> tab_pos} by walking
    the UIA row in display order and finding the clipboard tab position
    that holds the same value (handling BT/date normalisation).

    A tab position can only be claimed by ONE display index. Display
    indices whose UIA value is empty or whose match would collide are
    skipped — caller can still use whatever mapping we did establish.
    """
    mapping = {}
    used_tabs = set()
    # Sort by display index so the dump output reads top-to-bottom.
    for display_idx in sorted(uia_row.keys()):
        uia_val = _norm_for_match(uia_row[display_idx])
        if not uia_val:
            continue
        # Walk clipboard tabs in order and take the FIRST unclaimed match.
        for tab_pos in sorted(clip_row.keys()):
            if tab_pos in used_tabs:
                continue
            if _value_matches(uia_val, clip_row.get(tab_pos)):
                mapping[display_idx] = tab_pos
                used_tabs.add(tab_pos)
                break
    return mapping


def translate_clip_rows_to_display(clip_rows, display_to_tab):
    """Re-key every clipboard row from {tab_pos -> value} into
    {display_idx -> value} using the mapping. Columns with no mapping
    are dropped (we wouldn't be able to interpret them anyway)."""
    out = []
    for clip_row in clip_rows:
        translated = {}
        for display_idx, tab_pos in display_to_tab.items():
            v = clip_row.get(tab_pos)
            if v is None:
                continue
            translated[display_idx] = v
        out.append(translated)
    return out


def categorize_rows_three_way(rows, cols, view_name, rules=None):
    """
    Split rows into (flagged, accepted, not_eligible).

    Buckets, in the user-facing language Owen uses for data validation:
      • flagged ('Not Accepted')  — past Del Date AND has a reason
        (not delivered, Cust. Ref missing/TBC/chased, custom rule).
      • accepted ('Accepted')      — past Del Date AND clean (POD +
        valid Cust. Ref + no rule violations).
      • not_eligible ('Not Eligible') — Del Date is in the future
        OR can't be parsed. We still RECORD these so the captured
        count can be checked against the on-screen total; we just
        don't ever flag them as needing follow-up.

    Returns three lists; every input row lands in exactly one of them,
    so len(flagged) + len(accepted) + len(not_eligible) == len(rows).

    Note: the older two-bucket categorize_rows() (below) is preserved
    as a shim for callers that don't care about the not-eligible
    breakdown.
    """
    rules = rules or []
    today = date.today()
    flagged = []
    accepted = []
    not_eligible = []
    has_cust_ref_col = "Cust. Ref" in cols
    has_del_date = "Del Date Time" in cols
    has_status = "Status" in cols
    for row in rows:
        our_ref = (row.get(cols.get("Our Ref", -1)) or "").strip()
        if not our_ref:
            # Our Ref column wasn't detected (e.g., hidden on this view) —
            # fall back to scanning the row for a BT-prefixed value.
            our_ref = _find_ref_in_row(row)
        # Clipboard data has the raw numeric ref (e.g. "62882"); the displayed
        # form is "BT62882". Add the prefix back so the email reads naturally.
        if our_ref and our_ref.isdigit() and len(our_ref) <= 6:
            our_ref = "BT" + our_ref
        status = (row.get(cols.get("Status", -1)) or "").strip()
        cust_ref = (row.get(cols.get("Cust. Ref", -1)) or "").strip() if has_cust_ref_col else ""
        del_date_str = (row.get(cols.get("Del Date Time", -1)) or "").strip()
        customer = (row.get(cols.get("Customer", -1)) or "").strip()

        del_dt = parse_dm_date(del_date_str) if has_del_date else None
        is_past = bool(del_dt and del_dt.date() < today)

        reasons = []

        # All flag criteria — standard and custom — only apply when the
        # delivery date is in the past. Future-dated rows are out of scope
        # for the daily chase even if they would otherwise trip a rule
        # (Owen's call: "we are only looking for delivery deadline dates
        # that are in the past").
        if is_past:
            if not has_status or status not in DELIVERED_STATUSES:
                reasons.append(f"Not delivered (status: {status or 'unknown'})")
            if has_cust_ref_col:
                if not cust_ref:
                    reasons.append("No Cust. Ref")
                elif cust_ref.upper() == "TBC":
                    reasons.append("Cust. Ref = TBC")
                elif "chased" in cust_ref.lower():
                    reasons.append(f"Cust. Ref contains 'chased' ({cust_ref})")
            reasons.extend(evaluate_custom_rules(customer, cust_ref, rules))

        # ---- Whitelist (accept-rule) override ----
        # If any "accept" rule matches this row, wipe all reasons so the row
        # moves from flagged to accepted. Lets the user add positive exceptions
        # without rewriting the existing flag rules.
        if reasons:
            cust_lower = (customer or "").lower()
            for rule in rules:
                if rule.get("rule_type") != "accept":
                    continue
                cc = (rule.get("customer_contains") or "").lower()
                if cc and cc not in cust_lower:
                    continue
                ce = (rule.get("customer_excludes") or "").lower()
                if ce and ce in cust_lower:
                    continue
                if cust_ref in rule.get("values", []):
                    reasons = []
                    break

        record = {
            "view": view_name,
            "our_ref": our_ref,
            "customer": customer,
            "status": status,
            "del_date": del_date_str,
            "cust_ref": cust_ref,
        }
        if reasons:
            record["reasons"] = "; ".join(reasons)
            flagged.append(record)
        elif is_past:
            accepted.append(record)
        else:
            # Future-dated, or no parseable date — out of scope for
            # review. Record the reason so the data-validation pane
            # can show "Not Eligible: delivery date in future".
            if del_dt is None:
                record["reasons"] = (
                    f"Not eligible (no parseable date: {del_date_str!r})")
            else:
                record["reasons"] = (
                    f"Not eligible (delivery {del_date_str} is in the future)")
            not_eligible.append(record)
    return flagged, accepted, not_eligible


def categorize_rows(rows, cols, view_name, rules=None):
    """Two-bucket compatibility shim. Returns (flagged, accepted) and
    discards the not_eligible bucket - kept so existing callers that
    only care about chase-worthy rows keep working unchanged."""
    flagged, accepted, _not_eligible = categorize_rows_three_way(
        rows, cols, view_name, rules)
    return flagged, accepted


# ---------- View switching ----------

def switch_view(dm, view_name):
    """Click the filter button for a view. Returns True if it switched."""
    # If already showing this view, skip
    current_title = (dm.window_text() or "")
    if f"View {view_name}" in current_title or current_title.endswith(view_name):
        # Already there
        return True
    if view_name == "In Progress" and "In Progress" in current_title:
        return True

    btn = find_button(dm, view_name)
    if btn is None:
        print(f"    Couldn't find button {view_name!r}")
        return False
    try:
        wrapper = UIAWrapper(btn.element_info)
        try:
            wrapper.invoke()
        except Exception:
            wrapper.click_input()
        # Wait for the grid to refresh
        time.sleep(2)
        return True
    except Exception as e:
        print(f"    Switch error: {e}")
        return False


# ---------- Custom rules (dm_rules.xlsx) ----------

def ensure_rules_workbook(path):
    """Create a starter rules workbook if one doesn't exist yet."""
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Rules"
    headers = [
        "Customer Contains",
        "Rule Type",
        "Values",
        "Reason if violated",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    # Starter examples — illustrate each rule type
    ws.append([
        "United Utilities Scientific Services",
        "must equal",
        "3500279751",
        "Incorrect PO or wrong account",
        "Customer-specific: this customer must use this exact PO.",
    ])
    ws.append([
        "Brenntag UK Limited",
        "must start with",
        "BR",
        "Brenntag ref should start with BR",
        "Customer-specific prefix check.",
    ])
    ws.append([
        "",                       # empty = applies to all customers
        "must not contain",
        "needs PR",
        "Cust. Ref says 'needs PR'",
        "Catches the 'needs PR' placeholder across all customers.",
    ])
    ws.append([
        "",
        "must not contain",
        "chase, chsed, chesd, chasd, chaed",
        "Cust. Ref indicates chasing (incl. misspellings)",
        "Covers 'chase'/'chased' and typos missing a letter. 'hased'/'cased' omitted because they appear in 'purchased' etc.",
    ])
    ws.append([
        "",
        "must not equal",
        ".",
        "Cust. Ref is just '.'",
        "Catches lone-dot placeholders.",
    ])
    ws.append([
        "",
        "must not contain",
        "name, reg required",
        "Cust. Ref says 'name' or 'reg required'",
        "Catches missing-info placeholders.",
    ])
    ws.append([
        "",
        "must not contain",
        "TEST, PROVISIONAL, DRAFT",
        "Suspicious Cust. Ref",
        "Catches test bookings / placeholder refs across all customers.",
    ])
    widths = [40, 18, 32, 36, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Dropdown for Rule Type
    rule_type_dv = DataValidation(
        type="list",
        formula1=(
            '"must equal,must not equal,must contain,must not contain,'
            'must start with,must not start with"'
        ),
        allow_blank=True,
    )
    rule_type_dv.add("B2:B500")
    ws.add_data_validation(rule_type_dv)

    info = wb.create_sheet("Instructions", 0)
    info.append(["Delivery Master — Customer Rules"])
    info["A1"].font = Font(bold=True, size=14)
    info.append([])
    info.append(["Columns:"])
    info.append(["  - Customer Contains: text the Customer field must contain (case-insensitive)."])
    info.append(["                       LEAVE BLANK to apply the rule to every customer."])
    info.append(["  - Rule Type: one of"])
    info.append(["       must equal           -> Cust. Ref must be one of the listed Values"])
    info.append(["       must not equal       -> Cust. Ref must NOT be any of the listed Values"])
    info.append(["       must contain         -> Cust. Ref must contain at least one of the Values (substring)"])
    info.append(["       must not contain     -> Cust. Ref must NOT contain any of the Values (substring)"])
    info.append(["       must start with      -> Cust. Ref must start with one of the Values (prefix)"])
    info.append(["       must not start with  -> Cust. Ref must NOT start with any of the Values (prefix)"])
    info.append(["  - Values: comma-separated. Examples: '3500279751' or '3500279751, 3500279752' or 'TEST, DRAFT'"])
    info.append(["  - Reason if violated: the text that appears in the flagged email."])
    info.append([])
    info.append(["Rules fire regardless of delivery date, so wrong POs are caught even before delivery."])
    info.append([])
    info.append(["Examples:"])
    info.append(["  Customer X must use PO 12345:           Customer Contains=X     must equal           12345"])
    info.append(["  Customer X's PO must start with BR:     Customer Contains=X     must start with      BR"])
    info.append(["  No one should use TEST or DRAFT:        Customer Contains=blank must not contain     TEST, DRAFT"])
    info.append(["  Bad blacklist (any customer):           Customer Contains=blank must not equal       XXX, ZZZ"])
    info.append([])
    info.append(["To add a rule: append a new row on the 'Customer Rules' tab and save the file."])
    info.column_dimensions["A"].width = 110
    wb.save(path)
    print(f"Created starter rules workbook: {path}")


VALID_RULE_TYPES = {
    "must equal",
    "must not equal",
    "must contain",
    "must not contain",
    "must start with",
    "must not start with",
    # New format-shape rule types (added 2026-05 from consignment-log
    # analysis - lets us auto-generate rules for customers whose refs
    # are typically all-digit or all-letter):
    "must be numeric",      # cust_ref must contain only digits (allow / -)
    "must not be alpha",    # cust_ref must contain at least one digit
    "must not be name",     # cust_ref must not look like a person's name
    "accept",  # positive whitelist - wipes all reasons when matched
}


# Pattern: 1-3 Title-Case words, no digits, no punctuation. Same heuristic
# the consignment-log analysis used for the 'name_used_as_ref' detector.
import re as _re_for_name
_NAME_REF_RE = _re_for_name.compile(
    r"^[A-Z][a-z]+(?: [A-Z][a-z]+){0,2}$")
_NAME_REF_IGNORE = {
    "Tbc", "Need", "Needs", "Chase", "Chased", "Required",
    "Pending", "None", "Quote", "Quoted", "Draft", "Test", "Pod",
}


def _ref_looks_like_name(s: str) -> bool:
    s = (s or "").strip()
    if not s or any(c.isdigit() for c in s):
        return False
    if not _NAME_REF_RE.match(s):
        return False
    return not any(w in _NAME_REF_IGNORE for w in s.split())


def load_rules(path):
    """
    Return list of {customer_contains, rule_type, values, reason}.

    Schema (current): Customer Contains | Rule Type | Values | Reason | (notes)
    Schema (legacy):  Customer Contains | Values | Reason | (notes)   -> rule_type defaults to 'must equal'
    """
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  Could not read rules workbook: {e}")
        return []
    if "Customer Rules" not in wb.sheetnames:
        return []
    ws = wb["Customer Rules"]

    # Detect schema by looking at header row. We now read by HEADER NAME
    # (not fixed index) so optional columns like "Customer Excludes" can
    # be inserted anywhere without breaking older rule rows.
    header_row = [(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_lower = [str(h).strip().lower() for h in header_row]
    new_schema = "rule type" in headers_lower

    # Map header name (lower) -> column index
    hdr_idx = {h: i for i, h in enumerate(headers_lower)}

    def get(row, header_lower, default=""):
        idx = hdr_idx.get(header_lower)
        if idx is None or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip()

    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        if new_schema:
            cust_contains = get(row, "customer contains")
            cust_excludes = get(row, "customer excludes")
            rule_type = get(row, "rule type").lower()
            vals_raw = get(row, "values")
            reason = get(row, "reason if violated") or get(row, "reason")
        else:
            # legacy 4-column: Customer Contains | Values | Reason | Notes
            cust_contains = (str(row[0]) if row[0] is not None else "").strip()
            cust_excludes = ""
            rule_type = "must equal"
            vals_raw = (str(row[1]) if len(row) > 1 and row[1] is not None else "").strip()
            reason = (str(row[2]) if len(row) > 2 and row[2] is not None else "").strip()
        if not rule_type:
            rule_type = "must equal"
        if rule_type not in VALID_RULE_TYPES:
            print(f"  Skipping rule with unknown Rule Type {rule_type!r}")
            continue
        # Flag-style rule types don't take a values list - they apply
        # whenever a customer matches. For list-style rule types we still
        # require at least one value + a reason.
        FLAG_RULE_TYPES = {"must be numeric", "must not be alpha",
                           "must not be name"}
        is_flag = rule_type in FLAG_RULE_TYPES
        values = [v.strip() for v in vals_raw.split(",") if v.strip()]
        if not is_flag and (not values or not reason):
            continue
        if not reason:
            # Auto-generate a reason for flag rules so downstream display
            # always has something to print.
            reason = f"Cust. Ref violates '{rule_type}' for this customer"
        if is_flag and not values:
            values = ["*"]  # sentinel - we ignore it in the evaluator
        rules.append({
            "customer_contains": cust_contains,  # empty -> all customers
            "customer_excludes": cust_excludes,  # non-empty -> exclude matching
            "rule_type": rule_type,
            "values": values,
            "reason": reason,
        })
    return rules


def evaluate_custom_rules(customer, cust_ref, rules):
    """Return a list of reason strings for any custom rules this row violates."""
    reasons = []
    cust_lower = (customer or "").lower()
    ref_value = (cust_ref or "").strip()
    ref_lower = ref_value.lower()
    for rule in rules:
        # Customer match: empty 'customer_contains' = applies to all customers
        cc = rule["customer_contains"].lower()
        if cc and cc not in cust_lower:
            continue
        # Customer exclude: skip this rule if the customer matches this text.
        # Empty 'customer_excludes' = no exclusion.
        ce = rule.get("customer_excludes", "").lower()
        if ce and ce in cust_lower:
            continue

        rt = rule["rule_type"]
        values = rule["values"]
        if rt == "must equal":
            if ref_value not in values:
                expected = " or ".join(values)
                reasons.append(f"{rule['reason']} (expected {expected}, got {ref_value or 'empty'!r})")
        elif rt == "must not equal":
            if ref_value in values:
                reasons.append(f"{rule['reason']} (Cust. Ref is {ref_value!r})")
        elif rt == "must contain":
            if not any(v and v.lower() in ref_lower for v in values):
                wanted = " or ".join(values)
                reasons.append(f"{rule['reason']} (Cust. Ref must contain {wanted}, got {ref_value or 'empty'!r})")
        elif rt == "must not contain":
            for needle in values:
                if needle and needle.lower() in ref_lower:
                    reasons.append(f"{rule['reason']} (Cust. Ref contains {needle!r})")
                    break
        elif rt == "must start with":
            if not any(v and ref_lower.startswith(v.lower()) for v in values):
                wanted = " or ".join(values)
                reasons.append(f"{rule['reason']} (Cust. Ref must start with {wanted}, got {ref_value or 'empty'!r})")
        elif rt == "must not start with":
            for prefix in values:
                if prefix and ref_lower.startswith(prefix.lower()):
                    reasons.append(f"{rule['reason']} (Cust. Ref starts with {prefix!r})")
                    break
        elif rt == "must be numeric":
            # Reject if reference contains any letter. We tolerate digits,
            # spaces, hyphens, slashes, dots - common in formatted PO
            # numbers like '12345-67' or '345.789'. The 'values' column
            # is ignored for this rule type (a flag rule, not a list rule).
            if ref_value and any(c.isalpha() for c in ref_value):
                reasons.append(
                    f"{rule['reason']} (Cust. Ref contains letters: "
                    f"{ref_value!r})")
        elif rt == "must not be alpha":
            # Reject if reference is alpha-only (no digits at all).
            # Catches names-as-PO for numeric-only customers.
            if ref_value and not any(c.isdigit() for c in ref_value) \
                    and any(c.isalpha() for c in ref_value):
                reasons.append(
                    f"{rule['reason']} (Cust. Ref has no digits: "
                    f"{ref_value!r})")
        elif rt == "must not be name":
            if _ref_looks_like_name(ref_value):
                reasons.append(
                    f"{rule['reason']} (Cust. Ref looks like a person's "
                    f"name: {ref_value!r})")
    return reasons


# ---------- Excel review workbook ----------

def write_review_xlsx(all_flagged, all_accepted, out_path, diagnostics=None):
    """Write a multi-tab review workbook.

    Tabs:
      - Instructions: user-facing notes
      - Flagged: rows that need emailing (user edits Include? column)
      - Accepted: clean rows to optionally promote
      - Diagnostics: per-view column mapping + first 10 raw rows so we can
        debug column mis-detection (e.g. "Customer" picking up the driver
        column) without needing the user to send back raw JSON.
    """
    wb = Workbook()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # ---- Flagged sheet ----
    ws = wb.active
    ws.title = "Flagged"
    flagged_headers = [
        "Include?", "View", "Our Ref", "Customer", "Status", "Del Date", "Cust. Ref", "Reasons"
    ]
    ws.append(flagged_headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in all_flagged:
        ws.append([
            "Yes", r["view"], r["our_ref"], r["customer"], r["status"],
            r["del_date"], r["cust_ref"], r.get("reasons", ""),
        ])

    # Column widths only — DataValidation, freeze_panes, and auto_filter
    # have been observed to native-crash openpyxl on Python 3.14, so they're
    # disabled. The Include? column still works as plain text ("Yes"/"No").
    widths_flagged = [10, 14, 12, 32, 14, 18, 26, 60]
    for i, w in enumerate(widths_flagged, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Accepted sheet ----
    ws2 = wb.create_sheet("Accepted")
    accepted_headers = [
        "Promote?", "View", "Our Ref", "Customer", "Status", "Del Date", "Cust. Ref", "Promote Reason"
    ]
    ws2.append(accepted_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in all_accepted:
        ws2.append([
            "No", r["view"], r["our_ref"], r["customer"], r["status"],
            r["del_date"], r["cust_ref"], "",
        ])

    widths_accepted = [10, 14, 12, 32, 14, 18, 26, 40]
    for i, w in enumerate(widths_accepted, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Diagnostics sheet ----
    # Per-view column mapping plus the first 10 raw rows from each view.
    # Use this when something looks wrong (missing flags, wrong Customer
    # column, etc.) — you can see exactly what was in each column index
    # for that view and compare against what DM was actually displaying.
    if diagnostics:
        wsd = wb.create_sheet("Diagnostics")
        wsd.append(["View", "Total rows", "Column mapping (canonical name -> index)"])
        for cell in wsd[1]:
            cell.font = header_font
            cell.fill = header_fill
        for d in diagnostics:
            cols_str = ", ".join(
                f"{k}={v}" for k, v in sorted(d.get("cols", {}).items(), key=lambda x: x[1])
            )
            wsd.append([d["view"], d.get("row_count", 0), cols_str])
        wsd.append([])
        wsd.append(["First 10 raw rows per view (column index in the header row):"])
        wsd[wsd.max_row][0].font = Font(bold=True)

        # Figure out the widest row across all views so we can lay out a
        # consistent column-index header.
        max_idx = 0
        for d in diagnostics:
            for row in d.get("raw_sample", []):
                for k in row.keys():
                    try:
                        max_idx = max(max_idx, int(k))
                    except (TypeError, ValueError):
                        pass

        for d in diagnostics:
            wsd.append([])
            label = f"--- {d['view']} ({d.get('row_count', 0)} rows total) ---"
            wsd.append([label])
            wsd[wsd.max_row][0].font = Font(bold=True)
            # Build the column-header row. If we captured real UIA header
            # names (from the grid header-row walk), use those — much easier
            # to read than "col0, col1, col2". Fall back to col-index labels
            # for any column where we don't have a name.
            raw_headers = d.get("raw_headers") or {}
            header_row = ["View"]
            for i in range(max_idx + 1):
                name = raw_headers.get(str(i)) or raw_headers.get(i)
                if name:
                    header_row.append(f"[{i}] {name}")
                else:
                    header_row.append(f"col{i}")
            wsd.append(header_row)
            for cell in wsd[wsd.max_row]:
                cell.font = header_font
                cell.fill = header_fill
            # Up to 10 kept sample rows
            for row in d.get("raw_sample", []):
                row_cells = [d["view"]]
                for i in range(max_idx + 1):
                    row_cells.append(row.get(str(i), ""))
                wsd.append(row_cells)
            if not d.get("raw_sample"):
                wsd.append([d["view"], "(no rows captured)"])

            # Up to 10 DROPPED rows (filtered out as multidrop sub-rows etc.).
            # If a row the user expected to see flagged is missing from the
            # Flagged tab, check here — it may have been mistakenly filtered.
            dropped_sample = d.get("dropped_sample", [])
            if dropped_sample:
                wsd.append([])
                wsd.append([f"  Dropped rows for {d['view']} (filtered as non-bookings):"])
                wsd[wsd.max_row][0].font = Font(italic=True)
                for row in dropped_sample:
                    row_cells = [f"{d['view']} (dropped)"]
                    for i in range(max_idx + 1):
                        row_cells.append(row.get(str(i), ""))
                    wsd.append(row_cells)

        # Widths
        wsd.column_dimensions["A"].width = 16
        wsd.column_dimensions["B"].width = 14
        wsd.column_dimensions["C"].width = 60
        for i in range(3, max_idx + 3):
            wsd.column_dimensions[get_column_letter(i)].width = 22

    # ---- Instructions sheet ----
    ws3 = wb.create_sheet("Instructions", 0)  # insert as first sheet
    ws3.append(["Delivery Master daily review"])
    ws3.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws3.append([])
    ws3.append(["How to use:"])
    ws3.append(["1. Review the Flagged tab. Change 'Include?' to 'No' on any row that shouldn't be emailed."])
    ws3.append(["2. Scan the Accepted tab. Change 'Promote?' to 'Yes' on any row that SHOULD be emailed (and add a note)."])
    ws3.append(["3. Save and close this file."])
    ws3.append(["4. Run 'Generate Email.bat' to produce the .eml file ready to send."])
    ws3.append([])
    ws3.append([f"Totals: {len(all_flagged)} flagged, {len(all_accepted)} accepted."])
    ws3.append([])
    ws3.append(["If something looks wrong — e.g. a column reading as Customer that's"])
    ws3.append(["actually Driver, or a view showing 0 flagged when you expect some —"])
    ws3.append(["check the Diagnostics tab. It shows the column mapping per view and"])
    ws3.append(["the first 10 raw rows so we can see exactly which column index"])
    ws3.append(["contained which data."])
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.column_dimensions["A"].width = 120

    out_path = Path(out_path)
    try:
        wb.save(out_path)
    except PermissionError:
        # The file is open in Excel — save with a timestamped name instead
        ts = datetime.now().strftime("%H%M%S")
        alt = out_path.with_stem(out_path.stem + f"_{ts}")
        wb.save(alt)
        print(f"WARNING: {out_path.name} is open in Excel — saved as {alt.name} instead.")
        return alt
    return out_path


# ---------- HTML / EML output ----------

def _html_escape(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&#39;"))


def generate_html(all_flagged):
    today_str = date.today().strftime("%d %B %Y")
    parts = [
        "<!DOCTYPE html>",
        "<html><body style='font-family: Calibri, Arial, sans-serif; font-size: 11pt;'>",
        "<p>Hi Lauren,</p>",
        f"<p>Daily Delivery Master check — {today_str}. Jobs below have a delivery date "
        f"in the past <i>and</i> at least one issue (not delivered, or Cust. Ref empty / "
        f"TBC / chased):</p>",
    ]
    by_view = defaultdict(list)
    for f in all_flagged:
        by_view[f["view"]].append(f)

    for view in VIEWS:
        rows = by_view.get(view, [])
        parts.append(f"<h3 style='margin-bottom:4px'>{_html_escape(view)} ({len(rows)} flagged)</h3>")
        if not rows:
            parts.append("<p style='margin-top:0'><i>Nothing flagged.</i></p>")
            continue
        parts.append(
            "<table style='border-collapse:collapse; font-size:10.5pt; margin-top:0' "
            "border='1' cellpadding='5' cellspacing='0'>"
        )
        parts.append(
            "<tr style='background:#f0f0f0'>"
            "<th>Our Ref</th><th>Customer</th><th>Status</th>"
            "<th>Del Date</th><th>Cust. Ref</th><th>Reason(s)</th></tr>"
        )
        for r in rows:
            parts.append(
                "<tr>"
                f"<td>{_html_escape(r['our_ref'])}</td>"
                f"<td>{_html_escape(r['customer'])}</td>"
                f"<td>{_html_escape(r['status'])}</td>"
                f"<td>{_html_escape(r['del_date'])}</td>"
                f"<td>{_html_escape(r['cust_ref'])}</td>"
                f"<td>{_html_escape('; '.join(r['reasons']))}</td>"
                "</tr>"
            )
        parts.append("</table>")

    parts.append("<p>Thanks,<br>Owen</p>")
    parts.append("</body></html>")
    return "\n".join(parts)


def generate_eml(html_content):
    msg = email.mime.multipart.MIMEMultipart("alternative")
    msg["Subject"] = f"Delivery Master daily check — {date.today().strftime('%d %b %Y')}"
    msg["To"] = EMAIL_TO
    msg.attach(email.mime.text.MIMEText(html_content, "html"))
    return msg.as_bytes()


# ---------- Main ----------

SCRIPT_VERSION = "v46 (deterministic column mapping from chooser display order — no OCR/heuristics needed)"


VIEW_RESULTS_DIR = HERE / "view_results"


# ---------- Column chooser ("burger menu") control ----------
#
# Header detection on a screen-rendered grid is fragile: different users have
# different DM column layouts, screen sizes change which columns fit, and OCR
# misses headers on narrower viewports. The robust answer is to FORCE the grid
# into a known minimal layout before we read it. The "burger" menu at the top
# right of the grid lets us tick/untick every available column, and the
# discovery script confirmed that:
#   - The button is a RadDropDownButton at the top-right of the grid.
#   - Clicking it (via click_input) opens a Popup descendant of DM.
#   - Inside that Popup are CheckBox elements, one per column, each with
#     name = the exact header text and TogglePattern.CurrentToggleState
#     returning 0=OFF / 1=ON.
#   - click_input() on a CheckBox reliably toggles the state.
#
# We use that to leave only these five columns visible before each view's
# read. Everything else gets unticked. The five exactly match the names DM
# uses for the CheckBox.name property.
REQUIRED_VISIBLE_COLUMNS = {
    "Our Ref",
    "Customer",
    "Cust. Ref",
    "Status",
    "Del Date Time",
}


def find_column_chooser_button(grid):
    """Locate the burger button (RadDropDownButton at the top-right of the
    grid's header strip). Returns the pywinauto element or None."""
    try:
        gr = grid.element_info.rectangle
    except Exception:
        return None
    candidates = []
    with gc_paused():
        try:
            for el in grid.descendants():
                try:
                    ei = el.element_info
                    cn = ei.class_name or ""
                    name = ei.name or ""
                    if "RadDropDownButton" in cn or "ControlPanel" in name or "ControlPanel" in cn:
                        r = ei.rectangle
                        candidates.append((el, (int(r.left), int(r.top), int(r.right), int(r.bottom))))
                except Exception:
                    pass
        except Exception:
            pass
    if not candidates:
        return None
    top_right = (gr.right, gr.top)
    candidates.sort(key=lambda c: abs(c[1][0] - top_right[0]) + abs(c[1][1] - top_right[1]))
    return candidates[0][0]


def _click_button(button_elem):
    """Click a button via UIA invoke first, fall back to click_input. Returns
    True on success."""
    try:
        w = UIAWrapper(button_elem.element_info)
        try:
            w.invoke()
            return True
        except Exception:
            pass
        try:
            w.click_input()
            return True
        except Exception as e:
            print(f"      click_input failed: {e}", flush=True)
            return False
    except Exception as e:
        print(f"      button wrap failed: {e}", flush=True)
        return False


def find_chooser_popup(dm, timeout=2.5):
    """The column-chooser popup is a `class_name='Popup'` descendant of the
    DM window (not a top-level window on this setup — confirmed by the
    discovery run). Wait up to `timeout` seconds for one to appear with
    non-zero rect."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        with gc_paused():
            try:
                for el in dm.descendants():
                    try:
                        ei = el.element_info
                        if (ei.class_name or "") == "Popup":
                            r = ei.rectangle
                            if (r.right - r.left) > 50 and (r.bottom - r.top) > 50:
                                return el
                    except Exception:
                        pass
            except Exception:
                pass
        time.sleep(0.2)
    return None


def _read_state_via_patterns(iuia_element):
    """Try every UIA pattern that might tell us this CheckBox's tick state,
    in order of reliability. Returns 'ON' / 'OFF' / None if nothing worked.

    On v42 we used TogglePattern only; on the colleague's setup it returned
    nothing for every CheckBox (probably because the popup was already
    closing or the COM pointer had been recycled by the time we queried).
    The fallbacks let us recover state from SelectionItemPattern and from
    the LegacyIAccessible 0x10 (STATE_SYSTEM_CHECKED) bit.
    """
    try:
        from ctypes import POINTER
        from comtypes import cast  # type: ignore
        from comtypes.gen.UIAutomationClient import (  # type: ignore
            IUIAutomationTogglePattern,
            IUIAutomationSelectionItemPattern,
            IUIAutomationLegacyIAccessiblePattern,
        )
    except Exception:
        return None
    UIA_TogglePatternId = 10024
    UIA_SelectionItemPatternId = 10010
    UIA_LegacyIAccessiblePatternId = 10018
    try:
        ptr = iuia_element.GetCurrentPattern(UIA_TogglePatternId)
        if ptr:
            tp = cast(ptr, POINTER(IUIAutomationTogglePattern))
            s = tp.CurrentToggleState
            if s == 0:
                return "OFF"
            if s == 1:
                return "ON"
    except Exception:
        pass
    try:
        ptr = iuia_element.GetCurrentPattern(UIA_SelectionItemPatternId)
        if ptr:
            sip = cast(ptr, POINTER(IUIAutomationSelectionItemPattern))
            return "ON" if sip.CurrentIsSelected else "OFF"
    except Exception:
        pass
    try:
        ptr = iuia_element.GetCurrentPattern(UIA_LegacyIAccessiblePatternId)
        if ptr:
            lp = cast(ptr, POINTER(IUIAutomationLegacyIAccessiblePattern))
            state_bits = lp.CurrentState
            return "ON" if (state_bits & 0x10) else "OFF"
    except Exception:
        pass
    return None


def read_chooser_checkboxes(popup):
    """Walk the popup and return every CheckBox as
    [{'name': column_text, 'state': 'ON'|'OFF'|None, 'element': el}, ...].
    """
    out = []
    with gc_paused():
        try:
            for el in popup.descendants():
                try:
                    ei = el.element_info
                    if (ei.control_type or "") != "CheckBox":
                        continue
                    name = (ei.name or "").strip()
                    if not name:
                        continue
                    state = None
                    try:
                        state = _read_state_via_patterns(ei.element)
                    except Exception:
                        pass
                    out.append({"name": name, "state": state, "element": el})
                except Exception:
                    pass
        except Exception:
            pass
    return out


def toggle_chooser_checkbox(box_entry):
    """Flip one column-chooser CheckBox. The discovery confirmed click_input
    is the reliable method on this setup (TogglePattern.Toggle and
    InvokePattern.Invoke both no-op'd for reasons we don't fully understand,
    but a physical click works). Returns True on success."""
    try:
        UIAWrapper(box_entry["element"].element_info).click_input()
        time.sleep(0.12)
        return True
    except Exception as e:
        print(f"      toggle of {box_entry['name']!r} failed: {e}", flush=True)
        return False


def close_column_chooser():
    """Dismiss the column-chooser popup by sending Escape."""
    try:
        send_keys("{ESC}")
        time.sleep(0.4)
    except Exception:
        pass


def _build_chooser_order_mapping(boxes_in_order):
    """After applying the chooser plan, the clipboard exports VISIBLE columns
    in their data-model (chooser) order, with consecutive tab positions
    0..N-1.

    This gives us a deterministic mapping that doesn't need OCR or content
    heuristics: enumerate the chooser entries in display order, and each
    one in REQUIRED_VISIBLE_COLUMNS gets the next tab position. Empirically
    confirmed across views where OCR succeeded (Katie, Steven, Jamie C,
    Complete) — every one of them matched this mapping. The only views
    where the heuristic-based detection got it wrong (Kyle and In Progress
    in v45) would have come out right under this approach.

    Returns {canonical_name: tab_position} or {} if anything looks off
    (e.g. one of the required columns isn't present in the chooser at all).
    """
    mapping = {}
    names_seen = {b["name"] for b in boxes_in_order}
    missing = REQUIRED_VISIBLE_COLUMNS - names_seen
    if missing:
        print(
            f"    !!! these required columns aren't in the chooser at all: "
            f"{sorted(missing)} — falling back to OCR/heuristics",
            flush=True,
        )
        return {}
    tab = 0
    for box in boxes_in_order:
        if box["name"] in REQUIRED_VISIBLE_COLUMNS:
            mapping[box["name"]] = tab
            tab += 1
    return mapping


def prepare_minimal_layout(dm, grid):
    """Open the column chooser and ensure only REQUIRED_VISIBLE_COLUMNS are
    ticked. Returns (success_bool, chooser_mapping_dict). On success the
    mapping is {canonical_name: clipboard_tab_position} based on chooser
    order — the deterministic mapping the worker should prefer. Callers
    carry on with detection-based logic if either is empty / failed.

    Key design points learned from v42's failure:
      - Telerik's column-chooser popup is fragile. Re-reading it after the
        first walk can come back with empty / None states (the popup may
        have started to close, or COM pointers have been recycled). We
        therefore read state ONCE, build a complete plan, then apply every
        toggle in a single pass without touching the popup again.
      - State reading goes through three UIA patterns (Toggle, then
        SelectionItem, then the LegacyIAccessible 0x10 bit) so we don't
        come back with `None` because the first method happened to be
        unavailable.
      - Every CheckBox's name + observed state is logged so the next time
        something goes wrong we can see exactly which boxes the script
        saw and how it scored them.
    """
    print(f"  Enforcing minimal column layout for this view…", flush=True)
    button = find_column_chooser_button(grid)
    if button is None:
        print("    !!! column-chooser button not found — skipping layout reset", flush=True)
        return False, {}
    if not _click_button(button):
        print("    !!! couldn't open the column chooser — skipping", flush=True)
        return False, {}
    # Brief settle delay so Telerik can finish painting/wiring the popup
    # before we walk it. v42 read straight away and got None for every
    # state, suggesting the popup wasn't fully alive yet.
    time.sleep(0.6)
    popup = find_chooser_popup(dm)
    if popup is None:
        print("    !!! column-chooser popup didn't appear — skipping", flush=True)
        close_column_chooser()
        return False, {}

    boxes = read_chooser_checkboxes(popup)
    print(f"    Column chooser has {len(boxes)} entries", flush=True)
    if not boxes:
        close_column_chooser()
        return False, {}

    # Dump every box's state so we can see exactly what the script is
    # working from. Cheap and one-time per view.
    for b in boxes:
        marker = "+" if b["name"] in REQUIRED_VISIBLE_COLUMNS else " "
        print(f"      {marker}  {b['state']!s:<6}  {b['name']!r}", flush=True)

    # Build the plan up front. We don't re-read the popup between toggles
    # because that's exactly what broke v42 — the second walk silently
    # returned None states. Each `toggle_chooser_checkbox` click flips the
    # in-grid checkbox via Win32 click_input; we trust that click without
    # verifying state again.
    plan_on = []   # required column that isn't currently ON → tick it
    plan_off = []  # non-required column that's currently ON → untick it
    unknown_state = []  # state read failed for this box
    for box in boxes:
        name = box["name"]
        state = box["state"]
        if name in REQUIRED_VISIBLE_COLUMNS:
            if state == "ON":
                continue
            if state == "OFF":
                plan_on.append(box)
            else:
                unknown_state.append(box)
        else:
            if state == "ON":
                plan_off.append(box)
            elif state == "OFF":
                continue
            else:
                unknown_state.append(box)

    if unknown_state:
        # Couldn't read state for some boxes. We deliberately DON'T toggle
        # these — toggling blind could flip a column we wanted left alone.
        # The detection layer downstream will deal with whatever is shown.
        names = ", ".join(repr(b["name"]) for b in unknown_state[:12])
        more = "" if len(unknown_state) <= 12 else f" (+{len(unknown_state)-12} more)"
        print(f"    !!! couldn't read tick state for {len(unknown_state)} box(es): {names}{more}", flush=True)
        print(f"        (leaving those as-is — pattern lookups returned None)", flush=True)

    print(f"    Plan: turn ON {len(plan_on)} (required & off), "
          f"turn OFF {len(plan_off)} (extras currently on)", flush=True)

    # Apply ON toggles first so we never leave the grid temporarily empty.
    changes = 0
    for box in plan_on:
        print(f"      turning ON:  {box['name']!r}", flush=True)
        if toggle_chooser_checkbox(box):
            changes += 1
    for box in plan_off:
        print(f"      turning OFF: {box['name']!r}", flush=True)
        if toggle_chooser_checkbox(box):
            changes += 1

    close_column_chooser()
    # Let the grid repaint with the new layout before Phase 1 walks it.
    time.sleep(1.5)
    print(f"    Layout reset complete ({changes} toggles applied)", flush=True)

    # Build the deterministic column-to-tab-position mapping from the
    # chooser's display order. Telerik's clipboard exports VISIBLE columns
    # in their data-model (chooser) order, so the Nth required column in
    # the chooser is at clipboard tab position N. No OCR or content
    # heuristics needed — this works regardless of how the user has the
    # columns arranged on screen.
    chooser_mapping = _build_chooser_order_mapping(boxes)
    if chooser_mapping:
        order_str = ", ".join(
            f"{n}={i}" for n, i in sorted(chooser_mapping.items(), key=lambda kv: kv[1])
        )
        print(f"    Chooser-order column mapping: {order_str}", flush=True)
    return True, chooser_mapping


def _process_one_view_inline(view_name):
    """Worker mode: process ONE view and write its results to JSON.
    Designed to run in its own subprocess so a native crash here doesn't
    kill the orchestrator."""
    import json
    print(f"==== Worker for view: {view_name!r} ====", flush=True)
    dm = find_dm()
    if dm is None:
        print("DM not found.")
        sys.exit(2)
    rules = load_rules(RULES_PATH)
    # Load this company's known customer list - the strongest signal for
    # telling the Customer column from the Cust. Ref column (the real Customer
    # column matches the list at a high rate; the reference column doesn't).
    dm_company, tms_customer_names = _load_tms_customer_names()
    if tms_customer_names:
        print(f"  TMS customer list for '{dm_company}': {len(tms_customer_names)} names "
              f"(used to disambiguate Customer vs Cust. Ref)", flush=True)
    else:
        print(f"  No TMS customer list for '{dm_company}' - using content heuristics only",
              flush=True)
    if not switch_view(dm, view_name):
        print(f"Couldn't switch to {view_name}")
        sys.exit(3)
    grid, panel = find_grid_and_panel(dm)
    if grid is None or panel is None:
        print("Grid/panel not found.")
        sys.exit(4)

    # Phase 0 — force a minimal column layout via DM's column-chooser menu.
    # This guarantees the rest of the worker sees only Our Ref, Customer,
    # Cust. Ref, Status, Del Date Time — no matter how the user has the grid
    # arranged for everyday use. Eliminates the cross-machine layout drift
    # that was causing wrong column mappings on some setups.
    #
    # prepare_minimal_layout ALSO returns the deterministic column mapping
    # ({canonical_name: tab_position}) derived from chooser display order.
    # That's the most reliable answer for which clipboard tab is which
    # canonical column — no OCR or content heuristics involved. We carry
    # that through to Phase 2 and use it directly.
    chooser_mapping = {}
    try:
        _, chooser_mapping = prepare_minimal_layout(dm, grid)
    except Exception as e:
        print(f"  !!! layout-reset errored ({e}); carrying on with whatever's visible.", flush=True)
    # The grid may have been redrawn — re-find it before Phase 1 walks cells.
    grid, panel = find_grid_and_panel(dm)
    if grid is None or panel is None:
        print("Grid/panel disappeared after layout reset.")
        sys.exit(5)

    # Phase 1 — column detection from visible rows
    sample_rows = read_visible_rows(panel)
    if not sample_rows:
        print("Phase 1 sample empty (refreshing); waiting 4s...", flush=True)
        time.sleep(4)
        grid, panel = find_grid_and_panel(dm)
        if panel is not None:
            sample_rows = read_visible_rows(panel)
    # Brief settle delay before OCR — even if the view didn't switch (because
    # DM was already on this view), the grid may not have finished painting
    # when the worker spawned. Without this delay In Progress (the first view)
    # often comes back with 0 OCR words.
    time.sleep(1.0)

    # Header detection priority:
    #   1) Walk the grid's UIA header row directly (most reliable — gets EVERY
    #      header in display order, no OCR clipping, no content guessing)
    #   2) OCR the rendered header strip (works when (1) doesn't return enough)
    #   3) UIA TableItem pattern (Telerik usually returns nothing, but try)
    #   4) Content-based heuristics (fallback only)
    # v2: read the visible header row (left-to-right == clipboard tab order)
    # and build a {canonical: tab} map from the LABELS. This is screen- and
    # DPI-independent and resolves the columns content alone can't.
    header_seed, raw_headers = (
        detect_columns_via_header_row(grid) if grid is not None else ({}, {})
    )
    ordered_labels = [raw_headers[i] for i in sorted(raw_headers)] if raw_headers else []
    header_map = (
        _header_map_from_labels(ordered_labels)
        if (_header_map_from_labels and ordered_labels) else {}
    )
    if header_map:
        print(f"  header-row labels -> {header_map}", flush=True)
    else:
        print("  header-row read returned nothing usable — content mapping will carry it",
              flush=True)
    # OCR is kept for DIAGNOSTICS ONLY in v2. It is the single screen/DPI-
    # dependent step, so it must never set or override the mapping — it can
    # only be logged for reference.
    ocr_seed = {}
    try:
        ocr_seed = detect_columns_via_ocr(grid, panel, dm=dm) if panel is not None else {}
    except Exception as _oe:
        print(f"  (OCR diagnostic skipped: {_oe})", flush=True)
    if ocr_seed:
        print(f"  [diagnostic] OCR read (NOT used for mapping): {ocr_seed}", flush=True)
    # Provisional mapping from the Phase-1 sample (the authoritative mapping is
    # computed later on the full clipboard data).
    if _resolve_columns_v2 is not None:
        cols, _conf0, _diag0 = _resolve_columns_v2(
            sample_rows, header_map=header_map,
            customer_names=tms_customer_names, log=lambda *a, **k: None)
    else:
        cols, _diag0 = detect_columns(sample_rows)
    print(f"Columns ({len(sample_rows)}-row sample, header_map={header_map}): {cols}",
          flush=True)

    # Phase 2 — checkpoint to JSON after EACH page so a crash on the next-page
    # navigation doesn't lose what we've already gathered (this is what bit
    # us on the Complete view: page 1's 408 rows were thrown away when the
    # page 2 click crashed Python).
    VIEW_RESULTS_DIR.mkdir(exist_ok=True)
    out = VIEW_RESULTS_DIR / f"{view_name.replace(' ', '_')}.json"

    def _save_progress(page_num, rows_so_far):
        # Detect columns FRESH from whatever rows we're about to save. We
        # used to reuse the Phase 1 sample's `cols`, but Phase 1 rows are
        # keyed by UIA display index and clipboard rows are keyed by tab
        # position — feeding clipboard rows through Phase 1's cols was
        # giving wildly wrong flagged/accepted counts in per-page
        # checkpoints. Detecting from `rows_so_far` itself guarantees the
        # spaces match.
        cur_cols = detect_columns(rows_so_far)[0] if rows_so_far else {}
        flg, acc, neli = categorize_rows_three_way(
            rows_so_far, cur_cols, view_name, rules)
        try:
            with open(out, "w", encoding="utf-8") as f:
                json.dump({
                    "view": view_name,
                    "row_count": len(rows_so_far),
                    "cols": dict(cur_cols),
                    "flagged": flg,
                    "accepted": acc,
                    "not_eligible": neli,
                    "pages_completed": page_num,
                    "partial": True,
                }, f, default=str, indent=2)
            print(
                f"  [page-{page_num} checkpoint] {len(rows_so_far)} rows, "
                f"{len(flg)} not accepted, {len(acc)} accepted, "
                f"{len(neli)} not eligible -> {out.name}", flush=True)
        except Exception as e:
            print(f"  [page-{page_num} checkpoint] save failed: {e}", flush=True)

    # Save the Phase 1 sample as a fallback checkpoint BEFORE we touch anything
    # else. If Phase 2 native-crashes the worker (which happens on some setups),
    # the orchestrator will at least surface these ~25 visible rows instead of
    # nothing. We do this regardless of --no-clipboard mode, because scroll mode
    # can crash just as readily as clipboard mode.
    try:
        _save_progress(0, sample_rows or [])
    except Exception:
        pass

    skip_clipboard = "--no-clipboard" in sys.argv
    rows = None
    if not skip_clipboard:
        try:
            rows = read_all_rows_via_clipboard(dm, grid, panel, on_page=_save_progress)
        except Exception as e:
            print(f"Clipboard errored ({e}); falling back to scroll.", flush=True)
    else:
        print("  [no-clipboard mode] skipping clipboard, using scroll-based read", flush=True)
    if not rows:
        try:
            rows = read_all_rows_paged(dm, grid, panel, on_page=_save_progress)
        except Exception as e:
            print(f"Scroll-based read errored ({e}); using Phase 1 sample as fallback.", flush=True)
            rows = sample_rows or []
    print(f"Total rows: {len(rows) if rows else 0}", flush=True)

    # Note: v41 introduced a "rosetta stone" translation here that mapped
    # clipboard tab positions to UIA display indices. It was based on a
    # wrong premise — I'd misread legitimate operator placeholder text
    # ("tbc", "CHECKING", "Luton Regional Tariff", etc.) in the Customer
    # and Cust. Ref fields as evidence of a column-order mismatch. In
    # reality the clipboard tab order DOES match the display index order
    # on this setup, and that translation is unnecessary. The helper
    # functions (find_clipboard_anchor / build_display_to_tab /
    # translate_clip_rows_to_display) are still defined above in case
    # they're ever needed, but they're not invoked here.
    display_to_tab = {}

    # Column mapping decision.
    #
    # Three sources are available, in DECREASING ORDER OF RELIABILITY:
    #
    #   1. The chooser-order mapping (chooser_mapping) built in Phase 0.
    #      Telerik exports VISIBLE columns to the clipboard in their
    #      data-model order — the same order the chooser shows them in —
    #      so the Nth required column in the chooser is the Nth tab in
    #      the clipboard. This is deterministic: no OCR, no heuristics,
    #      no screen-size sensitivity. If the chooser pass succeeded
    #      (chooser_mapping has 5 entries), USE IT.
    #
    #   2. Content-only re-detection on the full clipboard data. Works
    #      well when only the 5 essentials are visible because the
    #      deterministic columns (Our Ref via BT-pattern, Del Date Time
    #      via dates, Status via status keywords) account for 3 of 5.
    #      The remaining 2 (Customer / Cust. Ref) are heuristic and
    #      occasionally swap.
    #
    #   3. Phase 1 cols from the UIA sample. Lowest signal-to-noise.
    #
    # We always try (1) first, fall through to (2) if the chooser failed,
    # and only use (3) if both higher sources failed. We also log a
    # warning if (2) disagrees with (1) — that would tell us something's
    # changed about how Telerik exports clipboards on a particular setup.
    # ---- v2 column resolution on the FULL clipboard data -----------------
    # Content-first and screen-independent: Our Ref / Status / dates identify
    # themselves from their values; the header map (if we read one) resolves
    # the columns content alone can't (Customer vs Cust. Ref, and the delivery
    # date when several date columns are present). The verification gate
    # refuses to emit a mapping that fails its own data — so a wrong guess is
    # caught and logged here instead of silently emailing the wrong column.
    if _resolve_columns_v2 is not None and rows:
        v2_cols, v2_conf, v2_diag = _resolve_columns_v2(
            rows, header_map=header_map, customer_names=tms_customer_names, log=print)
        print(f"v2 resolver mapping: {v2_cols} (confidence={v2_conf})", flush=True)
        if v2_conf == "high":
            cols = v2_cols
        else:
            # Low confidence — keep whatever v2 pinned (Our Ref / Status / a
            # lone date column stay trustworthy) and let the downstream manual
            # column-mapper resolve the rest. We deliberately do NOT fall back
            # to the chooser-order mapping: a live run proved the chooser's
            # display order does NOT match the clipboard tab order on this
            # setup (it put Status/Customer/Cust. Ref in the wrong tabs), so
            # trusting it would silently mislabel columns.
            problems = v2_diag.get("problems") or v2_diag.get("reasons", [])
            print("  !!! v2 mapping is LOW confidence: " + "; ".join(problems), flush=True)
            print("      keeping content-pinned columns; the manual mapper "
                  "should confirm the rest for this view.", flush=True)
            if v2_cols:
                cols = v2_cols
    elif chooser_mapping and len(chooser_mapping) == len(REQUIRED_VISIBLE_COLUMNS):
        print(f"Using chooser-order mapping (deterministic): {chooser_mapping}", flush=True)
        cols = chooser_mapping
    elif rows:
        full_cols, _ = detect_columns(rows)
        if full_cols != cols:
            print(f"Columns from full clipboard data ({len(rows)} rows): {full_cols}", flush=True)
            cols = full_cols

    # NOTE: previous versions of this script tried to filter out "non-booking"
    # rows based on the value at the Our-Ref column. That was a bad idea —
    # Owen confirmed every row in DM is a booking, so any filter at this layer
    # is hiding real data. We KEEP every row the clipboard returned.
    dropped_sample = []

    flagged, accepted, not_eligible = categorize_rows_three_way(
        rows or [], cols, view_name, rules)

    # Data-validation: the most authoritative source for "how many rows
    # DM said there are" is _LAST_RUN_MAX_TOTAL, which the read paths
    # update every time they see a pagination indicator. That value
    # survives a later failed page (which is exactly what bit us on the
    # Complete view: clipboard failed on page 2, then find_pagination
    # at the end returned 0, then the validation banner claimed
    # "300/0 - all accounted for"). Fall through to a fresh probe and
    # then to len(rows) for single-page views.
    expected_total = _LAST_RUN_MAX_TOTAL
    if expected_total is None:
        try:
            ps = find_pagination_status(dm)
            if ps is not None:
                _start, _end, total = ps
                expected_total = int(total) if total > 0 else None
        except Exception:
            expected_total = None
    captured_total = len(rows) if rows else 0
    if expected_total is None:
        expected_total = captured_total
    missing = max(0, expected_total - captured_total)
    print(
        f"Not Accepted: {len(flagged)}, Accepted: {len(accepted)}, "
        f"Not Eligible: {len(not_eligible)}", flush=True)
    if missing:
        print(
            f"  !!! data-validation: captured {captured_total} of "
            f"{expected_total} on-screen rows for view "
            f"{view_name!r} — {missing} appear to be missing. "
            "Re-running the view usually picks them up.", flush=True)
    else:
        print(
            f"  data-validation: captured {captured_total}/"
            f"{expected_total} rows for view {view_name!r} - all "
            "accounted for.", flush=True)

    # Capture a sample of raw rows (column-index -> value) for the
    # Diagnostics tab. Helps debug "why was this row not flagged?" — the user
    # can see exactly what was in each column index for this view and tell us
    # whether the column mapping is right.
    raw_sample = []
    for r in (rows or [])[:10]:
        raw_sample.append({str(k): v for k, v in r.items()})

    # Save full rows too so a column-mapper UI can re-categorise with a
    # user-supplied mapping without re-scraping Delivery Master.
    all_rows = [{str(k): v for k, v in r.items()} for r in (rows or [])]

    # Final write — overwrites any per-page checkpoint with the complete result
    with open(out, "w", encoding="utf-8") as f:
        json.dump({
            "view": view_name,
            "row_count": captured_total,
            "expected_total": expected_total,
            "missing_count": missing,
            "cols": dict(cols),
            "raw_headers": {str(k): v for k, v in (raw_headers or {}).items()},
            "display_to_tab": {str(k): v for k, v in (display_to_tab or {}).items()},
            "flagged": flagged,
            "accepted": accepted,
            "not_eligible": not_eligible,
            "raw_sample": raw_sample,
            "all_rows": all_rows,
            "dropped_sample": dropped_sample,
            "partial": False,
        }, f, default=str, indent=2)
    print(f"Wrote: {out}", flush=True)


def _combine_view_results():
    """Read every per-view JSON file and write the final review xlsx + summary."""
    import json
    print("\n=== Combining view results ===", flush=True)
    all_flagged = []
    all_accepted = []
    all_not_eligible = []
    summary = []
    diagnostics = []  # list of {view, cols, row_count, raw_sample}
    # Cross-view rollup for the data-validation banner.
    grand_captured = 0
    grand_expected = 0
    grand_missing = 0
    views_with_gaps: list[tuple[str, int, int]] = []
    for view in VIEWS:
        slug = view.replace(" ", "_")
        path = VIEW_RESULTS_DIR / f"{slug}.json"
        if not path.exists():
            print(f"  {view:>12}: NO RESULT (subprocess likely crashed)", flush=True)
            summary.append((view, 0, 0, 0, 0, "no result"))
            diagnostics.append({"view": view, "cols": {}, "row_count": 0, "raw_sample": []})
            continue
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            flagged = data.get("flagged", [])
            accepted = data.get("accepted", [])
            not_eligible = data.get("not_eligible", [])
            row_count = data.get("row_count", 0)
            expected_total = data.get("expected_total", row_count)
            missing = data.get(
                "missing_count", max(0, expected_total - row_count))
            cols = data.get("cols", {})
            cols_str = ", ".join(f"{k}={v}" for k, v in sorted(cols.items(), key=lambda x: x[1]))
            all_flagged.extend(flagged)
            all_accepted.extend(accepted)
            all_not_eligible.extend(not_eligible)
            grand_captured += row_count
            grand_expected += expected_total
            grand_missing += missing
            if missing:
                views_with_gaps.append((view, row_count, expected_total))
            summary.append(
                (view, row_count, len(flagged), len(accepted),
                 len(not_eligible), cols_str))
            diagnostics.append({
                "view": view,
                "cols": cols,
                "row_count": row_count,
                "expected_total": expected_total,
                "missing_count": missing,
                "raw_sample": data.get("raw_sample", []),
                "raw_headers": data.get("raw_headers", {}),
                "dropped_sample": data.get("dropped_sample", []),
            })
            gap_note = (
                f"  [!! MISSING {missing}]" if missing else "")
            print(
                f"  {view:>12}: {row_count:>3}/{expected_total:<3} rows  "
                f"({len(flagged):>3} not accepted, {len(accepted):>3} "
                f"accepted, {len(not_eligible):>3} not eligible)"
                f"{gap_note}   ({cols_str})", flush=True)
        except Exception as e:
            print(f"  {view:>12}: ERROR reading {path.name}: {e}", flush=True)
            summary.append((view, 0, 0, 0, 0, f"json error: {e}"))
            diagnostics.append({"view": view, "cols": {}, "row_count": 0, "raw_sample": []})

    # Cross-view data-validation banner.
    print("", flush=True)
    if grand_missing == 0:
        print(
            f"  data-validation: captured all "
            f"{grand_captured} rows shown across every view.",
            flush=True)
    else:
        print(
            f"  !!! data-validation: captured {grand_captured} of "
            f"{grand_expected} rows ({grand_missing} missing) across "
            "all views. Affected views:", flush=True)
        for v, got, want in views_with_gaps:
            print(f"      - {v}: {got} of {want}", flush=True)

    def _row_sort_key(r):
        try:
            view_idx = VIEWS.index(r["view"])
        except ValueError:
            view_idx = len(VIEWS)
        dt = parse_dm_date(r["del_date"]) or datetime.max
        return (view_idx, dt)
    all_flagged.sort(key=_row_sort_key)
    all_accepted.sort(key=_row_sort_key)

    xlsx_path = HERE / "dm_daily_review.xlsx"
    saved_to = None
    try:
        saved_to = write_review_xlsx(all_flagged, all_accepted, xlsx_path, diagnostics=diagnostics)
    except Exception as e:
        print(f"\n!!! xlsx write failed: {e}", flush=True)

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for v, n, f, a, ne, info in summary:
        print(
            f"  {v:>12}: {n:>3} rows  ({f:>3} not accepted, "
            f"{a:>3} accepted, {ne:>3} not eligible)   ({info})")
    print(
        f"\nTotal: {len(all_flagged)} not accepted, "
        f"{len(all_accepted)} accepted, "
        f"{len(all_not_eligible)} not eligible")
    if saved_to:
        print(f"\nReview workbook: {saved_to}")

    # ---- Record this run to the shared database (history + result file) ----
    # Local-first: cloud_sync queues the run if Supabase is unreachable, so a
    # failure here never affects the actual check output.
    if _cloud is not None:
        try:
            company = (os.environ.get("DM_COMPANY") or "").strip().lower()
            if company not in ("north", "south"):
                try:
                    ir_dir = SCRIPT_DIR.parent / "invoicing_rules"
                    if str(ir_dir) not in sys.path:
                        sys.path.insert(0, str(ir_dir))
                    import invoice_store  # type: ignore
                    company = invoice_store.get_active_company()
                except Exception:
                    company = "north"
            # Date period the run covered (min/max delivery date across results).
            dates = []
            for r in all_flagged + all_accepted:
                d = parse_dm_date(r.get("del_date"))
                if d:
                    dates.append(d)
            period_from = min(dates).date().isoformat() if dates else None
            period_to = max(dates).date().isoformat() if dates else None
            by_view = {
                v: {
                    "rows": n,
                    "flagged": f,
                    "accepted": a,
                    "not_eligible": ne,
                } for v, n, f, a, ne, _info in summary}
            views_done = [
                v for v, n, f, a, ne, _info in summary
                if n or f or a or ne]
            result_summary = {
                "total_flagged": len(all_flagged),
                "total_accepted": len(all_accepted),
                "total_not_eligible": len(all_not_eligible),
                "grand_captured": grand_captured,
                "grand_expected": grand_expected,
                "grand_missing": grand_missing,
                "by_view": by_view,
            }
            filters = {"company": company, "views": views_done, "mode": "auto"}
            run_id = _cloud.record_run(
                script="dm_daily_check", company=company,
                period_from=period_from, period_to=period_to,
                filters=filters, result_summary=result_summary,
                flags=all_flagged, result_file=str(saved_to) if saved_to else None,
            )
            # Drain any previously-queued runs while we have connectivity.
            try:
                drained = _cloud.drain_pending()
            except Exception:
                drained = 0
            print(f"\n[cloud] recorded run {run_id} "
                  f"({len(all_flagged)} flags, period {period_from}..{period_to})"
                  + (f"; flushed {drained} queued run(s)" if drained else ""),
                  flush=True)
        except Exception as e:
            print(f"\n[cloud] run recording skipped: {e}", flush=True)


def _process_one_view_subprocess(view_name, no_clipboard=False):
    """Spawn a fresh Python subprocess to process ONE view.

    Key flags:
      -u             force unbuffered stdio in the child. Without this a
                     mid-import crash loses every print() the child made
                     because Python's stdout buffer never flushed - which
                     is the "subprocess exited with code 1, no output"
                     case the user reported.
      PYTHONUNBUFFERED=1    belt-and-braces equivalent (some Python builds
                     ignore -u under certain stdio redirections).

    Output is captured so we can ECHO it back into our stdout AFTER the
    process exits - if the child died mid-print, capture_output preserves
    whatever made it into the pipe. Inheriting stdio was losing this on
    native crashes."""
    # Under a PyInstaller --onefile build, sys.executable is the listener
    # .exe and __file__ is a temp path that can't be re-launched directly.
    # Detect frozen mode and re-launch THIS same .exe with a sentinel arg;
    # cal_listener/__main__.py routes that back to _process_one_view_inline.
    if getattr(sys, "frozen", False):
        cmd = [sys.executable, "--engine-view", view_name]
    else:
        cmd = [sys.executable, "-u", str(Path(__file__)), "--view", view_name]
    if no_clipboard:
        cmd.append("--no-clipboard")
    env = dict(os.environ)
    env["PYTHONUNBUFFERED"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    print(f"  + launching: {' '.join(cmd[1:])}", flush=True)
    # On Windows, hide the per-view console window. Without this each
    # view spawns a black cmd flash on screen (6+ per run) which is
    # ugly and confusing. The output is captured via capture_output
    # anyway so we don't need a visible console for it.
    extra: dict = {}
    if sys.platform == "win32":
        try:
            extra["creationflags"] = subprocess.CREATE_NO_WINDOW
        except AttributeError:
            pass
    # LIVE-STREAM the subprocess output instead of buffering with
    # capture_output. Buffering meant each view ran silently for 10-30s
    # then dumped its full output in one go AFTER exiting — which made
    # the orchestrator look frozen and hid early-import crashes until
    # after the (potentially silent) crash. With Popen + line-iteration
    # we see exactly what's happening as it happens. We still preserve
    # the full transcript so a native crash that loses output before
    # the pipe drains has at least the captured lines.
    transcript: list = []
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            encoding="utf-8",
            errors="replace",
            env=env,
            **extra,
        )
        deadline = time.time() + 300
        assert proc.stdout is not None
        while True:
            line = proc.stdout.readline()
            if not line:
                # EOF — child closed stdout (either exited or crashed).
                break
            line = line.rstrip()
            if line:
                print(f"    {line}", flush=True)
                transcript.append(line)
            if time.time() > deadline:
                proc.kill()
                print(f"  !!! subprocess timed out after 300s", flush=True)
                return -1
        rc = proc.wait(timeout=10)
        if not transcript:
            print(f"    (subprocess produced NO output before exiting "
                  f"with code {rc} - probable native crash in module-level "
                  f"import / pywinauto init)", flush=True)
        return rc
    except subprocess.TimeoutExpired:
        print(f"  !!! subprocess timed out after 300s", flush=True)
        return -1
    except Exception as e:
        print(f"  !!! subprocess errored: {e}", flush=True)
        return -2


def _orchestrator():
    """Default mode: spawn a subprocess per view, then combine results."""
    dm = find_dm()
    if dm is None:
        print("Delivery Master not found. Open it and try again.")
        sys.exit(1)
    print(f"Connected: {dm.window_text()!r}")
    print("")
    print("HOW TO STOP:")
    print("  - Press Ctrl+C in this window, OR")
    print(f"  - Create a file named STOP.txt in {HERE}")
    print("  Either signal makes the script finish the current view, save,")
    print("  and exit cleanly. The xlsx will still be written for whatever")
    print("  views completed.")
    print("")

    ocr_backend = _check_ocr_backend()
    if ocr_backend:
        print(f"OCR backend available: {ocr_backend}")
        print("  -> column headers will be read from the rendered grid.")
    else:
        print("=" * 70)
        print("WARNING: NO OCR BACKEND FOUND.")
        print("Without OCR the script can't read the actual column headers, so")
        print("Customer / Cust. Ref / etc. are detected by guessing at the data,")
        print("which sometimes picks the wrong column.")
        print("")
        print("To fix this, install Tesseract (5-minute one-time setup):")
        print("  1. Download installer:")
        print("     https://github.com/UB-Mannheim/tesseract/wiki")
        print("     (pick the 64-bit installer near the top of that page)")
        print("  2. Run the installer with default settings")
        print("     (it'll go to C:\\Program Files\\Tesseract-OCR\\)")
        print("  3. Re-run this script. It auto-detects.")
        print("=" * 70)
    print("")

    ensure_rules_workbook(RULES_PATH)
    rules = load_rules(RULES_PATH)
    if rules:
        print(f"Loaded {len(rules)} custom rule(s) from {RULES_PATH.name}:")
        for r in rules:
            who = r["customer_contains"] or "(all customers)"
            print(f"  - {who}: Cust. Ref {r['rule_type']} {r['values']}")
    else:
        print(f"No custom rules loaded (check {RULES_PATH.name}).")

    # Clear any prior per-view JSON, including any leftover STOP.txt from a
    # previous run.
    if VIEW_RESULTS_DIR.exists():
        for p in VIEW_RESULTS_DIR.glob("*.json"):
            try:
                p.unlink()
            except Exception:
                pass
    stop_path = HERE / "STOP.txt"
    if stop_path.exists():
        try:
            stop_path.unlink()
            print("(removed leftover STOP.txt from a previous run)")
        except Exception:
            pass

    # Drop our COM/UIA references before we start spawning workers.
    # We do NOT call gc.collect() here on purpose — explicit collection would
    # walk reference cycles of pywinauto wrappers and segfault on stale COM
    # pointers on the broken-comtypes machines. Each subprocess is a fresh
    # Python process, so it doesn't inherit our heap; the orchestrator's
    # leaked wrappers are reclaimed when the whole process exits via os._exit
    # (see end of _orchestrator).
    dm = None
    rules = None

    interrupted = False
    # Once we see one native clipboard crash, stop using clipboard for the rest
    # of the run — saves time and avoids guaranteed crashes on the same PC.
    disable_clipboard_globally = False
    try:
        for view in VIEWS:
            if _check_stop_file():
                print(f"\n!!! STOP.txt detected — skipping remaining views.", flush=True)
                interrupted = True
                break
            print(f"\n=== Spawning subprocess for view: {view} ===", flush=True)
            rc = _process_one_view_subprocess(view, no_clipboard=disable_clipboard_globally)
            if rc != 0:
                print(f"  -> subprocess exited with code {rc} (likely a native crash)", flush=True)
                # On Windows, 3221225477 = 0xC0000005 = ACCESS_VIOLATION, almost
                # always from clipboard/UI automation on this user's PC.
                # Retry this view in scroll-only mode, and use scroll-only for
                # all remaining views so we don't crash 5 more times.
                if not disable_clipboard_globally:
                    print(f"  -> retrying {view} in scroll-only mode (no clipboard)...", flush=True)
                    disable_clipboard_globally = True
                    rc2 = _process_one_view_subprocess(view, no_clipboard=True)
                    if rc2 != 0:
                        print(f"  -> retry also failed (code {rc2})", flush=True)
    except KeyboardInterrupt:
        print("\n!!! Ctrl+C — finishing up and saving what we have...", flush=True)
        interrupted = True

    if interrupted:
        # Clean up the stop signal so the next run starts fresh.
        try:
            (HERE / "STOP.txt").unlink()
        except Exception:
            pass

    _combine_view_results()

    # Same reasoning as in main()'s --view branch: skip Py_Finalize so we
    # don't risk a crash during interpreter shutdown after everything has
    # already been saved. The "Press any key to continue . . ." that comes
    # from the .bat file's `pause` still runs because the batch script sees
    # exit code 0 and continues.
    sys.stdout.flush()
    sys.stderr.flush()
    os._exit(0)


def main():
    print(f"==== Delivery Master Daily Check — {SCRIPT_VERSION} ====")
    print(f"Script file: {__file__}")
    args = sys.argv[1:]
    if "--view" in args:
        idx = args.index("--view")
        if idx + 1 >= len(args):
            print("Usage: --view <view name>")
            sys.exit(1)
        try:
            _process_one_view_inline(args[idx + 1])
        finally:
            # CRITICAL: bypass Python's normal interpreter shutdown.
            #
            # Py_Finalize runs a full cyclic GC sweep, which on machines with
            # the bad pywinauto/comtypes combo will iterate over the reference
            # cycles of UIA wrappers we built during Phase 1, call __del__ on
            # them, and segfault on the stale COM Release. By the time we
            # reach this point the per-view JSON checkpoint is already on
            # disk, so the orchestrator has everything it needs. os._exit
            # tells the OS to reap the process immediately — no finalizers,
            # no GC sweep, no chance to crash. Exit code 0 so the orchestrator
            # treats this view as a clean run.
            sys.stdout.flush()
            sys.stderr.flush()
            os._exit(0)
    if "--combine" in args:
        return _combine_view_results()
    if "--legacy" in args:
        return _legacy_inline_main()
    return _orchestrator()


def _legacy_inline_main():
    """Old single-process main, kept around in case the new subprocess
    architecture has trouble. Run with --legacy to invoke."""
    dm = find_dm()
    if dm is None:
        print("Delivery Master not found. Open it and try again.")
        sys.exit(1)
    print(f"Connected: {dm.window_text()!r}")

    # Make sure the rules workbook exists; load custom rules
    ensure_rules_workbook(RULES_PATH)
    rules = load_rules(RULES_PATH)
    if rules:
        print(f"Loaded {len(rules)} custom rule(s) from {RULES_PATH.name}:")
        for r in rules:
            who = r["customer_contains"] or "(all customers)"
            print(f"  - {who}: Cust. Ref {r['rule_type']} {r['values']}")
    else:
        print(f"No custom rules loaded (check {RULES_PATH.name}).")

    all_flagged = []
    all_accepted = []
    summary = []
    saved_to = None
    xlsx_path = HERE / "dm_daily_review.xlsx"
    json_backup_path = HERE / "dm_daily_checkpoint.json"

    def _row_sort_key(r):
        try:
            view_idx = VIEWS.index(r["view"])
        except ValueError:
            view_idx = len(VIEWS)
        dt = parse_dm_date(r["del_date"]) or datetime.max
        return (view_idx, dt)

    def _checkpoint_json():
        """Write a JSON snapshot of all collected data. Pure Python, can't crash
        natively — guaranteed safety net."""
        import json
        try:
            with open(json_backup_path, "w", encoding="utf-8") as f:
                json.dump({
                    "flagged": all_flagged,
                    "accepted": all_accepted,
                    "summary": summary,
                }, f, default=str, indent=2)
            return True
        except Exception as e:
            print(f"    !!! JSON checkpoint failed: {e}", flush=True)
            return False

    def _checkpoint_csv():
        """Write CSV files as the most reliable backup."""
        import csv
        try:
            flagged_csv = HERE / "dm_flagged.csv"
            with open(flagged_csv, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Include?", "View", "Our Ref", "Customer", "Status",
                            "Del Date", "Cust. Ref", "Reasons"])
                for r in all_flagged:
                    w.writerow(["Yes", r.get("view", ""), r.get("our_ref", ""),
                                r.get("customer", ""), r.get("status", ""),
                                r.get("del_date", ""), r.get("cust_ref", ""),
                                r.get("reasons", "")])
            accepted_csv = HERE / "dm_accepted.csv"
            with open(accepted_csv, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Promote?", "View", "Our Ref", "Customer", "Status",
                            "Del Date", "Cust. Ref", "Promote Reason"])
                for r in all_accepted:
                    w.writerow(["No", r.get("view", ""), r.get("our_ref", ""),
                                r.get("customer", ""), r.get("status", ""),
                                r.get("del_date", ""), r.get("cust_ref", ""), ""])
            return True
        except Exception as e:
            print(f"    !!! CSV checkpoint failed: {e}", flush=True)
            return False

    def _checkpoint_xlsx():
        """Write the review workbook with whatever data we have so far."""
        all_flagged.sort(key=_row_sort_key)
        all_accepted.sort(key=_row_sort_key)
        try:
            return write_review_xlsx(all_flagged, all_accepted, xlsx_path)
        except Exception as e:
            print(f"    !!! workbook save failed: {e}", flush=True)
            return None

    try:
        for view in VIEWS:
            print(f"\n--- View: {view} ---")
            try:
                # Re-find DM each iteration in case any prior native call left
                # our reference in a bad state.
                dm = find_dm()
                if dm is None:
                    print(f"    Couldn't re-find DM window")
                    summary.append((view, 0, 0, 0, "DM not found"))
                    continue
                if not switch_view(dm, view):
                    print(f"    Skipping {view}")
                    summary.append((view, 0, 0, 0, "switch failed"))
                    continue

                grid, panel = find_grid_and_panel(dm)
                if grid is None or panel is None:
                    print(f"    Grid/panel not found")
                    summary.append((view, 0, 0, 0, "grid not found"))
                    continue

                # PHASE 1 — identify columns from the visible rows (instant).
                # If the grid is mid-refresh (no rows visible), wait and retry once.
                sample_rows = read_visible_rows(panel)
                if not sample_rows:
                    print("    Phase 1 sample empty (view may still be refreshing); waiting 4s...")
                    time.sleep(4)
                    grid, panel = find_grid_and_panel(dm)
                    if panel is not None:
                        sample_rows = read_visible_rows(panel)
                uia_seed = detect_columns_via_uia(panel) if panel is not None else {}
                cols, diag = detect_columns(sample_rows, uia_seed=uia_seed)
                col_summary = ", ".join(f"{k}={v}" for k, v in sorted(cols.items(), key=lambda x: x[1]))
                print(f"    Columns ({len(sample_rows)}-row UIA sample): {col_summary}")
                if "date_match_counts" in diag:
                    counts = diag["date_match_counts"]
                    shown = ", ".join(f"col{i}={c}" for i, c in counts.items() if c > 0)
                    print(f"    Date matches per column: {shown}")

                # PHASE 2 — bulk-read via Ctrl+A + Ctrl+C; fall back to scroll if needed.
                rows = None
                try:
                    rows = read_all_rows_via_clipboard(dm, grid, panel)
                except Exception as e:
                    print(f"    Clipboard read errored ({e}); falling back to scroll.")
                if not rows:
                    print("    Clipboard read produced nothing; falling back to scroll.")
                    rows = read_all_rows_paged(dm, grid, panel)
                print(f"    Total unique rows: {len(rows)}")
                if not rows:
                    summary.append((view, 0, 0, 0, "no rows"))
                    continue

                # If Phase 1 detected nothing (e.g., empty grid during refresh),
                # fall back to detecting from the full Phase 2 data.
                if not cols:
                    print("    Phase 1 found no columns — re-detecting from full data...", flush=True)
                    cols, diag = detect_columns(rows)
                    col_summary = ", ".join(f"{k}={v}" for k, v in sorted(cols.items(), key=lambda x: x[1]))
                    print(f"    Columns (re-detected from {len(rows)} rows): {col_summary}", flush=True)

                flagged, accepted = categorize_rows(rows, cols, view, rules)
                print(f"    Flagged: {len(flagged)}   Accepted (clean past-date): {len(accepted)}", flush=True)
                all_flagged.extend(flagged)
                all_accepted.extend(accepted)
                summary.append((view, len(rows), len(flagged), len(accepted), col_summary))

                # Drop UIA references and force COM cleanup BEFORE the heavier
                # openpyxl write — those stale COM deallocators were running
                # mid-write and crashing Python (the "no VTable" errors we saw).
                grid = None
                panel = None
                rows = None
                sample_rows = None
                cols_local = cols  # keep cols (already a plain dict)
                uia_seed = None
                gc.collect()
                time.sleep(0.3)
                cols = cols_local

                # CHECKPOINT 1 — JSON (pure Python, guaranteed safety net).
                if _checkpoint_json():
                    print(f"    json checkpoint: {json_backup_path.name}", flush=True)

                # CHECKPOINT 2 — CSV (also pure Python; opens in Excel, no fancy bits).
                if _checkpoint_csv():
                    print(f"    csv checkpoints: dm_flagged.csv / dm_accepted.csv", flush=True)

                # CHECKPOINT 3 — xlsx (the real review workbook, openpyxl).
                print(f"    [debug] writing xlsx...", flush=True)
                ckpt = _checkpoint_xlsx()
                if ckpt:
                    saved_to = ckpt
                    print(f"    xlsx checkpoint: {ckpt.name}", flush=True)
                else:
                    print(f"    xlsx checkpoint failed (CSV/JSON above are usable)", flush=True)

                # Final GC pass before next view, so any stale COM objects
                # accumulated this iteration are released here, not at a
                # bad moment during the next view's UIA work.
                gc.collect()
                time.sleep(0.3)
            except Exception as view_err:
                import traceback
                print(f"    !!! ERROR processing view {view}: {view_err}")
                traceback.print_exc()
                summary.append((view, 0, 0, 0, f"exception: {type(view_err).__name__}"))
                continue
    finally:
        # Final save (in case the for-loop body never completed a view).
        ckpt = _checkpoint_xlsx()
        if ckpt:
            saved_to = ckpt

        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)
        for v, n, f, a, info in summary:
            print(f"  {v:>12}: {n:>3} rows, {f:>3} flagged, {a:>3} accepted   ({info})")
        print(f"\nTotal: {len(all_flagged)} flagged, {len(all_accepted)} accepted")
        if saved_to:
            print(f"\nReview workbook: {saved_to}")
    print(f"\nNEXT STEPS:")
    print(f"  1. Open the review workbook in Excel.")
    print(f"  2. On the Flagged tab, change Include? to 'No' on any false positives.")
    print(f"  3. On the Accepted tab, change Promote? to 'Yes' on rows that should be emailed.")
    print(f"  4. Save and close Excel.")
    print(f"  5. Run 'Generate Email.bat' to produce the email.")


if __name__ == "__main__":
    main()
