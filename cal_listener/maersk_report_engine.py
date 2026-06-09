"""
Maersk Report engine.

Cross-references our Consignment Log against the customer's weekly
Proforma sheet from 'Courier and Logistics Limited TA Cal GBP SR ...'.

Join key:    our Reference (col H)        <->  their Transporeon ID (col A)
Compared:    our Revenue   (col AC)       <->  their Total Cost      (col U)

Status taxonomy (one row per unique reference):
  MATCH             — both sides exist, values agree to ±£0.005
  VALUE DIFFERENCE  — both sides exist, values disagree
  ONLY ON LOG       — reference only on our side
  ONLY ON CUSTOMER  — reference only on their side

Duplicates: a reference appearing more than once on a side is silently
summed (per Owen's spec — duplicates aren't a concern in themselves).
Only if the summed values still don't match do we surface the
duplication, in the Notes column, as a hint at why the row diverged.
"""
from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


# --- Data containers --------------------------------------------------------

@dataclass
class LogRow:
    reference: str
    revenue: Optional[float]
    job_no: str = ""
    date: object = None
    customer_name: str = ""
    raw_row_num: int = 0


@dataclass
class CustomerRow:
    reference: str
    total_cost: Optional[float]
    transport_no: str = ""
    completed: object = None
    origin: str = ""
    destination: str = ""
    raw_row_num: int = 0


@dataclass
class MatchedRef:
    reference: str
    log_rows: list[LogRow] = field(default_factory=list)
    cust_rows: list[CustomerRow] = field(default_factory=list)
    status: str = ""
    log_value: Optional[float] = None
    cust_value: Optional[float] = None
    diff: Optional[float] = None
    note: str = ""


# --- Helpers ---------------------------------------------------------------

def _to_str_ref(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# --- Readers ----------------------------------------------------------------

def read_consignment_log(path: str | Path) -> list[LogRow]:
    """Read our Consignment Log. Joins by Reference (col H).
    Tolerates schema drift: looks up columns by header name."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]

    def col(name: str) -> int:
        for i, h in enumerate(header):
            if h.lower() == name.lower():
                return i
        return -1

    i_ref = col("Reference")
    i_rev = col("Revenue")
    i_job = col("Job No.")
    i_date = col("Date")
    i_cust = col("Customer Name")
    if i_ref < 0 or i_rev < 0:
        raise ValueError(
            "Consignment Log is missing the Reference or Revenue column. "
            f"Found headers: {header[:10]}...")
    out: list[LogRow] = []
    for n, r in enumerate(rows[1:], start=2):
        ref = _to_str_ref(r[i_ref] if i_ref < len(r) else None)
        if not ref:
            continue
        out.append(LogRow(
            reference=ref,
            revenue=_to_float(r[i_rev] if i_rev < len(r) else None),
            job_no=_to_str_ref(r[i_job] if 0 <= i_job < len(r) else ""),
            date=r[i_date] if 0 <= i_date < len(r) else None,
            customer_name=_to_str_ref(
                r[i_cust] if 0 <= i_cust < len(r) else ""),
            raw_row_num=n,
        ))
    return out


def read_customer_file(path: str | Path) -> list[CustomerRow]:
    """Read the Proforma sheet from the customer's weekly file.
    Joins by Transporeon ID (col A)."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    if "Proforma" in wb.sheetnames:
        ws = wb["Proforma"]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]

    def col(name: str) -> int:
        for i, h in enumerate(header):
            if h.lower() == name.lower():
                return i
        return -1

    i_ref = col("Transporeon ID")
    i_cost = col("Total Cost")
    i_tno = col("Transport No")
    i_done = col("Completed")
    i_orig = col("Origin Company")
    i_dest = col("Destination Company")
    if i_ref < 0 or i_cost < 0:
        raise ValueError(
            "Customer file is missing the Transporeon ID or Total Cost "
            f"column. Found headers: {header[:10]}...")
    out: list[CustomerRow] = []
    for n, r in enumerate(rows[1:], start=2):
        ref = _to_str_ref(r[i_ref] if i_ref < len(r) else None)
        if not ref:
            continue
        out.append(CustomerRow(
            reference=ref,
            total_cost=_to_float(r[i_cost] if i_cost < len(r) else None),
            transport_no=_to_str_ref(
                r[i_tno] if 0 <= i_tno < len(r) else ""),
            completed=r[i_done] if 0 <= i_done < len(r) else None,
            origin=_to_str_ref(r[i_orig] if 0 <= i_orig < len(r) else ""),
            destination=_to_str_ref(
                r[i_dest] if 0 <= i_dest < len(r) else ""),
            raw_row_num=n,
        ))
    return out


# --- Cross-reference --------------------------------------------------------

STATUS_MATCH       = "MATCH"
STATUS_DIFF        = "VALUE DIFFERENCE"
STATUS_ONLY_LOG    = "ONLY ON LOG"
STATUS_ONLY_CUST   = "ONLY ON CUSTOMER"

ALL_STATUSES = (STATUS_MATCH, STATUS_DIFF, STATUS_ONLY_LOG, STATUS_ONLY_CUST)


def cross_reference(log_rows: list[LogRow],
                     cust_rows: list[CustomerRow],
                     tolerance: float = 0.005) -> list[MatchedRef]:
    """Build one MatchedRef per unique reference seen on either side.
    Duplicates on the same side are summed silently. Per Owen: only
    surface duplication in the Notes column when the summed values
    still disagree — the dup may explain the mismatch."""
    log_by: dict[str, list[LogRow]] = defaultdict(list)
    for r in log_rows:
        log_by[r.reference].append(r)
    cust_by: dict[str, list[CustomerRow]] = defaultdict(list)
    for r in cust_rows:
        cust_by[r.reference].append(r)

    all_refs = sorted(set(log_by) | set(cust_by))
    out: list[MatchedRef] = []
    for ref in all_refs:
        lrs = log_by.get(ref, [])
        crs = cust_by.get(ref, [])
        log_val = sum((lr.revenue or 0) for lr in lrs) if lrs else None
        cust_val = sum((cr.total_cost or 0) for cr in crs) if crs else None

        if lrs and not crs:
            status = STATUS_ONLY_LOG
        elif crs and not lrs:
            status = STATUS_ONLY_CUST
        elif (log_val is None) or (cust_val is None):
            status = STATUS_DIFF
        elif abs(log_val - cust_val) <= tolerance:
            status = STATUS_MATCH
        else:
            status = STATUS_DIFF

        diff = None
        if log_val is not None and cust_val is not None:
            diff = log_val - cust_val

        # Note: only flag duplication if the values disagree.
        note = ""
        if status == STATUS_DIFF:
            if len(lrs) > 1 and len(crs) > 1:
                note = (f"Duplicated on both sides "
                        f"(log×{len(lrs)}, customer×{len(crs)}) — "
                        f"may explain the difference.")
            elif len(lrs) > 1:
                note = (f"Duplicated on our log (×{len(lrs)}) — "
                        f"may explain the difference.")
            elif len(crs) > 1:
                note = (f"Duplicated on customer report (×{len(crs)}) — "
                        f"may explain the difference.")

        out.append(MatchedRef(
            reference=ref, log_rows=lrs, cust_rows=crs,
            status=status, log_value=log_val,
            cust_value=cust_val, diff=diff, note=note,
        ))
    return out


# --- Output -----------------------------------------------------------------

def _money(v) -> str:
    if v is None or v == "":
        return ""
    try:
        return f"£{float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def _signed_money(v) -> str:
    if v is None or v == "":
        return ""
    try:
        return f"{float(v):+,.2f}"
    except (TypeError, ValueError):
        return str(v)


OUTPUT_HEADERS = [
    "Reference",
    "Status",
    "Log Revenue (£)",
    "Customer Total Cost (£)",
    "Difference (Log − Customer, £)",
    "Notes",
    "Log Job No(s)",
    "Customer Transport No(s)",
    "Customer (from log)",
    "Origin (from customer)",
    "Destination (from customer)",
]


def to_output_rows(matched: list[MatchedRef]) -> list[list]:
    """Build the list of [value, ...] rows in OUTPUT_HEADERS order.
    Used by both the CSV writer and the in-UI preview table."""
    out_rows: list[list] = []
    for m in matched:
        log_jobs = " / ".join(lr.job_no for lr in m.log_rows if lr.job_no)
        cust_tnos = " / ".join(
            cr.transport_no for cr in m.cust_rows if cr.transport_no)
        cust_name = next(
            (lr.customer_name for lr in m.log_rows if lr.customer_name), "")
        origin = next(
            (cr.origin for cr in m.cust_rows if cr.origin), "")
        destination = next(
            (cr.destination for cr in m.cust_rows if cr.destination), "")
        out_rows.append([
            m.reference,
            m.status,
            _money(m.log_value),
            _money(m.cust_value),
            _signed_money(m.diff),
            m.note,
            log_jobs,
            cust_tnos,
            cust_name,
            origin,
            destination,
        ])
    return out_rows


def write_csv(matched: list[MatchedRef], output_path: str | Path) -> None:
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(OUTPUT_HEADERS)
        for row in to_output_rows(matched):
            w.writerow(row)


def summarise(matched: list[MatchedRef]) -> dict:
    from collections import Counter
    stats = Counter(m.status for m in matched)
    n_dup_diff = sum(1 for m in matched if m.note)
    log_total = sum((m.log_value or 0) for m in matched if m.log_value)
    cust_total = sum((m.cust_value or 0) for m in matched if m.cust_value)
    return {
        "by_status":          dict(stats),
        "total_refs":         len(matched),
        "dups_implicated":    n_dup_diff,
        "log_value_total":    log_total,
        "cust_value_total":   cust_total,
        "net_difference":     log_total - cust_total,
    }


def generate(log_path: str | Path, cust_path: str | Path,
              output_path: str | Path) -> dict:
    """End-to-end: read both files, cross-ref, write CSV, return stats."""
    log_rows = read_consignment_log(log_path)
    cust_rows = read_customer_file(cust_path)
    matched = cross_reference(log_rows, cust_rows)
    write_csv(matched, output_path)
    summary = summarise(matched)
    summary["log_rows_read"] = len(log_rows)
    summary["cust_rows_read"] = len(cust_rows)
    summary["output_path"] = str(output_path)
    summary["matched"] = matched  # caller may want this for the UI
    return summary
