"""
Consignment Cross-Ref engine — log + PDF parsing + matching.

Two parsers for the FedEx LPO PDF:

  1. PRIMARY: poppler-utils' `pdftotext -layout`. Owen's reference
     implementation uses this and it parses both real-world PDFs to
     the penny against the stated LPO Value. We try this first and
     prefer it whenever it's on PATH.

  2. FALLBACK: pdfplumber. Pure Python so it works without an
     external binary. Has one known weakness: when the Order
     Reference text wraps right up against the value column the
     two words can fuse into a single token like
     "SR4-EH£12235.00". The fallback detects this pattern and
     disambiguates by treating the right-most "£NNN.NN" (up to 4
     digits before the decimal) as the value and everything else
     as part of the order ref.

Matching logic and the "MATCH / VALUE DIFFERENCE / ONLY ON LOG /
ONLY ON PDF" taxonomy are spec-defined - see the module docstring
of report_builder.py for column meanings.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Add the Cal Toolkit root to sys.path so vendor_setup imports cleanly
# whether the plugin is loaded via the launcher (which already adds
# the root) or via a standalone script during testing.
_ROOT_FOR_VENDOR = Path(__file__).resolve().parent.parent.parent
if str(_ROOT_FOR_VENDOR) not in sys.path:
    sys.path.insert(0, str(_ROOT_FOR_VENDOR))

try:
    import vendor_setup  # type: ignore
except Exception:
    vendor_setup = None  # type: ignore


# ---------------------------------------------------------------------------
# Consignment log
# ---------------------------------------------------------------------------

# 1-indexed column positions in the consignment log (matches Owen's spec).
LOG_COLS = {
    "job_no":       2,    # B
    "date":         3,    # C
    "customer":     5,    # E
    "reference":    8,    # H  - the 8-digit join key
    "pu_company":   10,   # J
    "collect_from": 11,   # K
    "deliver_to":   17,   # Q
    "revenue":      29,   # AC - the value that matches the PDF
    "cust_vat":     30,   # AD
    "total":        31,   # AE - inc VAT
    # NB: there's a second 'Total' in col 34 (AH); that's the driver
    # side total, NOT what we compare against the PDF. Spec calls
    # it out explicitly; we don't read it.
}


@dataclass
class LogRow:
    job_no:       object
    date:         object
    customer:     object
    reference:    str
    pu_company:   object
    collect_from: object
    deliver_to:   object
    revenue:      object
    cust_vat:     object
    total:        object


def read_log(xlsx_path: str | Path) -> list[LogRow]:
    """Parse the consignment-log xlsx. Skips rows with a blank
    Reference (per spec). References coerced to stripped strings."""
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    try:
        # Spec calls Sheet1 but tolerate any name if there's only one.
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
        else:
            ws = wb.active
        out: list[LogRow] = []
        max_col = max(LOG_COLS.values())
        for r in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
            # iter_rows returns a tuple; index is 0-based, but our
            # LOG_COLS are 1-based — subtract 1.
            ref = r[LOG_COLS["reference"] - 1]
            if ref is None:
                continue
            ref_s = str(ref).strip()
            if not ref_s:
                continue
            out.append(LogRow(
                job_no=r[LOG_COLS["job_no"] - 1],
                date=r[LOG_COLS["date"] - 1],
                customer=r[LOG_COLS["customer"] - 1],
                reference=ref_s,
                pu_company=r[LOG_COLS["pu_company"] - 1],
                collect_from=r[LOG_COLS["collect_from"] - 1],
                deliver_to=r[LOG_COLS["deliver_to"] - 1],
                revenue=r[LOG_COLS["revenue"] - 1],
                cust_vat=r[LOG_COLS["cust_vat"] - 1],
                total=r[LOG_COLS["total"] - 1],
            ))
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# PDF parsing
# ---------------------------------------------------------------------------

@dataclass
class PdfRow:
    source_pdf: str
    depot:      str
    order:      str
    svc:        str
    date:       str
    account:    str
    customer:   str
    order_ref:  str
    value:      Optional[float]


# Module-level cache: the install handler in run.py can pre-set this
# to skip discovery entirely when it just installed/found a binary.
_PDFTOTEXT_OVERRIDE: Optional[str] = None


def set_pdftotext_path(path: Optional[str]) -> None:
    """Pin a known-good pdftotext.exe path so subsequent calls to
    have_pdftotext() / find_pdftotext_path() return it without
    re-scanning the filesystem. Pass None to clear the override."""
    global _PDFTOTEXT_OVERRIDE
    if path and Path(path).exists():
        _PDFTOTEXT_OVERRIDE = path
        _debug_log(f"override set: {path}")
    else:
        _PDFTOTEXT_OVERRIDE = None


def have_pdftotext() -> bool:
    """Return True if a pdftotext.exe is reachable — either bundled
    under CalToolkit/vendor/poppler/, somewhere on PATH, or in a
    well-known install location. Returns False only if we'd need to
    fall back to pdfplumber."""
    return find_pdftotext_path() is not None


def _debug_log(msg: str) -> None:
    """Append a line to data/pdftotext_lookup.log — best-effort, never raises.
    Lets us diagnose why have_pdftotext() returns False on a machine
    where the binary clearly exists."""
    try:
        root = Path(__file__).resolve().parent.parent.parent
        log_dir = root / "data"
        log_dir.mkdir(parents=True, exist_ok=True)
        with open(log_dir / "pdftotext_lookup.log", "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except Exception:
        pass


def _candidate_homes() -> list[Path]:
    """All plausible "user home" roots — Path.home() can return a
    OneDrive-redirected folder on some Windows setups, so we cast a
    wide net using every env var Windows might use."""
    seen: set[str] = set()
    out: list[Path] = []

    def add(p) -> None:
        if not p:
            return
        try:
            s = str(Path(p))
        except Exception:
            return
        if s in seen:
            return
        seen.add(s)
        out.append(Path(s))

    try:
        add(Path.home())
    except Exception:
        pass
    add(os.environ.get("USERPROFILE"))
    hd, hp = os.environ.get("HOMEDRIVE"), os.environ.get("HOMEPATH")
    if hd and hp:
        add(hd + hp)
    uname = os.environ.get("USERNAME")
    if uname:
        add(f"C:/Users/{uname}")
    return out


def find_pdftotext_path() -> Optional[str]:
    """Locate pdftotext.exe — defensive against:

    * Stale .pyc caches (we duplicate the search inline instead of
      relying solely on vendor_setup).
    * Path.home() returning a OneDrive folder (we use multiple
      home-detection methods).
    * PATH not refreshed in a running GUI process (we never rely on
      shutil.which alone).
    """
    # 0. Module-level override (set by run.py when install succeeds).
    if _PDFTOTEXT_OVERRIDE and Path(_PDFTOTEXT_OVERRIDE).exists():
        return _PDFTOTEXT_OVERRIDE

    # 1. Delegate to vendor_setup if importable — keeps a single
    # source of truth when the cache is fresh.
    if vendor_setup is not None:
        try:
            p = vendor_setup.find_pdftotext()
            if p and Path(p).exists():
                _debug_log(f"via vendor_setup: {p}")
                return p
        except Exception as e:
            _debug_log(f"vendor_setup raised: {e!r}")

    # 2. Inline search across vendor folder + every plausible home +
    # well-known absolute roots. Each home is tried via multiple
    # detection methods so OneDrive redirection can't hide the binary.
    root = Path(__file__).resolve().parent.parent.parent
    bases: list[Path] = [root / "vendor" / "poppler"]
    for h in _candidate_homes():
        bases.append(h / "poppler-windows")
        bases.append(h / "poppler")
    bases.extend([
        Path("C:/poppler"),
        Path("C:/poppler-windows"),
        Path("C:/Program Files/poppler"),
        Path("C:/Program Files/poppler-windows"),
    ])
    _debug_log(f"checking bases: {[str(b) for b in bases]}")
    for c in bases:
        try:
            if not c.exists():
                continue
            for p in c.rglob("pdftotext.exe"):
                _debug_log(f"hit: {p}")
                return str(p)
        except Exception as e:
            _debug_log(f"rglob {c} raised: {e!r}")

    # 3. PATH (least likely to work for a fresh install since
    # env updates from PowerShell don't reach a running GUI app).
    on_path = shutil.which("pdftotext")
    if on_path:
        _debug_log(f"via PATH: {on_path}")
        return on_path

    _debug_log("nothing found")
    return None


def parse_pdf(pdf_path: str | Path,
              label: str) -> tuple[list[PdfRow], float | None]:
    """Parse one FedEx LPO PDF. Returns (rows, stated_lpo_total).

    Tries pdftotext first (faithful to Owen's reference impl); falls
    back to pdfplumber if poppler-utils isn't installed."""
    if have_pdftotext():
        return _parse_via_pdftotext(pdf_path, label)
    return _parse_via_pdfplumber(pdf_path, label)


# ---------- pdftotext (primary) ----------

# Header data row: depot, 8-digit order, service code, dd/mm/yyyy date,
# account (typically 10 digits but tolerate any non-whitespace token),
# then "rest" (customer + order ref + value).
_ROW_RE = re.compile(
    r"^\s*([A-Z0-9]{2,4})\s+(\d{8})\s+(\S+)\s+"
    r"(\d{2}/\d{2}/\d{4})\s+(\S+)\s+(.+)$"
)
_VALUE_TAIL_RE = re.compile(r"£([\d,]+\.\d{2})\s*$")
_BARE_VALUE_RE = re.compile(r"^\s*£([\d,]+\.\d{2})\s*$")


def _parse_via_pdftotext(pdf_path: str | Path,
                          label: str) -> tuple[list[PdfRow], float | None]:
    """Shell out to pdftotext -layout. Uses the absolute path returned
    by find_pdftotext_path() so the bundled vendor binary works even
    when nothing is on the system PATH.

    Encoding nuance: pdftotext emits UTF-8 by default. On Windows the
    Python default text-mode encoding is cp1252, which mis-decodes
    the multi-byte £ (\xc2\xa3) into "Â£" — breaking our '£value'
    regexes for any line whose £ wasn't preceded by whitespace.
    Explicitly setting encoding='utf-8' fixes this and is harmless on
    macOS / Linux where it's already the default.
    """
    exe = find_pdftotext_path() or "pdftotext"
    txt = subprocess.run(
        [exe, "-layout", str(pdf_path), "-"],
        capture_output=True, text=True, check=True,
        encoding="utf-8", errors="replace",
    ).stdout
    return _parse_text_lines(txt.splitlines(), label)


def _parse_text_lines(lines, label: str) -> tuple[list[PdfRow], float | None]:
    rows: list[PdfRow] = []
    lpo_total: float | None = None
    for raw in lines:
        if not raw.strip():
            continue
        if "FedEx Classification" in raw:
            continue
        if "Order Reference" in raw and "Order Value" in raw:
            continue
        if "LPO Value" in raw:
            v = re.search(r"£([\d,]+\.\d{2})", raw)
            if v:
                lpo_total = float(v.group(1).replace(",", ""))
            continue
        m = _ROW_RE.match(raw)
        if m:
            depot, order, svc, date, account, rest = m.groups()
            value, ref_part = _extract_value(rest)
            customer, order_ref = _split_customer_and_ref(ref_part)
            rows.append(PdfRow(
                source_pdf=label, depot=depot, order=order, svc=svc,
                date=date, account=account, customer=customer,
                order_ref=order_ref, value=value,
            ))
        else:
            # Continuation line. Three things can happen here:
            #   (a) line is just "£value" alone -> attach to last row
            #   (b) line is wrapped customer/order-ref text WITH the
            #       value at the end ("  OVERFLOW TEXT  £123.45") ->
            #       still attach, and append the leading text as
            #       order-ref overflow so the row is complete
            #   (c) line is wrapped text WITHOUT a value -> append
            #       to order-ref overflow only
            # The original code only handled (a) via a strict
            # whole-line match, which is why some Wk50/Wk51 rows
            # ended up with value=None and showed as VALUE DIFFERENCE
            # without a value on screen.
            if not rows:
                continue
            last = rows[-1]
            v = _VALUE_TAIL_RE.search(raw)
            if v and last.value is None:
                last.value = float(v.group(1).replace(",", ""))
                head = raw[:v.start()].strip()
                if head:
                    last.order_ref = (
                        f"{last.order_ref} {head}".strip()
                        if last.order_ref else head)
            elif not v:
                # Pure text continuation — append to order_ref so the
                # full reference is preserved. Skip if the line looks
                # like a footer/page-break or other noise.
                text = raw.strip()
                if text and not text.startswith("Page "):
                    last.order_ref = (
                        f"{last.order_ref} {text}".strip()
                        if last.order_ref else text)
    return rows, lpo_total


def _extract_value(rest: str) -> tuple[float | None, str]:
    """Pull the trailing £value off the row's "rest" text. Returns
    (value or None, remaining_text). Handles the FUJI-style glued
    case "REF£140.00" by allowing the £ to be flush against text."""
    m = _VALUE_TAIL_RE.search(rest)
    if m:
        value = float(m.group(1).replace(",", ""))
        return value, rest[:m.start()].rstrip()
    return None, rest.rstrip()


def _split_customer_and_ref(text: str) -> tuple[str, str]:
    """Split the customer name (left) from the order reference (right)
    on 2-or-more spaces. The FedEx PDF leaves the order-ref column
    blank for most rows, so the typical result is (customer, '')."""
    parts = re.split(r"\s{2,}", text)
    customer = parts[0].strip()
    order_ref = " ".join(p.strip() for p in parts[1:]) if len(parts) > 1 else ""
    return customer, order_ref


# ---------- pdfplumber (fallback) ----------

# Disambiguate fused "SR4-EH£12235.00" → order_ref ends "SR4-EH12",
# value is "£235.00". Heuristic: the value is at most 4 digits before
# the decimal. Anything beyond that means the digits belong to the
# order ref.
_GLUED_RE = re.compile(
    r"^(?P<head>.*?)£(?P<digits>\d{1,4})(?P<value_tail>\d{0,4}\.\d{2})$"
)


def _parse_via_pdfplumber(pdf_path: str | Path,
                            label: str) -> tuple[list[PdfRow], float | None]:
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError(
            "Neither pdftotext (poppler-utils) nor pdfplumber is "
            "available. Install poppler-utils (recommended) or run:\n"
            "  pip install pdfplumber") from e
    rows: list[PdfRow] = []
    lpo_total: float | None = None
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(layout=True) or ""
            for raw in txt.splitlines():
                if not raw.strip():
                    continue
                if "FedEx Classification" in raw:
                    continue
                if "Order Reference" in raw and "Order Value" in raw:
                    continue
                if "LPO Value" in raw:
                    v = re.search(r"£([\d,]+\.\d{2})", raw)
                    if v:
                        lpo_total = float(v.group(1).replace(",", ""))
                    continue
                m = _ROW_RE.match(raw)
                if m:
                    depot, order, svc, date, account, rest = m.groups()
                    value, ref_part = _extract_value_with_glue_fix(rest)
                    customer, order_ref = _split_customer_and_ref(ref_part)
                    rows.append(PdfRow(
                        source_pdf=label, depot=depot, order=order, svc=svc,
                        date=date, account=account, customer=customer,
                        order_ref=order_ref, value=value,
                    ))
                else:
                    v = _BARE_VALUE_RE.match(raw)
                    if v and rows and rows[-1].value is None:
                        rows[-1].value = float(v.group(1).replace(",", ""))
    return rows, lpo_total


def _extract_value_with_glue_fix(rest: str) -> tuple[float | None, str]:
    """pdfplumber sometimes fuses the order-ref text and the £value
    into a single token. Recover from that by detecting >4 digits
    immediately after £ and re-splitting."""
    # Fast path: clean trailing value.
    m = _VALUE_TAIL_RE.search(rest)
    if m:
        # Was the £ preceded by text without whitespace? If so, this
        # is the glued case.
        before = rest[:m.start()]
        if before and not before[-1].isspace():
            # We have "<text>£<digits>" right at the end. Split.
            glue_m = _GLUED_RE.match(rest)
            if glue_m:
                head, digits, value_tail = (
                    glue_m["head"], glue_m["digits"], glue_m["value_tail"])
                # Cap value at 4 digits before the decimal; rest is
                # part of the order ref.
                value = float((digits + value_tail))
                if value > 9999.99:
                    # Take last 6 chars (NNN.NN) as value, rest as ref
                    suffix = digits + value_tail
                    cut = max(0, len(suffix) - 6)
                    ref_extra = suffix[:cut]
                    value_str = suffix[cut:]
                    value = float(value_str)
                    ref_part = (head + ref_extra).rstrip()
                    return value, ref_part
        value = float(m.group(1).replace(",", ""))
        return value, rest[:m.start()].rstrip()
    return None, rest.rstrip()


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

@dataclass
class VerifyResult:
    ok:              bool
    label:           str
    n_rows:          int
    line_sum:        float
    lpo_total:       float | None
    missing_value:   list[str] = field(default_factory=list)
    message:         str = ""


def verify_pdf_parse(rows: list[PdfRow],
                      lpo_total: float | None,
                      label: str,
                      tolerance: float = 0.005) -> VerifyResult:
    """Per-PDF invariant: line items sum to the stated LPO Value."""
    missing = [r.order for r in rows if r.value is None]
    line_sum = sum((r.value or 0) for r in rows)
    if missing:
        return VerifyResult(
            ok=False, label=label, n_rows=len(rows),
            line_sum=line_sum, lpo_total=lpo_total,
            missing_value=missing,
            message=(f"{label}: {len(missing)} row(s) have no value: "
                     f"{missing[:5]}{'…' if len(missing) > 5 else ''}"))
    if lpo_total is None:
        return VerifyResult(
            ok=False, label=label, n_rows=len(rows),
            line_sum=line_sum, lpo_total=None,
            message=f"{label}: could not find an LPO Value line in the PDF.")
    diff = line_sum - lpo_total
    if abs(diff) > tolerance:
        return VerifyResult(
            ok=False, label=label, n_rows=len(rows),
            line_sum=line_sum, lpo_total=lpo_total,
            message=(f"{label}: line items sum to £{line_sum:.2f} but PDF "
                     f"states LPO Value £{lpo_total:.2f} "
                     f"(diff £{diff:+.2f})."))
    return VerifyResult(
        ok=True, label=label, n_rows=len(rows),
        line_sum=line_sum, lpo_total=lpo_total,
        message=(f"{label}: {len(rows)} rows · sum £{line_sum:.2f} "
                 f"matches LPO £{lpo_total:.2f}"))


def find_duplicate_orders(all_pdf_rows: list[PdfRow]) -> dict[str, list[str]]:
    """Cross-PDF: no order_number should appear in more than one PDF.
    Returns {order: [list of source_pdf labels]} for any duplicates."""
    seen: dict[str, list[str]] = {}
    for r in all_pdf_rows:
        seen.setdefault(r.order, []).append(r.source_pdf)
    return {k: v for k, v in seen.items() if len(set(v)) > 1}


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

MATCH = "MATCH"
VALUE_DIFFERENCE = "VALUE DIFFERENCE"
ONLY_ON_LOG = "ONLY ON LOG"
ONLY_ON_PDF = "ONLY ON PDF"


@dataclass
class MatchedRef:
    """One entry in the cross-reference table: every unique reference
    found across the log + every PDF, with its status and source row(s)."""
    reference:    str
    log_row:      LogRow | None
    pdf_row:      PdfRow | None
    status:       str
    log_revenue:  float | None
    pdf_value:    float | None
    diff:         float | None


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def cross_reference(log_rows: list[LogRow],
                     pdf_rows: list[PdfRow],
                     tolerance: float = 0.005) -> list[MatchedRef]:
    """Build the spec's MATCH / VALUE DIFFERENCE / ONLY ON LOG /
    ONLY ON PDF taxonomy for every reference seen across either side."""
    log_by_ref: dict[str, LogRow] = {r.reference: r for r in log_rows}
    pdf_by_ref: dict[str, PdfRow] = {r.order: r for r in pdf_rows}
    all_refs = sorted(set(log_by_ref) | set(pdf_by_ref))

    out: list[MatchedRef] = []
    for ref in all_refs:
        l = log_by_ref.get(ref)
        p = pdf_by_ref.get(ref)
        log_rev = _to_float(l.revenue) if l is not None else None
        pdf_val = p.value if p is not None else None
        diff: float | None = None
        if l is None and p is not None:
            status = ONLY_ON_PDF
        elif l is not None and p is None:
            status = ONLY_ON_LOG
        elif log_rev is None or pdf_val is None:
            # Both rows exist but one side has no value. Treat as
            # value-difference so the user sees it on the review tab.
            status = VALUE_DIFFERENCE
        else:
            diff = log_rev - pdf_val
            status = MATCH if abs(diff) <= tolerance else VALUE_DIFFERENCE
        out.append(MatchedRef(
            reference=ref, log_row=l, pdf_row=p, status=status,
            log_revenue=log_rev, pdf_value=pdf_val, diff=diff,
        ))
    return out
