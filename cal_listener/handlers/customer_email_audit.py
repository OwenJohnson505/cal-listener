"""Customer Email Audit — listener handler.

Walks DM's Customer grid, opens each customer dialog, scrapes
Trading Name / Invoice email / Notes, cross-checks against a
ClearBooks CSV (uploaded to listener_inputs) and produces a 3-sheet
xlsx (All / Mismatches / Summary) uploaded to listener_results.

params:
  cb_csv_storage_path   str, key in listener_inputs. Required.
  limit                 int, optional smoke-test cap (default: walk all)
"""
from __future__ import annotations

import io
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List

import requests

from .. import dm


def _setup_engine_imports(ctx):
    from cal_listener import customer_profile_store as _cps
    from cal_listener import cloud_sync as _cs
    sys.modules.setdefault("customer_profile_store", _cps)
    sys.modules.setdefault("cloud_sync", _cs)
    _cs.bind_supabase(ctx.sb, user_id=ctx.settings.listener_id)


def _download_input(ctx, storage_path: str, dest: Path) -> None:
    url = (f"{ctx.settings.supabase_url.rstrip('/')}"
           f"/storage/v1/object/listener_inputs/{storage_path}")
    h = {"apikey": ctx.settings.supabase_service_key,
         "Authorization": f"Bearer {ctx.settings.supabase_service_key}"}
    r = requests.get(url, headers=h, timeout=120)
    r.raise_for_status()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(r.content)


def _build_xlsx(records: List[dict], summary: dict) -> bytes:
    """3-sheet xlsx (All / Mismatches / Summary), same shape as desktop."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Records"
    headers = ["TMS Customer", "DM Invoice Email",
               "DM Additional Emails", "DM Notes Emails",
               "ClearBooks Name", "ClearBooks Email",
               "Match Status", "Detail"]
    ws_all.append(headers)
    for c in ws_all[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0EA5A4")
        c.alignment = Alignment(horizontal="left", vertical="center")
    for rec in records:
        ws_all.append([
            rec.get("trading", ""),
            rec.get("invoice", ""),
            rec.get("additional_emails", ""),
            ", ".join(rec.get("notes_emails", []) or []),
            rec.get("cb_name", ""),
            rec.get("cb_email", ""),
            rec.get("status", ""),
            rec.get("detail", ""),
        ])

    ws_mm = wb.create_sheet("Mismatches")
    ws_mm.append(headers)
    for c in ws_mm[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="DC2626")
    for rec in records:
        if rec.get("status") == "mismatched":
            ws_mm.append([
                rec.get("trading", ""), rec.get("invoice", ""),
                rec.get("additional_emails", ""),
                ", ".join(rec.get("notes_emails", []) or []),
                rec.get("cb_name", ""), rec.get("cb_email", ""),
                rec.get("status", ""), rec.get("detail", "")])

    ws_sum = wb.create_sheet("Summary")
    for k, v in (summary or {}).items():
        ws_sum.append([k, v])
    for c in ws_sum["A"]:
        c.font = Font(bold=True)

    for ws in (ws_all, ws_mm):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run(params: Dict[str, Any], on_progress: Callable[..., None],
        ctx) -> Dict[str, Any]:
    cb_path = params.get("cb_csv_storage_path")
    if not cb_path:
        return {"ok": False, "error": "cb_csv_storage_path required",
                "record_count": 0}

    on_progress("Ensuring DM is logged in", percent=5)
    app = dm.ensure_logged_in(ctx, on_progress=on_progress, timeout=120)
    on_progress("Navigating to Customers", percent=8)
    dm.navigate_to_page(app, "Customers", on_progress=on_progress)

    tmpdir = Path(tempfile.mkdtemp(prefix="email_audit_"))
    local_csv = tmpdir / Path(cb_path).name
    on_progress(f"Downloading {cb_path}", percent=12)
    _download_input(ctx, cb_path, local_csv)

    _setup_engine_imports(ctx)
    from cal_listener import customer_email_audit_engine as _engine

    limit = params.get("limit")
    if limit is not None:
        try:
            limit = int(limit)
        except Exception:
            limit = None

    records: List[dict] = []

    def _on_record(rec: dict) -> None:
        records.append(rec)

    def _on_engine_progress(done: int, total: int, msg: str) -> None:
        pct = 15 + int(75 * (done / max(total, 1)))
        on_progress(f"{msg} ({done}/{total})",
                    percent=min(pct, 90))

    def _stop_check() -> bool:
        # Cancellation propagates via on_progress raising JobCancelled
        # on the next call from the runner; engine doesn't poll this
        # often enough to be the primary path.
        return False

    on_progress("Running audit (walking every customer)", percent=15)
    summary = _engine.run_audit(
        csv_path=local_csv,
        on_record=_on_record,
        on_progress=_on_engine_progress,
        stop_check=_stop_check,
        logger=lambda m: on_progress(m, percent=None),
        limit=limit,
    )

    on_progress(
        f"Engine finished: {summary.get('processed')} processed, "
        f"{summary.get('mismatched')} mismatched", percent=92)

    on_progress("Building Excel output", percent=94)
    xlsx_bytes = _build_xlsx(records, summary)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    key = f"customer_email_audit/customer_email_audit_{stamp}.xlsx"
    on_progress(f"Uploading {key}", percent=97)
    ok = ctx.sb.storage_upload(
        "listener_results", key, xlsx_bytes,
        content_type=("application/vnd.openxmlformats-"
                      "officedocument.spreadsheetml.sheet"))
    public_url = (ctx.sb.storage_public_url("listener_results", key)
                  if ok else None)

    on_progress("Done", percent=100)
    return {
        "ok": True,
        "record_count": len(records),
        "result_url": public_url,
        "summary": summary or {},
    }
