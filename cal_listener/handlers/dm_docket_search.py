"""DM Docket Search — listener handler.

Opens DM's Docket Search dialog, applies the user's search filter,
iterates every result row, scrapes the wizard, and writes the lot
to an xlsx uploaded to Supabase Storage.

params:
  search    dict, forwarded verbatim to dm_driver.apply_search().
            Shape comes from the web's Docket Search inputs panel.
            See plugins/dm_docket_search/dm_driver.py:apply_search for
            the accepted keys (date_from, date_to, customer_filter,
            status flags, etc.).
  max_jobs  int, safety cap on number of bookings to scrape (None = all).
"""
from __future__ import annotations

import io
import sys
from datetime import datetime, timezone
from typing import Any, Callable, Dict, List

from .. import dm


def _build_xlsx(rows: List[dict]) -> bytes:
    """Replicates desktop listener_entry.py's column ordering."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Docket Search"
    cols: List[str] = []
    seen: set = set()
    preferred = ["our_ref", "status", "date_time", "customer",
                 "customer_ref", "tariff_name", "driver", "value_gbp",
                 "collect_company", "deliver_company", "invoice_no"]
    for k in preferred:
        if any(k in r for r in rows):
            cols.append(k)
            seen.add(k)
    for r in rows:
        for k in r.keys():
            if k not in seen:
                cols.append(k)
                seen.add(k)
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0EA5A4")
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run(params: Dict[str, Any], on_progress: Callable[..., None], ctx) -> Dict[str, Any]:
    search_payload = params.get("search") or {}
    max_jobs = params.get("max_jobs")
    if max_jobs is not None:
        try: max_jobs = int(max_jobs)
        except Exception: max_jobs = None

    on_progress("Ensuring DM is logged in", percent=5)
    dm.ensure_logged_in(ctx, on_progress=on_progress, timeout=120)

    # dm_driver.search_and_scrape() handles its own DM connection +
    # opens Docket Search itself, so the handler just needs DM running
    # and logged in.
    from cal_listener import dm_driver as _drv

    def _engine_progress(i: int, n: int, our_ref: str) -> None:
        if n == 0:
            on_progress(str(our_ref), percent=None)
        else:
            pct = 15 + int(80 * (i / max(n, 1)))
            on_progress(f"[{i}/{n}] {our_ref}", percent=min(pct, 95))

    on_progress("Running docket search", percent=10)
    rows = _drv.search_and_scrape(
        search_payload,
        on_progress=_engine_progress,
        should_cancel=lambda: False,
        max_jobs=max_jobs,
    )

    if not rows:
        on_progress("Search returned 0 results", level="info", percent=100)
        return {"ok": True, "record_count": 0, "result_url": None}

    on_progress(f"Building xlsx for {len(rows)} rows", percent=96)
    xlsx = _build_xlsx(rows)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    key = f"dm_docket_search/dm_docket_search_{stamp}.xlsx"
    ok = ctx.sb.storage_upload(
        "listener_results", key, xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    public_url = ctx.sb.storage_public_url("listener_results", key) if ok else None

    on_progress(f"Done — {len(rows)} rows", percent=100)
    return {"ok": True, "record_count": len(rows), "result_url": public_url}
