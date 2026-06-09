"""Revenue Breakdown Scraper — listener handler.

Walks DM's Customer Invoice grid, scrapes each booking's Revenue
Breakdown via the bundled `breakdown_driver.dry_run()`, uploads a
3-sheet xlsx to Supabase Storage and returns the public URL.

params:
  max_rows   int, how many bookings to walk (default 50).
"""
from __future__ import annotations

import io
import sys
from datetime import datetime, timezone
from typing import Any, Callable, Dict, List

from .. import dm


def _setup_engine_imports():
    """Make `import dm_driver` work inside the bundled engine. The
    listener bundles dm_driver as `cal_listener.dm_driver`; the engine
    expects it at top level."""
    from cal_listener import dm_driver as _dmd
    sys.modules.setdefault("dm_driver", _dmd)


def _build_xlsx(rows: List[dict]) -> bytes:
    """Identical schema to desktop listener_entry.py so existing
    web-app download links keep working."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Breakdown"
    headers = ["Our Ref", "Customer", "Consignment Fee", "Waiting",
               "Other", "Surcharge Total", "Total Revenue",
               "Surcharge Rows"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0EA5A4")
    for r in rows:
        ws.append([
            r.get("our_ref", ""), r.get("customer", ""),
            r.get("consignment_fee", ""), r.get("waiting_charge", ""),
            r.get("other_charge", ""), r.get("surcharge_total", ""),
            r.get("total_revenue", ""),
            len(r.get("surcharges") or []),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run(params: Dict[str, Any], on_progress: Callable[..., None], ctx) -> Dict[str, Any]:
    max_rows = int(params.get("max_rows") or 50)

    on_progress("Ensuring DM is logged in", percent=5)
    app = dm.ensure_logged_in(ctx, on_progress=on_progress, timeout=120)
    on_progress("Navigating to Invoicing", percent=10)
    dm.navigate_to_page(app, "Invoicing", on_progress=on_progress)

    _setup_engine_imports()
    from cal_listener import revenue_breakdown_driver as _drv

    rows: List[dict] = []

    def _engine_progress(i: int, n: int, row: dict) -> None:
        rows.append(row)
        pct = 15 + int(75 * (i / max(n, 1)))
        on_progress(
            f"[{i}/{n}] {row.get('our_ref','?')} {row.get('customer','')}".strip(),
            percent=min(pct, 90))

    def _engine_log(level: str, msg: str) -> None:
        on_progress(msg, level=level.lower() if level != "INFO" else "info")

    on_progress(f"Walking up to {max_rows} bookings", percent=14)
    result = _drv.dry_run(
        max_rows=max_rows,
        on_progress=_engine_progress,
        log_callback=_engine_log,
        should_stop=lambda: False,
    )

    if not rows and not result.get("ok"):
        on_progress(
            f"Engine returned no rows: {result.get('error', 'unknown')}",
            level="warning", percent=100,
        )
        return {
            "ok": False,
            "error": result.get("error") or "no rows scraped",
            "summary": result.get("summary") or {},
            "record_count": 0,
        }

    on_progress(f"Building xlsx for {len(rows)} rows", percent=92)
    xlsx = _build_xlsx(rows)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    key = f"revenue_breakdown_scraper/revenue_breakdown_{stamp}.xlsx"
    ok = ctx.sb.storage_upload(
        "listener_results", key, xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    public_url = ctx.sb.storage_public_url("listener_results", key) if ok else None

    on_progress(f"Done — {len(rows)} rows scraped", percent=100)
    return {
        "ok": True,
        "record_count": len(rows),
        "result_url": public_url,
        "summary": result.get("summary") or {},
    }
