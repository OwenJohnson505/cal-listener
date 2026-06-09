"""
Customer Weekly Bookings Report — engine.

Replaces the Excel-template workflow (Consignment Log + Bookings With
Other Charges Summary -> manual VLOOKUPs -> Staci-style CSV) with a
deterministic transformation driven by per-customer config.

The Staci output schema (28 columns) is shipped as a preset; per-
customer variations live in Customer 360 (carrier label, vehicle tariff
table, charge-keyword -> output-column rules).
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional


# --- 28-column Staci output schema -------------------------------------------
# Each entry: (header, source_kind, *source_args)
#   source_kind = 'fixed' | 'log' | 'computed' | 'charge'
#   log:      arg is the Consignment-Log column header
#   fixed:    arg is the literal value
#   computed: arg is one of {'tariff', 'from_addr', 'to_addr',
#                            'mileage_price', 'quoted_price',
#                            'price_change', 'blank'}
#   charge:   arg is the charge-target ID; engine pulls the matched
#             amount from the row's charge-target totals

STACI_OUTPUT_SCHEMA = [
    ("Carrier",                                  "fixed",    None),  # populated post-hoc if Carrier Ref present
    ("Invoice number",                           "log",      "Invoice Number"),
    ("mda reference ",                           "log",      "Reference"),
    ("Invoice price",                            "log",      "Revenue"),
    ("Collection Date ",                         "log",      "Ready At"),
    ("Carrier Reference",                        "log",      "Job No."),
    ("Booked By",                                "log",      "Booked By"),
    ("Vehicle Type",                             "log",      "Vehicle"),
    ("Tariff",                                   "computed", "tariff"),
    ("From",                                     "computed", "from_addr"),
    ("To",                                       "computed", "to_addr"),
    ("Delivery Deadline (Date & Time)",          "log",      "Delivery Deadline"),
    ("POD ",                                     "log",      "Name"),
    ("Delivery Time ",                           "log",      "POD Time"),
    ("Miles",                                    "log",      "Miles"),
    ("Mileage Price ",                           "computed", "mileage_price"),
    ("Out of Hours",                             "charge",   "out_of_hours"),
    ("Out of Area ",                             "charge",   "out_of_area"),
    ("Drop Charge",                              "charge",   "drop_charge"),
    ("Planned Return ",                          "charge",   "planned_return"),
    ("Other Costs (Tolls, Parking etc)",         "charge",   "tolls_parking"),
    ("Quoted Price ",                            "computed", "quoted_price"),
    ("Waiting Time           (after 30 mins)",   "charge",   "waiting_after_30"),
    ("Waiting Time Charge",                      "computed", "blank"),
    ("Unplanned Returns ",                       "charge",   "unplanned_returns"),
    ("Other costs ",                             "charge",   "other_misc"),
    ("Price Change ",                            "computed", "price_change"),
    ("Additional Information    ",               "charge",   "additional_info"),
]

# Charge target IDs the schema references. The 'additional_info' target
# is the catch-all for charge lines whose keyword doesn't match any rule.
CHARGE_TARGETS = {
    "out_of_hours", "out_of_area", "drop_charge", "planned_return",
    "tolls_parking", "waiting_after_30", "unplanned_returns",
    "other_misc", "additional_info",
}


@dataclass
class BookingsReportConfig:
    carrier_label: str = "CAL"
    # vehicle name (case-insensitive match) -> tariff label
    vehicle_tariff_table: dict[str, str] = field(default_factory=lambda: {
        "Small van": "0.94", "SWB": "1.25", "LWB": "1.5", "XLWB": "1.6",
        "Luton": "1.8", "7.5t": "POA", "18t": "POA", "40ft": "POA",
        # Customer-specific aliases (Staci):
        "Small Van (AB InBev)": "1.25",
    })
    # ordered list of (keyword_substring, target_id). First match wins.
    charge_rules: list[tuple[str, str]] = field(default_factory=lambda: [
        ("Waiting Time",     "waiting_after_30"),
        ("Congestion Charge", "tolls_parking"),
        ("TV Returns",       "planned_return"),
        ("Third Party",      "out_of_area"),
        ("Extra Charge",     "other_misc"),
    ])
    # Fallback target for charge lines that match no rule. Owen's
    # convention: anything that doesn't fit a specific named bucket
    # (Out of Hours, Drop Charge, Waiting Time, etc.) lands in the
    # 'Other costs' column. The 'Additional Information' column
    # still echoes the keyword + amount text for audit visibility —
    # see build_rows() which populates it from charge_text_unmatched.
    default_charge_target: str = "other_misc"


# --- Source-file readers ------------------------------------------------------

def read_consignment_log(path: str | Path) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(v is not None for v in r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


def read_other_charges(
        path: str | Path) -> tuple[dict[str, dict], list[dict]]:
    """Job No -> {'description': str, 'other_charge': float, 'revenue': float}.

    Returns (data, warnings). Each warning is a dict with `where`,
    `row`, `column`, `raw_value`, `reason`, `action` — surface them
    to the user; the report processes the file tolerantly but every
    row that didn't parse cleanly is recorded so nothing slips
    through unseen.

    Header-aware: looks up each field by column name rather than
    blindly trusting positions. The DM "Bookings With Other Charge
    Summary" export has historically shifted columns around (some
    exports include Customer / Booked By columns ahead of the charge
    column), which used to crash with `could not convert string to
    float: 'Staci (Blackburn)@'`. Now we find the right columns from
    the header row and coerce numbers tolerantly — with a warning
    recorded for every value that needed coercion."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    warnings: list[dict] = []
    if not rows:
        warnings.append({
            "where": "read_other_charges",
            "row": 0, "column": "", "raw_value": "",
            "reason": "Workbook has no rows",
            "action": "Skipped file entirely",
        })
        return {}, warnings
    headers = [str(h or "").strip() for h in rows[0]]

    def _col(*candidates: str) -> int:
        """Find the index of the first header matching any candidate
        (case-insensitive, whitespace-collapsed)."""
        norm_headers = [
            " ".join(h.lower().split()) for h in headers
        ]
        for cand in candidates:
            target = " ".join(cand.lower().split())
            for i, h in enumerate(norm_headers):
                if h == target:
                    return i
        # Substring fallback — useful when the header has trailing
        # bracketed annotations like 'Other Charges (£)'.
        for cand in candidates:
            target = " ".join(cand.lower().split())
            for i, h in enumerate(norm_headers):
                if target and target in h:
                    return i
        return -1

    def _to_float(v, *, row_num: int, col_name: str,
                    job: str = "") -> float:
        """Tolerant numeric coercion. Records a warning when the cell
        couldn't be parsed cleanly — Owen specifically asked that we
        never fall back silently. Returns 0.0 so the report continues
        but the caller sees exactly which rows were affected."""
        if v is None or v == "":
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "").replace("£", "")
        if not s:
            return 0.0
        try:
            return float(s)
        except (ValueError, TypeError):
            warnings.append({
                "where":     "read_other_charges",
                "row":       row_num,
                "column":    col_name,
                "raw_value": str(v),
                "reason":    "Not a number — couldn't convert to "
                              "float",
                "action":    f"Treated as 0.00 for job {job!r}",
            })
            return 0.0

    # Resolve column positions ONCE, then index into each row.
    # The DM "Bookings With Other Charge Summary" has BOTH:
    #   - 'Other Charges'    (plural)  — text like "Congestion - 18"
    #   - 'Other Charge (£)' (singular, £-suffixed) — the numeric total
    # We want col_desc = the text column, col_other = the numeric one.
    # That means the £-suffixed names MUST come first in the candidate
    # list for numeric fields, and the plural-without-£ name comes
    # first for the description.
    col_job = _col("Job No.", "Job No", "Job", "Job Number",
                    "BT Reference", "Booking Number")
    col_desc = _col(
        "Other Charges", "Other Charge Description",
        "Other Charges Description", "Description", "Charges")
    col_other = _col(
        "Other Charge (£)", "Other Charges (£)",
        "Other Charge Amount", "Other Charges Amount",
        "Total Other Charges", "Charges Amount",
        "Other Charge Total", "Charges Total")
    col_revenue = _col(
        "Revenue (£)", "Total Revenue (£)", "Net Revenue (£)",
        "Revenue", "Total Revenue", "Net Revenue",
        "Invoice Price", "Sale Price")
    # If col_desc collided with col_other (same column matched
    # because the user's header is ambiguous), force them apart by
    # picking the next plausible column.
    if (col_desc >= 0 and col_other >= 0
            and col_desc == col_other):
        col_desc = -1
    # Last-resort substring match for the numeric column — look for
    # a header containing "(£)" or "(GBP)" + "charge".
    if col_other < 0:
        for i, h in enumerate(headers):
            low = h.lower()
            if ("£" in h or "(gbp)" in low) \
                    and "charge" in low:
                col_other = i
                break
    # Only emit a mapping warning if a NUMERIC column is missing
    # (description is optional). For clean runs we stay silent so
    # the warnings dialog only pops when there's something real to
    # review.
    if col_other < 0 or col_revenue < 0:
        chosen = {
            "Job No.": headers[col_job] if col_job >= 0
                        else "(missing)",
            "Other Charge": headers[col_other] if col_other >= 0
                             else "(missing)",
            "Revenue":      headers[col_revenue] if col_revenue >= 0
                             else "(missing)",
        }
        warnings.append({
            "where": "read_other_charges",
            "row": 1, "column": "(header mapping)",
            "raw_value": "; ".join(headers),
            "reason": "Couldn't find one or more expected columns "
                       "in the header row",
            "action": " · ".join(
                f"{k}→'{v}'" for k, v in chosen.items()),
        })
    if col_job < 0:
        col_job = 0  # fall back to column A; covered by mapping warning above

    out: dict[str, dict] = {}
    for row_idx, r in enumerate(rows[1:], start=2):
        if not r:
            continue
        job_v = r[col_job] if col_job < len(r) else None
        if not job_v:
            continue
        job = str(job_v).strip()
        desc = ""
        if col_desc >= 0 and col_desc < len(r):
            desc = str(r[col_desc] or "")
        elif len(r) > 1 and r[1] and not _to_float(
                r[1], row_num=row_idx, col_name="(col B)", job=job):
            desc = str(r[1])
        other = 0.0
        if col_other >= 0 and col_other < len(r):
            other = _to_float(
                r[col_other], row_num=row_idx,
                col_name=headers[col_other], job=job)
        revenue = 0.0
        if col_revenue >= 0 and col_revenue < len(r):
            revenue = _to_float(
                r[col_revenue], row_num=row_idx,
                col_name=headers[col_revenue], job=job)
        out[job] = {
            "description":  desc,
            "other_charge": other,
            "revenue":      revenue,
        }
    return out, warnings


# --- Charge parsing -----------------------------------------------------------

# pdftotext / DM mangles newlines to _x000D_ sometimes; normalise first.
_LINE_BREAK = re.compile(r"_x000[dD]_\s*|\r\n?|\n")
_CHARGE_LINE = re.compile(r"^(?P<kw>.+?)\s*-\s*(?P<amt>-?[\d,]+\.?\d*)\s*$")


def parse_charge_text(text: str) -> list[tuple[str, float]]:
    """Parse 'Waiting Time - 50.00\\nCongestion Charge - 18.00' into
    [('Waiting Time', 50.0), ('Congestion Charge', 18.0)]."""
    if not text:
        return []
    out: list[tuple[str, float]] = []
    for line in _LINE_BREAK.split(text):
        line = line.strip()
        if not line:
            continue
        m = _CHARGE_LINE.match(line)
        if not m:
            continue
        try:
            amt = float(m.group("amt").replace(",", ""))
        except ValueError:
            continue
        kw = m.group("kw").strip()
        out.append((kw, amt))
    return out


def assign_to_target(keyword: str,
                      rules: list[tuple[str, str]],
                      default: str) -> str:
    """Match `keyword` against rules (substring, case-insensitive).
    First match wins; falls back to `default` if none match."""
    kw_low = keyword.lower()
    for needle, target in rules:
        if needle.lower() in kw_low:
            return target
    return default


# --- Output builder -----------------------------------------------------------

def _fmt_money(v) -> str:
    if v is None or v == "":
        return ""
    try:
        return f"£{float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_dt(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y %H:%M")
    return str(v)


def _fmt_int(v) -> str:
    if v is None or v == "":
        return ""
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return str(v)


def _lookup_vehicle(vehicle: str, table: dict[str, str]) -> str:
    if not vehicle:
        return ""
    vk = str(vehicle).strip().lower()
    for k, v in table.items():
        if k.lower() == vk:
            s = str(v).strip()
            # Format numeric tariffs as money; passthrough text like "POA".
            try:
                return f"£{float(s):,.2f}"
            except ValueError:
                return s
    return ""


def _normalise_dm_text(s: str) -> str:
    """Flatten DM's _x000D_ carriage-return artifacts (and any real
    CR/LF) into single spaces so multi-stop addresses render as one
    line in the CSV cell. Owen prefers no embedded newlines."""
    if s is None:
        return ""
    text = str(s)
    text = text.replace("_x000D_", " ").replace("\r\n", " ")
    text = text.replace("\r", " ").replace("\n", " ")
    # Collapse runs of whitespace down to a single space.
    import re as _re
    return _re.sub(r"\s+", " ", text).strip()


def _addr(parts: list) -> str:
    """Concatenate parts with no separator, matching the template's
    `=A&""&B&""&C` pattern. Skips None / empty parts. Normalises
    DM's _x000D_ artifacts to real CR/LF."""
    return "".join(_normalise_dm_text(p).strip() if isinstance(p, str)
                   else (str(p).strip() if p not in (None, "") else "")
                   for p in parts if p not in (None, ""))


def build_rows(log_rows: list[dict],
                charges_by_job: dict[str, dict],
                cfg: BookingsReportConfig
                ) -> tuple[list[list[str]], list[dict]]:
    """Build the output rows. Returns (rows, warnings).

    Warnings flag:
      - Log rows where Revenue couldn't be parsed as a number.
      - Charge entries with no matching consignment-log row (an entry
        in the Other Charges file with a Job No that doesn't appear
        in the consignment log — silently dropped before, now surfaced).
    """
    warnings: list[dict] = []
    out_rows = []
    seen_jobs: set[str] = set()
    for log_idx, r in enumerate(log_rows, start=2):
        job = str(r.get("Job No.", "") or "").strip()
        seen_jobs.add(job)
        raw_revenue = r.get("Revenue")
        try:
            revenue = float(raw_revenue or 0)
        except (ValueError, TypeError):
            warnings.append({
                "where":     "build_rows (consignment_log)",
                "row":       log_idx,
                "column":    "Revenue",
                "raw_value": str(raw_revenue),
                "reason":    "Not a number — couldn't convert to "
                              "float",
                "action":    f"Treated as 0.00 for job {job!r}",
            })
            revenue = 0.0

        # Parse + assign charges
        charge_totals: dict[str, float] = {t: 0.0 for t in CHARGE_TARGETS}
        charge_text_unmatched: list[str] = []
        ch_entry = charges_by_job.get(job)
        if ch_entry and ch_entry["other_charge"]:
            for kw, amt in parse_charge_text(ch_entry["description"]):
                target = assign_to_target(kw, cfg.charge_rules,
                                            cfg.default_charge_target)
                charge_totals[target] += amt
                if target == cfg.default_charge_target:
                    charge_text_unmatched.append(f"{kw} - {amt:.2f}")

        # Mileage Price = Revenue - sum of all charge-target amounts
        charge_sum = sum(charge_totals.values())
        mileage_price = revenue - charge_sum

        # Quoted Price = Mileage Price + cols 16..20 (Out of Hours .. Tolls)
        quoted_price = mileage_price + sum(charge_totals[t] for t in (
            "out_of_hours", "out_of_area", "drop_charge",
            "planned_return", "tolls_parking"))

        # Price Change = cols 22..25 (Waiting after 30, Waiting Charge (blank),
        # Unplanned Returns, Other costs). Waiting Charge is always blank
        # per Owen, so it contributes 0.
        price_change = sum(charge_totals[t] for t in (
            "waiting_after_30", "unplanned_returns", "other_misc"))

        # Build row from schema
        row: list = []
        for header, kind, arg in STACI_OUTPUT_SCHEMA:
            if kind == "fixed":
                # Carrier: only show label if Carrier Reference exists
                row.append(cfg.carrier_label if job else "")
            elif kind == "log":
                v = r.get(arg)
                # Format money / dates / ints by header convention
                if arg in ("Revenue", "Total", "Cost"):
                    row.append(_fmt_money(v))
                elif arg in ("Ready At", "Delivery Deadline", "POD Time"):
                    row.append(_fmt_dt(v))
                elif arg == "Miles":
                    row.append(_fmt_int(v))
                else:
                    row.append("" if v is None else str(v))
            elif kind == "computed":
                if arg == "tariff":
                    row.append(_lookup_vehicle(r.get("Vehicle"),
                                                 cfg.vehicle_tariff_table))
                elif arg == "from_addr":
                    row.append(_addr([
                        r.get("PU Company Name"),
                        r.get("Collecting From"),
                        r.get("Collection Postcode"),
                    ]))
                elif arg == "to_addr":
                    row.append(_addr([
                        r.get("Deliver To"),
                        r.get("Town"),
                        r.get("Delivery Postcode"),
                    ]))
                elif arg == "mileage_price":
                    row.append(_fmt_money(mileage_price))
                elif arg == "quoted_price":
                    row.append(_fmt_money(quoted_price))
                elif arg == "price_change":
                    row.append(_fmt_money(price_change))
                elif arg == "blank":
                    row.append("")
                else:
                    row.append("")
            elif kind == "charge":
                amt = charge_totals.get(arg, 0.0)
                if arg == "additional_info":
                    # For the catch-all, show the unmatched text rather
                    # than the £ figure (otherwise the user can't see
                    # what the unrecognised charge was).
                    row.append("; ".join(charge_text_unmatched))
                else:
                    row.append(_fmt_money(amt) if amt else "")
            else:
                row.append("")
        out_rows.append(row)
    # Surface every charges entry whose Job No didn't appear in the
    # consignment log — these are charges that would otherwise be
    # silently dropped from the output.
    for job, entry in charges_by_job.items():
        if job in seen_jobs:
            continue
        if not entry.get("other_charge") and not (
                entry.get("description") or "").strip():
            continue  # genuinely empty row, nothing to flag
        warnings.append({
            "where":     "build_rows (orphan charges)",
            "row":       0,
            "column":    "Job No",
            "raw_value": job,
            "reason":    "Charges entry has no matching row in the "
                          "consignment log",
            "action":    f"£{entry.get('other_charge') or 0:.2f} "
                          "of charges NOT included in the report — "
                          "manually review",
        })
    return out_rows, warnings


def write_csv(rows: list[list[str]], output_path: str | Path,
               schema=STACI_OUTPUT_SCHEMA) -> None:
    """Write the report as CSV with the schema headers."""
    headers = [h for h, _, _ in schema]
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)


# --- End-to-end --------------------------------------------------------------

def generate_report(consignment_log_path: str | Path,
                     other_charges_path: str | Path,
                     output_path: str | Path,
                     cfg: Optional[BookingsReportConfig] = None) -> dict:
    cfg = cfg or BookingsReportConfig()
    log_rows = read_consignment_log(consignment_log_path)
    charges, charge_warnings = read_other_charges(other_charges_path)
    rows, build_warnings = build_rows(log_rows, charges, cfg)
    write_csv(rows, output_path)
    return {
        "log_rows": len(log_rows),
        "charged_jobs": sum(1 for c in charges.values() if c["other_charge"]),
        "output_rows": len(rows),
        "warnings":  charge_warnings + build_warnings,
        "output_path": str(output_path),
    }


# ---------------------------------------------------------------------------
# Customer 360 integration
# ---------------------------------------------------------------------------

# Human-readable label for each charge target (= the output column it
# fills). Used by the Customer 360 'Bookings Report' tab to populate
# the target dropdown when editing charge rules.
CHARGE_TARGET_LABELS = {
    "out_of_hours":      "Out of Hours",
    "out_of_area":       "Out of Area",
    "drop_charge":       "Drop Charge",
    "planned_return":    "Planned Return",
    "tolls_parking":     "Other Costs (Tolls, Parking etc)",
    "waiting_after_30":  "Waiting Time (after 30 mins)",
    "unplanned_returns": "Unplanned Returns",
    "other_misc":        "Other costs",
    "additional_info":   "Additional Information (catch-all)",
}


def _config_from_profile(profile: dict) -> "BookingsReportConfig":
    """Build a config from a profile dict's `bookings_report` block.
    Missing keys / empty lists fall back to the class defaults."""
    br = (profile or {}).get("bookings_report") or {}
    defaults = BookingsReportConfig()
    carrier = (br.get("carrier_label") or "").strip()
    # vehicle_tariff_table from profile is list-of-dicts; convert to
    # the engine's dict shape for case-insensitive lookup.
    vt_rows = br.get("vehicle_tariff_table") or []
    vt_dict: dict[str, str] = {}
    for row in vt_rows:
        if isinstance(row, dict):
            v = (row.get("vehicle") or "").strip()
            t = row.get("tariff")
            if v:
                vt_dict[v] = str(t).strip() if t is not None else ""
    # charge_rules list-of-dicts -> list-of-tuples
    cr_rows = br.get("charge_rules") or []
    cr_tuples: list[tuple[str, str]] = []
    for row in cr_rows:
        if isinstance(row, dict):
            kw = (row.get("keyword") or "").strip()
            tgt = (row.get("target") or "").strip()
            if kw and tgt:
                cr_tuples.append((kw, tgt))
    return BookingsReportConfig(
        carrier_label=carrier or defaults.carrier_label,
        vehicle_tariff_table=vt_dict or defaults.vehicle_tariff_table,
        charge_rules=cr_tuples or defaults.charge_rules,
    )


def config_for_customer(customer_name: str) -> tuple["BookingsReportConfig", str]:
    """Look up the customer's profile and build a BookingsReportConfig.

    Returns (config, source) where source is 'profile' (custom config
    found) or 'defaults' (no profile / no bookings_report block).

    Safe to call even if customer_profile_store isn't importable — falls
    back to defaults silently."""
    try:
        import customer_profile_store as store  # type: ignore
    except Exception:
        return BookingsReportConfig(), "defaults"
    prof = None
    try:
        prof = store.find_by_alias(customer_name)
    except Exception:
        prof = None
    if not prof:
        try:
            prof = store.get_profile(customer_name)
        except Exception:
            prof = None
    if not prof:
        return BookingsReportConfig(), "defaults"
    br = prof.get("bookings_report") if isinstance(prof, dict) else None
    if not br or not isinstance(br, dict):
        return BookingsReportConfig(), "defaults"
    # Treat an all-empty bookings_report block as 'defaults' so the
    # plugin's status line is honest.
    if (not br.get("carrier_label")
            and not br.get("vehicle_tariff_table")
            and not br.get("charge_rules")):
        return BookingsReportConfig(), "defaults"
    return _config_from_profile(prof), "profile"
