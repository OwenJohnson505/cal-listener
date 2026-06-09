"""
LPO Report workbook builder.

Produces the 7-sheet review xlsx defined in Owen's spec:

  1. Summary       — counts + totals, mostly formula-driven so it
                     updates if anyone edits a data row
  2. LPO Report — one row per unique reference, with the
                  MATCH/VALUE DIFFERENCE/ONLY ON LOG/ONLY ON PDF
                  status and conditional formatting
  3. Customer LPOs (combined) — every line item from every LPO PDF,
                                 tagged with which LPO it came from
  4. Delivery Master — every relevant row from the log, with
                        On Customer LPOs? + LPO Week lookup columns
  5. Value Differences — just the rows where Status = VALUE DIFFERENCE
  6. Only on Delivery Master — references absent from any LPO
  7. Only on LPO — references absent from the Delivery Master log

All sheets use Arial, dark-blue header rows (#305496), thin grey
borders, frozen top row and auto-filter.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
TITLE_FONT  = Font(name="Arial", size=14, bold=True, color="305496")
NORMAL      = Font(name="Arial", size=10)
BOLD        = Font(name="Arial", size=10, bold=True)
CENTER      = Alignment(horizontal="center", vertical="center")
RIGHT       = Alignment(horizontal="right", vertical="center")
THIN        = Side(style="thin", color="B4B4B4")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OK_FILL     = PatternFill("solid", start_color="C6EFCE")
DIFF_FILL   = PatternFill("solid", start_color="FFE699")
MISS_FILL   = PatternFill("solid", start_color="F8CBAD")

MONEY        = "£#,##0.00"
MONEY_RED    = "£#,##0.00;[Red](£#,##0.00)"
LOG_DATE_FMT = "dd/mm/yyyy hh:mm"


def _style_header_row(ws, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _apply_body(ws, r1: int, r2: int, cols: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if cell.font.name != "Arial":
                cell.font = NORMAL


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(log_rows: list,
                    pdf_rows: list,
                    pdf_meta: list[tuple[str, float]],
                    matched: list,
                    output_path: str | Path) -> Path:
    """`log_rows` from engine.read_log(),
    `pdf_rows` is the combined list from every engine.parse_pdf(),
    `pdf_meta` is [(label, stated_lpo_total), ...] in display order,
    `matched` from engine.cross_reference()."""

    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "LPO Report — Delivery Master vs Customer LPO Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = (f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}")
    ws["A2"].font = Font(name="Arial", italic=True, color="6b7280")

    row = 4
    ws.cell(row, 1, "Counts").font = BOLD
    row += 1
    ws.cell(row, 1, "Delivery Master — references")
    ws.cell(row, 2, len(log_rows))
    row += 1
    for label, _ in pdf_meta:
        ws.cell(row, 1, f"LPO {label} — references")
        ws.cell(row, 2,
                 sum(1 for r in pdf_rows if r.source_pdf == label))
        row += 1
    ws.cell(row, 1, "LPO combined — references")
    ws.cell(row, 2, len(pdf_rows))
    row += 1
    ws.cell(row, 1, "References on BOTH (Delivery Master & LPO)")
    ws.cell(row, 2,
             "=COUNTIF('LPO Report'!E:E,\"MATCH\")"
             "+COUNTIF('LPO Report'!E:E,\"VALUE DIFFERENCE\")")
    row += 1
    ws.cell(row, 1, "  …values agree")
    ws.cell(row, 2, "=COUNTIF('LPO Report'!E:E,\"MATCH\")")
    row += 1
    ws.cell(row, 1, "  …values differ")
    ws.cell(row, 2,
             "=COUNTIF('LPO Report'!E:E,\"VALUE DIFFERENCE\")")
    row += 1
    ws.cell(row, 1, "On Delivery Master only (not on any LPO)")
    ws.cell(row, 2, "=COUNTA('Only on Delivery Master'!A:A)-1")
    row += 1
    ws.cell(row, 1, "On LPO only (not on Delivery Master)")
    ws.cell(row, 2, "=COUNTA('Only on LPO'!A:A)-1")
    row += 2
    ws.cell(row, 1, "Totals (£, ex VAT)").font = BOLD
    row += 1
    # Revenue sum — use an explicit range to avoid the SUM(I:I)
    # double-counting gotcha if a totals row ever lands at the
    # bottom of the log sheet.
    last_log_row = len(log_rows) + 1
    ws.cell(row, 1, "Delivery Master — sum of Value")
    ws.cell(row, 2,
             f"=SUM('Delivery Master'!H2:H{last_log_row})")
    ws.cell(row, 2).number_format = MONEY
    row += 1
    for label, lpo in pdf_meta:
        ws.cell(row, 1, f"LPO {label} — stated LPO total")
        ws.cell(row, 2, lpo or 0.0)
        ws.cell(row, 2).number_format = MONEY
        row += 1
    last_pdf_row = len(pdf_rows) + 1
    ws.cell(row, 1, "LPO combined — sum of line values")
    ws.cell(row, 2,
             f"=SUM('Customer LPOs (combined)'!I2:I{last_pdf_row})")
    ws.cell(row, 2).number_format = MONEY
    row += 1
    ws.cell(row, 1, "Matched refs — Delivery Master side")
    ws.cell(row, 2,
             "=SUMIFS('LPO Report'!C:C,'LPO Report'!E:E,\"MATCH\")"
             "+SUMIFS('LPO Report'!C:C,'LPO Report'!E:E,"
             "\"VALUE DIFFERENCE\")")
    ws.cell(row, 2).number_format = MONEY
    matched_log_row = row
    row += 1
    ws.cell(row, 1, "Matched refs — LPO side")
    ws.cell(row, 2,
             "=SUMIFS('LPO Report'!D:D,'LPO Report'!E:E,\"MATCH\")"
             "+SUMIFS('LPO Report'!D:D,'LPO Report'!E:E,"
             "\"VALUE DIFFERENCE\")")
    ws.cell(row, 2).number_format = MONEY
    matched_pdf_row = row
    row += 1
    ws.cell(row, 1, "Net difference on matched (Delivery Master − LPO)")
    ws.cell(row, 2, f"=B{matched_log_row}-B{matched_pdf_row}")
    ws.cell(row, 2).number_format = MONEY_RED

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 26

    # ---------------- LPO Report ----------------
    ws = wb.create_sheet("LPO Report")
    headers = [
        "Customer Reference", "Sources", "Delivery Master Value (£)",
        "LPO Value (£)", "Status", "Difference (DM−LPO, £)",
        "LPO Week", "Customer Name (LPO)", "LPO Order Ref",
        "Delivery Master Job No.", "Delivery Master Date",
        "LPO Pickup Date", "LPO Depot", "LPO Service",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    r = 2
    for m in matched:
        l = m.log_row
        p = m.pdf_row
        ws.cell(r, 1, m.reference)
        srcs = (["Log"] if l else []) + ([p.source_pdf] if p else [])
        ws.cell(r, 2, " + ".join(srcs))
        if l is not None and m.log_revenue is not None:
            ws.cell(r, 3, m.log_revenue)
        if p is not None and p.value is not None:
            ws.cell(r, 4, p.value)
        # Status — formula-driven so the workbook recalcs after edits.
        ws.cell(r, 5,
                 f"=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),"
                 f"IF(ABS(C{r}-D{r})<=0.005,\"MATCH\","
                 f"\"VALUE DIFFERENCE\"),"
                 f"IF(AND(NOT(ISNUMBER(C{r})),ISNUMBER(D{r})),"
                 f"\"ONLY ON PDF\","
                 f"IF(AND(ISNUMBER(C{r}),NOT(ISNUMBER(D{r}))),"
                 f"\"ONLY ON LOG\",\"NO VALUES\")))")
        ws.cell(r, 6,
                 f"=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),C{r}-D{r},\"\")")
        if p is not None:
            ws.cell(r, 7, p.source_pdf)
            ws.cell(r, 8, p.customer)
            ws.cell(r, 9, p.order_ref)
        if l is not None:
            ws.cell(r, 10, l.job_no)
            if isinstance(l.date, datetime):
                cell = ws.cell(r, 11, l.date)
                cell.number_format = LOG_DATE_FMT
            else:
                ws.cell(r, 11, l.date)
        if p is not None:
            ws.cell(r, 12, p.date)
            ws.cell(r, 13, p.depot)
            ws.cell(r, 14, p.svc)
        r += 1
    last = r - 1
    for rr in range(2, last + 1):
        ws.cell(rr, 3).number_format = MONEY
        ws.cell(rr, 4).number_format = MONEY
        ws.cell(rr, 6).number_format = MONEY_RED
    _apply_body(ws, 1, last, len(headers))
    _set_widths(ws, [20, 16, 16, 18, 22, 22, 18, 36, 26, 14, 18, 16, 12, 12])
    ws.freeze_panes = "A2"
    if last >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
        sr = f"E2:E{last}"
        ws.conditional_formatting.add(
            sr, FormulaRule(formula=['$E2="MATCH"'], fill=OK_FILL))
        ws.conditional_formatting.add(
            sr, FormulaRule(formula=['$E2="VALUE DIFFERENCE"'],
                             fill=DIFF_FILL))
        ws.conditional_formatting.add(
            sr, FormulaRule(
                formula=['OR($E2="ONLY ON LOG",$E2="ONLY ON PDF")'],
                fill=MISS_FILL))

    # ---------------- Customer LPOs (combined) ----------------
    ws = wb.create_sheet("Customer LPOs (combined)")
    ph = [
        "Customer Reference", "LPO Week", "Depot", "Service",
        "Pickup Date", "Customer Account", "Customer Name",
        "Order Reference", "Order Value (£)", "On Delivery Master?",
    ]
    ws.append(ph)
    _style_header_row(ws, len(ph))
    for i, p in enumerate(pdf_rows, start=2):
        ws.cell(i, 1, p.order)
        ws.cell(i, 2, p.source_pdf)
        ws.cell(i, 3, p.depot)
        ws.cell(i, 4, p.svc)
        ws.cell(i, 5, p.date)
        ws.cell(i, 6, p.account)
        ws.cell(i, 7, p.customer)
        ws.cell(i, 8, p.order_ref)
        ws.cell(i, 9, p.value)
        ws.cell(i, 10,
                 f"=IF(ISNUMBER(MATCH(A{i},'Delivery Master'!A:A,0)),"
                 f"\"Yes\",\"No\")")
    last = len(pdf_rows) + 1
    for rr in range(2, last + 1):
        ws.cell(rr, 9).number_format = MONEY
    _apply_body(ws, 1, last, len(ph))
    _set_widths(ws, [20, 14, 8, 10, 14, 16, 36, 28, 16, 22])
    ws.freeze_panes = "A2"
    if last >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(ph))}{last}"
        ws.conditional_formatting.add(
            f"J2:J{last}",
            FormulaRule(formula=['$J2="Yes"'], fill=OK_FILL))
        ws.conditional_formatting.add(
            f"J2:J{last}",
            FormulaRule(formula=['$J2="No"'], fill=MISS_FILL))
    # Totals row (blank row first so SUM doesn't accidentally include it
    # via SUM(I:I) elsewhere; we sum I2:I{last} explicitly anyway).
    tr = last + 2
    ws.cell(tr, 8, "Total:").font = BOLD
    ws.cell(tr, 8).alignment = RIGHT
    ws.cell(tr, 9, f"=SUM(I2:I{last})").font = BOLD
    ws.cell(tr, 9).number_format = MONEY

    # ---------------- Delivery Master ----------------
    ws = wb.create_sheet("Delivery Master")
    lh = [
        "Customer Reference", "Job No.", "Date", "Customer",
        "PU Company", "Collect From", "Deliver To",
        "Value (£)", "Customer VAT (£)", "Total inc VAT (£)",
        "On Customer LPOs?", "LPO Week",
    ]
    ws.append(lh)
    _style_header_row(ws, len(lh))
    for i, l in enumerate(log_rows, start=2):
        ws.cell(i, 1, l.reference)
        ws.cell(i, 2, l.job_no)
        if isinstance(l.date, datetime):
            cell = ws.cell(i, 3, l.date)
            cell.number_format = LOG_DATE_FMT
        else:
            ws.cell(i, 3, l.date)
        ws.cell(i, 4, l.customer)
        ws.cell(i, 5, l.pu_company)
        ws.cell(i, 6, l.collect_from)
        ws.cell(i, 7, l.deliver_to)
        ws.cell(i, 8, l.revenue)
        ws.cell(i, 9, l.cust_vat)
        ws.cell(i, 10, l.total)
        ws.cell(i, 11,
                 f"=IF(ISNUMBER(MATCH(A{i},"
                 f"'Customer LPOs (combined)'!A:A,0)),\"Yes\",\"No\")")
        ws.cell(i, 12,
                 f"=IFERROR(VLOOKUP(A{i},"
                 f"'Customer LPOs (combined)'!A:B,2,FALSE),\"\")")
    last = len(log_rows) + 1
    for rr in range(2, last + 1):
        for c in (8, 9, 10):
            ws.cell(rr, c).number_format = MONEY
    _apply_body(ws, 1, last, len(lh))
    _set_widths(ws, [20, 12, 18, 26, 30, 22, 26, 14, 16, 18, 20, 14])
    ws.freeze_panes = "A2"
    if last >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(lh))}{last}"
        ws.conditional_formatting.add(
            f"K2:K{last}",
            FormulaRule(formula=['$K2="Yes"'], fill=OK_FILL))
        ws.conditional_formatting.add(
            f"K2:K{last}",
            FormulaRule(formula=['$K2="No"'], fill=MISS_FILL))

    # ---------------- Value Differences ----------------
    ws = wb.create_sheet("Value Differences")
    vh = [
        "Customer Reference", "LPO Week", "Customer (LPO)",
        "Delivery Master Value (£)", "LPO Value (£)",
        "Difference (DM−LPO, £)", "Delivery Master Job No.",
        "LPO Pickup Date", "LPO Order Ref",
    ]
    ws.append(vh)
    _style_header_row(ws, len(vh))
    vr = 2
    for m in matched:
        if m.status != "VALUE DIFFERENCE":
            continue
        l = m.log_row
        p = m.pdf_row
        if l is None or p is None:
            continue
        ws.cell(vr, 1, m.reference)
        ws.cell(vr, 2, p.source_pdf)
        ws.cell(vr, 3, p.customer)
        ws.cell(vr, 4, m.log_revenue)
        ws.cell(vr, 5, p.value)
        ws.cell(vr, 6, f"=D{vr}-E{vr}")
        ws.cell(vr, 7, l.job_no)
        ws.cell(vr, 8, p.date)
        ws.cell(vr, 9, p.order_ref)
        vr += 1
    vlast = vr - 1
    for rr in range(2, max(vlast, 1) + 1):
        ws.cell(rr, 4).number_format = MONEY
        ws.cell(rr, 5).number_format = MONEY
        ws.cell(rr, 6).number_format = MONEY_RED
    _apply_body(ws, 1, max(vlast, 1), len(vh))
    _set_widths(ws, [20, 14, 36, 16, 18, 22, 14, 16, 28])
    ws.freeze_panes = "A2"
    if vlast >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(vh))}{vlast}"

    # ---------------- Only on Delivery Master ----------------
    ws = wb.create_sheet("Only on Delivery Master")
    oh = [
        "Customer Reference", "Job No.", "Date", "Customer",
        "Collect From", "Deliver To",
        "Value (£)", "Total inc VAT (£)",
    ]
    ws.append(oh)
    _style_header_row(ws, len(oh))
    only_log_refs = [m for m in matched if m.status == "ONLY ON LOG"]
    for i, m in enumerate(only_log_refs, start=2):
        l = m.log_row
        ws.cell(i, 1, m.reference)
        ws.cell(i, 2, l.job_no)
        if isinstance(l.date, datetime):
            cell = ws.cell(i, 3, l.date)
            cell.number_format = LOG_DATE_FMT
        else:
            ws.cell(i, 3, l.date)
        ws.cell(i, 4, l.customer)
        ws.cell(i, 5, l.collect_from)
        ws.cell(i, 6, l.deliver_to)
        ws.cell(i, 7, l.revenue)
        ws.cell(i, 8, l.total)
    olast = len(only_log_refs) + 1
    for rr in range(2, max(olast, 1) + 1):
        for c in (7, 8):
            ws.cell(rr, c).number_format = MONEY
    _apply_body(ws, 1, max(olast, 1), len(oh))
    _set_widths(ws, [20, 12, 20, 26, 24, 26, 14, 18])
    ws.freeze_panes = "A2"
    if olast >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(oh))}{olast}"
    tr = olast + 2
    ws.cell(tr, 6, "Total:").font = BOLD
    ws.cell(tr, 6).alignment = RIGHT
    ws.cell(tr, 7, f"=SUM(G2:G{olast})").font = BOLD
    ws.cell(tr, 7).number_format = MONEY
    ws.cell(tr, 8, f"=SUM(H2:H{olast})").font = BOLD
    ws.cell(tr, 8).number_format = MONEY

    # ---------------- Only on LPO ----------------
    ws = wb.create_sheet("Only on LPO")
    oph = [
        "Customer Reference", "LPO Week", "Depot", "Service",
        "Pickup Date", "Customer Account", "Customer Name",
        "Order Reference", "Order Value (£)",
    ]
    ws.append(oph)
    _style_header_row(ws, len(oph))
    only_pdf_refs = [m for m in matched if m.status == "ONLY ON PDF"]
    for i, m in enumerate(only_pdf_refs, start=2):
        p = m.pdf_row
        ws.cell(i, 1, p.order)
        ws.cell(i, 2, p.source_pdf)
        ws.cell(i, 3, p.depot)
        ws.cell(i, 4, p.svc)
        ws.cell(i, 5, p.date)
        ws.cell(i, 6, p.account)
        ws.cell(i, 7, p.customer)
        ws.cell(i, 8, p.order_ref)
        ws.cell(i, 9, p.value)
    oplast = len(only_pdf_refs) + 1
    for rr in range(2, max(oplast, 1) + 1):
        ws.cell(rr, 9).number_format = MONEY
    _apply_body(ws, 1, max(oplast, 1), len(oph))
    _set_widths(ws, [20, 14, 8, 10, 14, 16, 36, 28, 16])
    ws.freeze_panes = "A2"
    if oplast >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(oph))}{oplast}"
    tr = oplast + 2
    ws.cell(tr, 8, "Total:").font = BOLD
    ws.cell(tr, 8).alignment = RIGHT
    ws.cell(tr, 9, f"=SUM(I2:I{oplast})").font = BOLD
    ws.cell(tr, 9).number_format = MONEY

    # Force Arial on every populated cell to defeat Excel's default
    # Calibri leaking onto cells we didn't explicitly style.
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                f = cell.font
                if f.name != "Arial":
                    cell.font = Font(name="Arial", size=f.size or 10,
                                      bold=f.bold, color=f.color,
                                      italic=f.italic)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
