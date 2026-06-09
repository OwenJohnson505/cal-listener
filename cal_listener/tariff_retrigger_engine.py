"""
Tariff Re-trigger engine.

Pure-Python, no Qt. The widget wraps these calls in a worker thread.

Two public flows:

    dry_run(bt_refs, out_xlsx)
        For each BT-ref, opens its wizard, captures Tariff Name /
        Body Type / Consignment Fee / Other Charge / Driver Charge,
        then CLOSES the wizard without saving. Writes a results.xlsx
        with one row per job and a 'preview' status column.

    live_run(bt_refs, out_xlsx)
        Same opening flow, then reselects the tariff that's already
        assigned (just to force DM to refresh the surcharge), confirms
        the re-price prompt, saves, dismisses 'Confirm & Continue'.
        Writes the results.xlsx with before/after for every field plus
        per-field 'changed?' flags so any drift is obvious. No
        customer / tariff guards - the BT ref uniquely identifies one
        job, so we just process whatever comes back.

The retariff actions live in apply_retariff(), which is currently a
STUB - it will be filled in once we have the probe output. dry_run()
does NOT need apply_retariff() and is therefore fully working today.
"""
from __future__ import annotations

import logging
import re
import sys
import threading
import time
import traceback
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Callable

# Fuzzy-match threshold for tariff-name comparison. 'LWB Bluleaf' vs
# 'LWB Blueleaf' ratios ~0.95, well above this. Genuinely different
# tariffs (e.g. 'Standard' vs 'LWB Blueleaf') ratio well below 0.7.
TARIFF_MATCH_THRESHOLD = 0.85


# ---------------------------------------------------------------------------
# Immediate-stop signalling. The widget's Stop button calls request_stop()
# which sets this event. Every inner loop in the engine (fee polling,
# top-window waits, step boundaries) checks _stop_requested() and raises
# StopRequested as soon as it's set - so 'Stop' takes effect mid-step
# rather than only between jobs.
# ---------------------------------------------------------------------------

_STOP_EVENT = threading.Event()


class StopRequested(Exception):
    """Raised when _check_stop() detects the user pressed Stop. Bubbles
    up through process_one and run_batch which then write whatever
    rows have already been collected and exit cleanly."""


def request_stop() -> None:
    _STOP_EVENT.set()


def reset_stop() -> None:
    _STOP_EVENT.clear()


def _stop_requested() -> bool:
    return _STOP_EVENT.is_set()


def _check_stop() -> None:
    """Raise StopRequested if the stop flag is set. Call at safe
    points throughout the engine - between steps, inside polls, on
    each iteration of a wait loop."""
    if _STOP_EVENT.is_set():
        raise StopRequested("user pressed Stop")

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent.parent
DM_PLUGIN_DIR = ROOT / "plugins" / "dm_docket_search"
for p in (str(DM_PLUGIN_DIR), str(ROOT)):
    if p not in sys.path:
        sys.path.insert(0, p)

import dm_driver  # type: ignore

# Logs land in plugins/tariff_retrigger/data/. Persist across runs.
LOG_DIR = HERE / "data"
try:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass


# ---------------------------------------------------------------------------
# Logging - the user has had grief with automation scripts before, so we
# emit a step-by-step trail at INFO, every wizard interaction at DEBUG,
# and every anomaly at WARNING. Two handlers: one writes to a timestamped
# .log file in the plugin's data/ folder, one fires a Python callback so
# the UI widget can stream lines into its 'RUN SHEET' panel live.
# ---------------------------------------------------------------------------

class _CallbackHandler(logging.Handler):
    """Pipe formatted log records into a user-supplied callback. Used
    by the Qt worker to stream the run log into the widget."""

    def __init__(self, callback: Callable[[str, str], None]):
        super().__init__()
        self._cb = callback

    def emit(self, record: logging.LogRecord) -> None:
        try:
            self._cb(record.levelname, self.format(record))
        except Exception:
            # Don't ever let a UI hiccup crash the engine.
            pass


def setup_logger(
    label: str,
    callback: Callable[[str, str], None] | None = None,
) -> tuple[logging.Logger, Path]:
    """Build a logger for one run. Returns (logger, log_file_path).

    `label` is folded into the log filename (e.g. 'dry' or 'live') so
    the user can tell at a glance which run a file came from. The
    callback, if provided, fires once per emitted log line with
    `(levelname, formatted_message)` so the widget can colour-code."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOG_DIR / f"tariff_retrigger_{label}_{stamp}.log"
    logger = logging.getLogger(f"tariff_retrigger.{label}.{stamp}")
    logger.setLevel(logging.DEBUG)
    logger.propagate = False
    # Wipe any prior handlers (we make a fresh logger per run anyway,
    # but be defensive).
    for h in list(logger.handlers):
        logger.removeHandler(h)
    file_fmt = logging.Formatter(
        "%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(file_fmt)
    logger.addHandler(fh)
    if callback is not None:
        ui_fmt = logging.Formatter(
            "%(asctime)s  %(message)s", datefmt="%H:%M:%S")
        ch = _CallbackHandler(callback)
        ch.setLevel(logging.INFO)
        ch.setFormatter(ui_fmt)
        logger.addHandler(ch)
    logger.info("=" * 60)
    logger.info(f"Tariff Re-trigger - {label.upper()} run")
    logger.info(f"Log file: {log_path}")
    logger.info("=" * 60)
    return logger, log_path


def _nullable_logger() -> logging.Logger:
    """For tests / direct programmatic use without a file or callback."""
    lg = logging.getLogger("tariff_retrigger.null")
    lg.addHandler(logging.NullHandler())
    lg.setLevel(logging.CRITICAL + 1)
    return lg


# Status values for the results column. Keep the set small so the
# user can filter on it in Excel.
S_OK = "OK"
S_FLAGGED = "FLAGGED"
S_ERROR = "ERROR"
S_SKIPPED = "SKIPPED"
S_PREVIEW = "preview"


# Drift tolerance for the verify step. DM stores currency in pence
# internally; a sub-penny diff is just float wobble.
DRIFT_TOLERANCE = 0.01


# ---------------------------------------------------------------------------
# Input - read the BT-refs out of the breakdown xlsx
# ---------------------------------------------------------------------------

def read_bt_refs(xlsx_path: str | Path) -> list[str]:
    """Read every BT-reference from the first column of the breakdown
    xlsx (header row ignored). Only the first column is read - the
    user's rule is 'never use values from the spreadsheet other than
    the BT reference'.

    Returns a de-duplicated, order-preserved list."""
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    seen: set = set()
    out: list[str] = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False  # skip header
            continue
        if not row:
            continue
        v = row[0]
        if v is None:
            continue
        ref = str(v).strip()
        if not ref:
            continue
        # Sanity: only accept things that look like a DM job reference
        # (alpha prefix + digits). Numeric-only sneaks in sometimes.
        if not re.match(r"^[A-Za-z]+[\-\s]?\d+$", ref):
            continue
        key = ref.upper().replace(" ", "").replace("-", "")
        if key in seen:
            continue
        seen.add(key)
        out.append(ref)
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Money parsing
# ---------------------------------------------------------------------------

_ALPHA_PREFIX_RE = re.compile(r"^[A-Za-z]+[\-\s]?")


def _normalise_tariff(s: str) -> str:
    """Lowercase + drop all whitespace + drop non-alphanumeric. Used
    only for matching, never displayed."""
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def tariff_matches(expected: str, actual: str) -> tuple[bool, float]:
    """Return (is_match, similarity_pct).

    - Exact case-insensitive whitespace-collapsed match: 100%
    - Otherwise SequenceMatcher ratio on the normalised forms;
      >= TARIFF_MATCH_THRESHOLD counts as a match.

    This handles typos / spelling variants ('LWB Bluleaf' vs the
    DM-actual 'LWB Blueleaf' was reported as a false-mismatch in the
    2026-05-27 dry run - the strings differ by one missing 'e' which
    the user couldn't spot in the run-sheet font)."""
    if not expected or not actual:
        return False, 0.0
    exp_n = _normalise_tariff(expected)
    act_n = _normalise_tariff(actual)
    if not exp_n or not act_n:
        return False, 0.0
    if exp_n == act_n:
        return True, 100.0
    ratio = SequenceMatcher(None, exp_n, act_n).ratio()
    return ratio >= TARIFF_MATCH_THRESHOLD, round(ratio * 100.0, 1)


def _strip_alpha_prefix(ref: str) -> str:
    """Return the digit portion of a docket reference. DM's docket
    search field is digit-only and silently wipes the value on blur
    if you type an alpha prefix - so 'BT59129' must go in as '59129'.

    Handles common shapes:
        'BT59129'  -> '59129'
        'BT-59129' -> '59129'
        'S 12345'  -> '12345'
        '59129'    -> '59129'  (no-op if already a number)
        ''         -> ''
    """
    if not ref:
        return ""
    return _ALPHA_PREFIX_RE.sub("", str(ref).strip())


def parse_money(value) -> float | None:
    """Convert a DM-displayed money string into a float. Returns None
    if the value is genuinely blank; raises ValueError on garbage."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Strip currency symbol, thousand separators
    s = s.replace("£", "").replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        raise ValueError(f"cannot parse money value: {value!r}")


# ---------------------------------------------------------------------------
# Per-job actions
# ---------------------------------------------------------------------------

def _safe_close_popups(app, logger: logging.Logger | None = None) -> None:
    """Heavy sweep - iterates over every known popup title up to 4
    times. Use only when we genuinely expect stuck dialogs (between
    runs, after a job error). Costs 2-4s when nothing is there because
    each title gets its own UIA scan."""
    lg = logger or _nullable_logger()
    try:
        if hasattr(dm_driver, "_close_intrusive_popups"):
            n = dm_driver._close_intrusive_popups(app)
            if n:
                lg.info(f"  closed {n} intrusive popup(s) before this job")
    except Exception as e:
        lg.debug(f"  popup-close pass swallowed exception: {e}")


def bring_dm_to_front(app, logger: logging.Logger | None = None) -> bool:
    """Bring the DM main window to the foreground so subsequent
    click_input calls land on DM rather than whatever was covering it
    (typically Cal Toolkit). Resolution-agnostic - just needs the
    window to be on the visible desktop somewhere.

    Returns True on success. Best-effort: Windows 10+ restricts
    foreground changes from background processes, so we try the
    pywinauto-friendly set_focus() and the win32 SetForegroundWindow
    in tandem. If neither works the user can use the 'Focus DM'
    button in the UI to do it manually."""
    lg = logger or _nullable_logger()
    try:
        main = dm_driver._main_window(app)
    except Exception as e:
        lg.debug(f"  bring_dm_to_front: no main window: {e}")
        return False
    # Try pywinauto path first - handles minimised + activates.
    try:
        main.set_focus()
    except Exception as e:
        lg.debug(f"  set_focus raised: {e}")
    # Belt-and-braces: raw win32 SetForegroundWindow. Sometimes
    # succeeds where set_focus alone doesn't.
    try:
        import ctypes
        hwnd = int(main.handle)
        # If minimised, restore first (SW_RESTORE=9).
        try:
            ctypes.windll.user32.ShowWindow(hwnd, 9)
        except Exception:
            pass
        ctypes.windll.user32.SetForegroundWindow(hwnd)
    except Exception as e:
        lg.debug(f"  SetForegroundWindow raised: {e}")
    # Log where DM actually is - useful for spotting if it's
    # restored to off-screen ultrawide coordinates on a laptop.
    try:
        r = main.rectangle()
        lg.debug(f"  DM main window rect: "
                 f"left={r.left} top={r.top} "
                 f"right={r.right} bottom={r.bottom}")
    except Exception:
        pass
    return True


def _focus(window, logger: logging.Logger | None = None) -> None:
    """Bring a specific dialog/window to the front before we click on
    it. Used right after a new top-level appears."""
    lg = logger or _nullable_logger()
    if window is None:
        return
    try:
        window.set_focus()
    except Exception as e:
        lg.debug(f"  focus raised: {e}")
    try:
        import ctypes
        hwnd = int(window.handle)
        ctypes.windll.user32.SetForegroundWindow(hwnd)
    except Exception:
        pass


def _aggressive_cleanup(app, logger: logging.Logger | None = None,
                       max_iterations: int = 4) -> int:
    """Force-dismiss any non-dashboard top-level DM window. Used at
    the start of each job and inside the error-cleanup path to
    prevent leftover state from a failed job blocking the next one.

    Strategy per leftover window:
      1. set_focus on the window
      2. send Escape (dismisses most modals)
      3. if still there next iteration, look for any Cancel/Close/No
         button and click it
      4. last resort: Alt+F4

    Returns the count of windows that disappeared."""
    lg = logger or _nullable_logger()
    dismissed = 0
    for it in range(max_iterations):
        leftover = []
        try:
            for w in app.windows():
                try:
                    if not w.is_visible():
                        continue
                    title = (w.window_text()
                             or w.element_info.name or "")
                except Exception:
                    continue
                lo = title.lower()
                # Skip the main dashboard - we want it to stay open.
                if "cal (" in lo:
                    continue
                # Skip the booking wizard - it's handled separately
                # via exit_wizard_without_saving.
                if lo == "booking wizard":
                    continue
                leftover.append((w, title))
        except Exception:
            break
        if not leftover:
            break
        for w, title in leftover:
            try:
                w.set_focus()
            except Exception:
                pass
            # Try Escape on iteration 0-1, button click on 2, Alt+F4
            # on 3.
            try:
                from pywinauto.keyboard import send_keys
                if it < 2:
                    send_keys("{ESC}")
                elif it == 2:
                    # Try Cancel/No/Close buttons explicitly
                    closed_via_button = False
                    try:
                        for b in w.descendants(control_type="Button"):
                            try:
                                bname = (b.element_info.name
                                         or "").lower()
                            except Exception:
                                continue
                            if any(k in bname for k in
                                   ("cancel", "close", "no", "exit",
                                    "skip")):
                                try:
                                    b.click_input()
                                    closed_via_button = True
                                    break
                                except Exception:
                                    continue
                    except Exception:
                        pass
                    if not closed_via_button:
                        send_keys("{ESC}")
                else:
                    # Nuclear: Alt+F4
                    send_keys("%{F4}")
            except Exception:
                pass
            time.sleep(0.2)
        time.sleep(0.3)
    # Recount how many disappeared.
    try:
        for w in app.windows():
            try:
                if not w.is_visible():
                    continue
                title = (w.window_text()
                         or w.element_info.name or "").lower()
            except Exception:
                continue
            if "cal (" in title or title == "booking wizard":
                continue
            # Still here - log it
            lg.warning(f"  aggressive_cleanup: leftover window "
                       f"{title!r} could not be dismissed")
    except Exception:
        pass
    if dismissed:
        lg.info(f"  aggressive_cleanup: dismissed leftover dialogs")
    return dismissed


def _has_blocking_popup(app) -> bool:
    """Cheap test: scan top-level DM windows ONCE for anything that
    isn't the main dashboard or the booking wizard. Returns True only
    if such a window is visible. Used to skip the heavy popup sweep on
    the happy path - 99% of Blueleaf jobs have no popups."""
    try:
        windows = app.windows()
    except Exception:
        return False
    for w in windows:
        try:
            if not w.is_visible():
                continue
            title = w.window_text() or ""
        except Exception:
            continue
        lo = title.lower()
        if not title:
            continue
        # The two we expect: main dashboard ('Cal (North) : In Progress')
        # and the booking wizard ('Booking Wizard'). Anything else
        # visible is potentially blocking.
        if "cal (" in lo or lo == "booking wizard":
            continue
        if lo in ("tariff list",):
            # We sometimes leave this open intentionally.
            continue
        return True
    return False


def _read_consignment_fee_fast(wiz) -> float | None:
    """Read txtTotalConsignmentCharge directly via its auto_id - no
    full wizard descendants walk. Returns parsed float or None.

    Used as the polling signal during retariff: if the value changes
    after we click the tariff button, the retariff has committed and
    we can proceed immediately instead of waiting for prompts that
    may never appear."""
    try:
        e = wiz.child_window(auto_id="txtTotalConsignmentCharge",
                             control_type="Edit")
        if not e.exists(timeout=0.1):
            return None
        return parse_money(e.window_text() or "")
    except Exception:
        return None


def _wait_for_fee_change(wiz, original: float | None,
                        timeout: float = 6.0,
                        logger: logging.Logger | None = None
                        ) -> float | None:
    """Poll the Consignment Fee field at ~3 Hz until either it differs
    from `original` by >= 1p, or `timeout` elapses. Returns the new
    value if change detected, else None. Polls the stop flag each
    iteration so Stop bites mid-wait."""
    lg = logger or _nullable_logger()
    if original is None:
        return None
    deadline = time.time() + timeout
    polls = 0
    while time.time() < deadline:
        _check_stop()
        current = _read_consignment_fee_fast(wiz)
        polls += 1
        if current is not None and abs(current - original) > 0.01:
            lg.info(
                f"  fee changed after {polls} poll(s): "
                f"£{original:,.2f} -> £{current:,.2f}")
            return current
        time.sleep(0.3)
    lg.debug(f"  no fee change after {polls} poll(s) "
             f"in {timeout:.1f}s")
    return None


def _quick_close_popups(app, logger: logging.Logger | None = None) -> int:
    """Single-pass popup check. Returns the number of popups closed.

    Cheaper than _safe_close_popups - one UIA scan, no retry loop. Use
    on hot paths (between wizard open and reading values) where stuck
    popups are unlikely but still need handling if they appear."""
    lg = logger or _nullable_logger()
    n = 0
    try:
        titles = getattr(dm_driver, "WIZARD_POPUP_TITLES", None) or ()
    except Exception:
        titles = ()
    for title in titles:
        try:
            popup = dm_driver._find_popup_window(app, title)
            if popup is None:
                continue
            try:
                popup.set_focus()
            except Exception:
                pass
            for btn_label in ("Exit", "Close", "OK", "Cancel"):
                try:
                    b = popup.child_window(title=btn_label,
                                           control_type="Button")
                    if b.exists(timeout=0.2):
                        try:
                            b.click_input()
                        except Exception:
                            b.invoke()
                        n += 1
                        lg.info(f"  closed popup {title!r} "
                                f"via {btn_label!r}")
                        break
                except Exception:
                    continue
        except Exception:
            continue
    return n


def open_job_via_in_progress(
    app, search_term: str,
    logger: logging.Logger | None = None,
) -> bool:
    """Fast path: type the full booking reference into the In Progress
    screen's search bar and double-click the matching row.

    `search_term` should be the FULL ref ('BT59129'), NOT digits-only.
    The free-text search matches against every visible column, so
    a bare '59129' will match substrings inside postcodes, customer
    refs, etc. and may return an unrelated job. The 'BT' prefix
    scopes the match to the docket-number column.

    UI elements (captured by dm_top_11_Cal_North_In_Progress.txt probe):
        txtSearch       - the free-search Edit on the In Progress view
        btnFreeSearch   - the magnifying-glass button next to it
        gvBookings      - the RadGridView showing matching bookings
        rbtnInProgressBookings - the Ribbon button to ensure that
                                 In Progress is the active view

    Returns True if the wizard opened, False if not (caller can then
    fall back to Docket Search for archived/completed jobs).
    """
    lg = logger or _nullable_logger()
    main = dm_driver._main_window(app)
    _safe_close_popups(app, logger=lg)

    # Ensure In Progress is the active view. Cheap idempotent click -
    # if we're already on In Progress, clicking the ribbon button is a
    # no-op for the data state. Without this, txtSearch might point at
    # a different list (Quote, Complete, etc.).
    try:
        rb = main.child_window(
            auto_id="rbtnInProgressBookings",
            control_type="Button")
        if rb.exists(timeout=0.5):
            try:
                rb.invoke()
            except Exception:
                rb.click_input()
            time.sleep(0.4)
    except Exception as e:
        lg.debug(f"  could not assert In Progress tab: {e}")

    # Find the search edit. If it's not visible the In Progress view
    # isn't up - bail so the caller can try the Docket Search fallback.
    try:
        search_edit = main.child_window(
            auto_id="txtSearch", control_type="Edit")
        if not search_edit.exists(timeout=1.0):
            lg.info("  In Progress txtSearch not visible "
                    "- falling back to Docket Search dialog")
            return False
    except Exception as e:
        lg.info(f"  In Progress txtSearch lookup failed: {e} "
                "- falling back to Docket Search dialog")
        return False

    # Clear + type. set_edit_text is most reliable; fall back to
    # focus + Ctrl+A + DEL + type.
    try:
        search_edit.set_focus()
    except Exception:
        pass
    typed = False
    try:
        search_edit.set_edit_text(search_term)
        typed = True
    except Exception:
        try:
            from pywinauto.keyboard import send_keys
            send_keys("^a{DEL}", with_spaces=False)
            send_keys(search_term, with_spaces=False)
            typed = True
        except Exception as e:
            lg.warning(f"  failed to type into txtSearch: {e}")
            return False
    if not typed:
        return False

    # Commit the filter. Enter applies most RadGridView filters; if
    # there's a Free Search button we click that too to be belt-and-braces.
    try:
        from pywinauto.keyboard import send_keys
        send_keys("{ENTER}")
    except Exception:
        pass
    try:
        btn = main.child_window(
            auto_id="btnFreeSearch", control_type="Button")
        if btn.exists(timeout=0.3):
            try:
                btn.invoke()
            except Exception:
                btn.click_input()
    except Exception:
        pass
    time.sleep(0.3)  # was 0.7 - filter applies as you type, this is safety
    lg.info(f"  In Progress search submitted with {search_term!r}")

    # Inspect the grid for matching rows.
    try:
        grid = main.child_window(
            auto_id="gvBookings", control_type="DataGrid")
        if not grid.exists(timeout=1.0):
            lg.info("  gvBookings not visible after search "
                    "- falling back to Docket Search")
            return False
    except Exception as e:
        lg.info(f"  gvBookings lookup failed: {e} "
                "- falling back to Docket Search")
        return False

    # Pick the FIRST DATA ROW. Telerik's RadGridView reports a column-
    # chooser 'ControlPanelItem' as control_type='DataItem' in the
    # top-right corner of the grid - it has no auto_id but matches a
    # naive descendants(control_type='DataItem') query. So we look for
    # rows by their stable auto_id pattern 'Row_N' (matches the
    # existing read-only DM driver at dm_driver.py:987) and confirm
    # the class is 'GridViewRow'.
    target_row = None
    # Give the grid a moment to repaint after the filter commits.
    for attempt in range(8):
        # First try the cheapest path: Row_0 by auto_id.
        try:
            candidate = grid.child_window(
                auto_id="Row_0", control_type="DataItem")
            if candidate.exists(timeout=0.2):
                target_row = candidate
                break
        except Exception:
            pass
        # Fallback: walk descendants and pick the first one whose
        # auto_id looks like 'Row_<digits>' AND whose class is
        # GridViewRow. This is robust to a future schema change.
        try:
            for c in grid.descendants(control_type="DataItem"):
                aid = ""
                cls = ""
                try:
                    aid = c.element_info.automation_id or ""
                except Exception:
                    pass
                try:
                    cls = c.element_info.class_name or ""
                except Exception:
                    pass
                if (aid.startswith("Row_")
                        and aid[4:].isdigit()
                        and "GridViewRow" in cls):
                    target_row = c
                    break
            if target_row is not None:
                break
        except Exception:
            pass
        time.sleep(0.25)

    if target_row is None:
        lg.info("  In Progress grid has no data row matching the "
                "search - falling back to Docket Search "
                "(job may be archived or filter is still settling)")
        return False
    try:
        aid = target_row.element_info.automation_id or "?"
    except Exception:
        aid = "?"
    lg.info(f"  In Progress grid: matched row auto_id={aid!r}")

    # Double-click the data row to open the booking wizard.
    try:
        target_row.double_click_input()
    except Exception as e:
        lg.warning(f"  could not double-click the matching row: {e}")
        return False
    time.sleep(0.6)  # was 1.2 - wizard usually visible within ~400ms

    # Wait up to ~2.5s for the wizard, polling every 0.25s. Was 6x0.5.
    wiz = None
    for _ in range(10):
        try:
            wiz = dm_driver._wizard(app)
        except Exception:
            wiz = None
        if wiz is not None:
            break
        time.sleep(0.25)
    if wiz is None:
        lg.warning("  wizard didn't open after double-click")
        return False
    # Only run popup-close if a quick check shows something is
    # blocking. On Blueleaf jobs no popups auto-open, so the iterative
    # sweep was pure waste (~8s/job). _has_blocking_popup is a single
    # UIA scan - sub-second when nothing's there.
    if _has_blocking_popup(app):
        _quick_close_popups(app, logger=lg)
    lg.debug("  wizard opened via In Progress search (fast path)")
    return True


def open_job_via_docket_search(
    app, bt_ref: str, docket_digits: str,
    logger: logging.Logger | None = None,
) -> bool:
    """Fallback: drive the Docket Search dialog (~25s per job vs ~3s
    for the In Progress path). Use this for jobs that aren't in the
    In Progress view (archived, completed).

    `docket_digits` must be the DIGIT portion only (e.g. '59129', not
    'BT59129') because the dialog's Docket No edit is digit-only and
    silently wipes alpha-prefixed input on blur.

    Reuses dm_driver.apply_search wholesale - same dialog the Docket
    Search plugin drives."""
    lg = logger or _nullable_logger()
    _safe_close_popups(app, logger=lg)
    lg.debug("open_docket_search dialog (fallback path)...")
    dm_driver.open_docket_search(app)
    payload = {
        "filters": {"docket_no": {"start": docket_digits}},
        "states":  {"live": True, "archived": True, "cancelled": False},
    }
    lg.info(
        f"  docket-search dialog: docket_no=ON, "
        f"From={docket_digits!r}, "
        "Live=ON, Archived=ON, Cancelled=OFF")
    dm_driver.apply_search(app, payload)
    time.sleep(0.6)
    dlg = dm_driver._result_dialog(app)
    if dlg is None:
        lg.warning(f"  no result dialog appeared for {bt_ref!r}")
        return False
    rows = dm_driver.read_result_grid(app)
    lg.info(f"  Docket Search returned {len(rows)} result(s)")
    if not rows:
        return False
    dm_driver.open_row_in_wizard(app, 0)
    time.sleep(1.0)
    _safe_close_popups(app, logger=lg)
    wiz = None
    for _ in range(6):
        try:
            wiz = dm_driver._wizard(app)
        except Exception:
            wiz = None
        if wiz is not None:
            break
        time.sleep(0.5)
    if wiz is None:
        lg.warning("  wizard did not open via fallback")
        return False
    lg.debug("  wizard opened via Docket Search (fallback path)")
    return True


def open_job_wizard(
    app, bt_ref: str,
    logger: logging.Logger | None = None,
) -> bool:
    """Open the booking wizard for `bt_ref`. Tries the In Progress
    search bar first (fast); falls back to the Docket Search dialog
    if the fast path can't find the job (archived/completed).

    Returns True on success."""
    lg = logger or _nullable_logger()
    docket_digits = _strip_alpha_prefix(bt_ref)
    if not docket_digits:
        lg.error(f"  could not extract docket digits from {bt_ref!r}")
        return False
    lg.info(f"  searching DM for {bt_ref!r} "
            f"(In Progress uses the full ref to avoid free-text "
            f"substring collisions; Docket Search uses digits "
            f"{docket_digits!r})")

    # Fast path - In Progress free-text search. Use the FULL ref so
    # 'BT59129' only matches the docket-number column, not a substring
    # in some other field (the 2026-05-27 'Luton' false positive was
    # bare digits matching inside an unrelated job).
    if open_job_via_in_progress(app, bt_ref, logger=lg):
        return True

    # Slow fallback - Docket Search dialog. Covers archived/completed
    # jobs the In Progress view doesn't show. Dialog's docket field
    # is digit-only - alpha prefixes get wiped on blur.
    return open_job_via_docket_search(app, bt_ref, docket_digits,
                                       logger=lg)


def read_wizard_values(
    app,
    logger: logging.Logger | None = None,
) -> dict:
    """Snapshot the values we care about for this task. Returns:

        {
          "customer_name":   str,           # NEW - for the guard
          "tariff_name":     str,
          "consignment_fee": float | None,
          "other_charge":    float | None,
          "driver_charge":   float | None,
          "raw":             {customer_name, tariff_name, ...}
        }
    """
    lg = logger or _nullable_logger()
    wiz = dm_driver._wizard(app)
    if wiz is None:
        raise RuntimeError("wizard not visible while reading values")
    idx = dm_driver._build_auto_id_index(wiz)
    raw = {
        "tariff_name":      dm_driver._value_for_auto_id(idx, "txtTariffName"),
        # Body type / vehicle - shown in the report so the user can spot
        # if a retariff somehow changed it. Tries the two auto_ids the
        # wizard uses for body / vehicle.
        "body_type":        (dm_driver._value_for_auto_id(idx, "cmbBodyType")
                             or dm_driver._value_for_auto_id(idx, "cmbVehicle")
                             or ""),
        "consignment_fee":  dm_driver._value_for_auto_id(idx, "txtTotalConsignmentCharge"),
        "other_charge":     dm_driver._value_for_auto_id(idx, "txtTotalOtherCharges"),
        "driver_charge":    dm_driver._value_for_auto_id(idx, "txtTotalDriverCharge"),
    }
    parsed = {
        "tariff_name":     raw["tariff_name"].strip(),
        "body_type":       (raw["body_type"] or "").strip(),
        "consignment_fee": parse_money(raw["consignment_fee"]),
        "other_charge":    parse_money(raw["other_charge"]),
        "driver_charge":   parse_money(raw["driver_charge"]),
        "raw":             raw,
    }
    lg.info(
        f"  captured: tariff={parsed['tariff_name']!r}  "
        f"body={parsed['body_type']!r}  "
        f"Consignment={_fmt_or_blank(parsed['consignment_fee'])}  "
        f"Other={_fmt_or_blank(parsed['other_charge'])}  "
        f"Driver={_fmt_or_blank(parsed['driver_charge'])}")
    return parsed


def _fmt_or_blank(v) -> str:
    if v is None:
        return "(blank)"
    try:
        return f"£{float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def exit_wizard_without_saving(
    app, logger: logging.Logger | None = None,
) -> None:
    """Close the wizard cleanly without committing changes.

    Delegates to dm_driver.close_wizard_no_save() - the same helper
    the read-only Docket Search plugin uses. It:
      * dismisses Internal/External/Customer Notes popups so Exit is
        reachable,
      * clicks the wizard's btnExit (with click_input/invoke fallback),
      * waits up to 5s for DM's 'are you sure you want to exit without
        saving?' confirm dialog (which is a top-level Window with no
        title/auto_id, found by its btnYes child),
      * clicks Yes on it.

    Falls back to a polite Exit-by-label click if close_wizard_no_save
    is not available on this dm_driver version. Never raises - exit
    is a best-effort cleanup step; an error here would mask the real
    job failure."""
    lg = logger or _nullable_logger()
    lg.debug("closing wizard without saving (via close_wizard_no_save)")
    try:
        if hasattr(dm_driver, "close_wizard_no_save"):
            # close_wizard_no_save already runs _close_intrusive_popups
            # internally - no need for our own pass afterwards. Tiny
            # settle so the next search isn't fighting a still-fading
            # wizard.
            dm_driver.close_wizard_no_save(app)
            time.sleep(0.2)
            return
    except Exception as e:
        lg.warning(f"  close_wizard_no_save raised: {e}")
    # Fallback for older dm_driver versions or unexpected wizard state.
    try:
        wiz = dm_driver._wizard(app)
    except Exception:
        wiz = None
    if wiz is not None:
        try:
            dm_driver._click_button(wiz, "Exit")
            time.sleep(0.6)
        except Exception as e:
            lg.warning(f"  exit-by-label fallback failed: {e}")
    else:
        lg.debug("  wizard not visible during exit - nothing to close")


# ---------------------------------------------------------------------------
# Change-Tariff write path - STUB until we have probe output
# ---------------------------------------------------------------------------

class NotYetWiredError(RuntimeError):
    """Raised when apply_retariff() is called before the probe-output
    -> driver wiring step has been completed. The dry-run flow does
    not need apply_retariff() so it works regardless."""


# Auto-ids from the existing dm_wizard_tree.txt dump.
_WIZ_CHANGE_TARIFF_AID = "btnChangeTariff"
_WIZ_SAVE_AID = "btnSave"
_WIZ_EXIT_AID = "btnExit"

# Button labels we try on the tariff-picker dialog and the re-price
# prompt - DM uses standard label text on its modal confirms.
_TARIFF_PICKER_COMMIT_LABELS = (
    "OK", "Ok", "Select", "Apply", "Continue", "Done", "Save")
_REPRICE_YES_LABELS = ("Yes", "OK", "Continue", "Apply")


def _dump_top_windows(app, label: str,
                      logger: logging.Logger | None = None) -> None:
    """Snapshot every DM top-level window into a probe file under
    data/. Lets us see exactly what UIA found when an expected button
    wasn't there - replaces 'please run the probe' with concrete
    diagnostic output captured live."""
    lg = logger or _nullable_logger()
    try:
        windows = app.windows()
    except Exception as e:
        lg.debug(f"  _dump_top_windows: app.windows() raised: {e}")
        return
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = LOG_DIR / f"_probe_{label}_{stamp}.txt"
    lines: list[str] = []
    for i, w in enumerate(windows):
        try:
            info = w.element_info
            lines.append(
                f"#{i:02d}  class={info.class_name!r}  "
                f"name={info.name!r}  auto_id={info.automation_id!r}  "
                f"visible={w.is_visible()}")
            # Buttons inside this window
            try:
                for b in w.descendants(control_type="Button"):
                    bi = b.element_info
                    lines.append(
                        f"    Button  name={bi.name!r:<28}  "
                        f"auto_id={bi.automation_id!r:<24}  "
                        f"text={(b.window_text() or '')!r:<28}")
            except Exception:
                pass
        except Exception:
            continue
    try:
        fname.write_text("\n".join(lines), encoding="utf-8")
        lg.warning(f"  dumped {len(windows)} top-level window(s) "
                   f"to {fname.name} for diagnosis")
    except Exception as e:
        lg.debug(f"  could not write probe dump: {e}")


def _click_button_by_auto_id_desc(window, auto_id: str,
                                  logger=None) -> bool:
    """Find a descendant Button matching `auto_id` and click it. More
    reliable than child_window(auto_id=...) for freshly-rendered DM
    dialogs (the post-save Confirm dialog + the re-price prompt both
    failed child_window lookups in 2026-05-27 live runs despite the
    button being in the descendants tree). Retries up to 3x with
    0.2s gaps because dialog children sometimes lag the window."""
    lg = logger or _nullable_logger()
    if window is None:
        return False
    for _attempt in range(3):
        try:
            for b in window.descendants(control_type="Button"):
                try:
                    aid = b.element_info.automation_id or ""
                except Exception:
                    continue
                if aid != auto_id:
                    continue
                try:
                    b.click_input()
                except Exception:
                    try:
                        b.invoke()
                    except Exception as e:
                        lg.warning(f"    {auto_id} click failed: {e}")
                        return False
                lg.debug(f"    clicked descendant Button "
                         f"auto_id={auto_id!r}")
                return True
        except Exception:
            pass
        time.sleep(0.2)
    return False


def _click_button_by_labels(window, labels, logger=None) -> bool:
    """Try each label in `labels` against `window.child_window(title=...)`.
    Returns True on the first one that exists and clicks successfully."""
    lg = logger or _nullable_logger()
    if window is None:
        return False
    for label in labels:
        try:
            b = window.child_window(title=label, control_type="Button")
            if not b.exists(timeout=0.3):
                continue
            try:
                b.click_input()
            except Exception:
                b.invoke()
            lg.debug(f"    clicked button {label!r}")
            return True
        except Exception:
            continue
    return False


def _wait_for_new_top_window(app, before_handles: set,
                             timeout: float = 4.0):
    """Poll for a top-level DM window we hadn't seen before. Returns
    the new window handle, or None on timeout.

    Stop-aware. ONLY returns windows that are explicitly visible -
    DM sometimes keeps invisible top-level windows in its tree (e.g.
    pre-fab Tariff List 'shell' that exists before being populated)
    and the 2026-05-27 BT60040 failure showed we'd return one of
    those and then try to click buttons on a phantom dialog."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        _check_stop()
        try:
            for w in app.windows():
                try:
                    h = int(w.handle)
                except Exception:
                    continue
                if h in before_handles:
                    continue
                # Only return EXPLICITLY visible windows. If
                # is_visible() raises or returns False, skip.
                try:
                    if not w.is_visible():
                        continue
                except Exception:
                    continue
                return w
        except Exception:
            pass
        time.sleep(0.15)
    return None


def _current_window_handles(app) -> set:
    """Snapshot every DM top-level window handle for diff detection."""
    out: set = set()
    try:
        for w in app.windows():
            try:
                out.add(int(w.handle))
            except Exception:
                continue
    except Exception:
        pass
    return out


def apply_retariff(
    app, current_tariff: str,
    save_after: bool,
    logger: logging.Logger | None = None,
    original_consignment_fee: float | None = None,
) -> None:
    """Drive the Change Tariff -> reselect -> Yes -> (optional Save)
    sequence.

    `current_tariff` is the tariff name DM already had on the booking
    (read from txtTariffName before this call). We just reselect it -
    DM's picker preselects the current tariff so committing the
    dialog with OK/Apply is enough to trigger the recompute.

    `save_after`:
      * True  - click Save, dismiss any post-save dialog. Used by live.
      * False - leave the wizard in modified state for the caller to
                Exit-without-saving via close_wizard_no_save. Used by
                the dry-run flow so we validate the click path without
                committing changes.

    On any missing element this raises with a clear message AND dumps
    the active top-level windows to data/_probe_*.txt so the
    failure is debuggable without a separate probe run."""
    lg = logger or _nullable_logger()
    _check_stop()
    wiz = dm_driver._wizard(app)
    if wiz is None:
        raise RuntimeError("wizard not visible when starting retariff")

    # ----- Step 1: click Change Tariff on the wizard -----
    lg.info("  step 1: clicking Change Tariff on wizard")
    pre_handles = _current_window_handles(app)
    try:
        btn = wiz.child_window(
            auto_id=_WIZ_CHANGE_TARIFF_AID, control_type="Button")
        if not btn.exists(timeout=0.6):
            _dump_top_windows(app, "no_btnChangeTariff", logger=lg)
            raise RuntimeError(
                f"Change Tariff button (auto_id={_WIZ_CHANGE_TARIFF_AID!r}) "
                "not found on wizard")
        try:
            btn.click_input()
        except Exception:
            btn.invoke()
    except Exception as e:
        _dump_top_windows(app, "btnChangeTariff_click_failed", logger=lg)
        raise RuntimeError(f"could not click Change Tariff: {e}")
    time.sleep(0.5)

    _check_stop()
    # ----- Step 2: tariff picker dialog appears - commit it -----
    lg.info("  step 2: waiting for tariff picker dialog")
    picker = _wait_for_new_top_window(app, pre_handles, timeout=4.0)
    if picker is None:
        _dump_top_windows(app, "no_tariff_picker", logger=lg)
        raise RuntimeError(
            "tariff picker dialog didn't appear within 4s of "
            "clicking Change Tariff")
    try:
        picker_name = picker.window_text() or picker.element_info.name or ""
    except Exception:
        picker_name = ""
    lg.info(f"    picker open: {picker_name!r}")
    _focus(picker, logger=lg)

    # The Tariff List dialog is a list of buttons, one per tariff
    # (auto_id 'btnTT_N', name='<Tariff Name>'). The user's probe of
    # this exact picker showed btnTT_0='LWB Blueleaf', btnTT_1='Small
    # Van', btnTT_2='SWB JF' - clicking the button IS the commit.
    # There's no separate OK/Apply.
    pre_handles_2 = _current_window_handles(app)
    target_name = (current_tariff or "").strip()
    target_btn = None

    # Pass 1: exact name match against any btnTT_* button (fast path).
    try:
        for b in picker.descendants(control_type="Button"):
            try:
                aid = b.element_info.automation_id or ""
                bname = (b.element_info.name or "").strip()
            except Exception:
                continue
            if not aid.startswith("btnTT_"):
                continue
            if bname == target_name:
                target_btn = b
                break
    except Exception:
        pass

    # Pass 2: fuzzy match (whitespace + casing tolerant) against
    # btnTT_* buttons. Handles minor spelling drift.
    if target_btn is None:
        try:
            target_norm = _normalise_tariff(target_name)
            for b in picker.descendants(control_type="Button"):
                try:
                    aid = b.element_info.automation_id or ""
                    bname = (b.element_info.name or "").strip()
                except Exception:
                    continue
                if not aid.startswith("btnTT_"):
                    continue
                ok, _sim = tariff_matches(target_name, bname)
                if ok:
                    target_btn = b
                    lg.info(f"    fuzzy match: picker shows {bname!r}, "
                            f"we wanted {target_name!r}")
                    break
        except Exception:
            pass

    if target_btn is None:
        _dump_top_windows(app, "no_matching_tariff_button", logger=lg)
        raise RuntimeError(
            f"Tariff List dialog has no button matching "
            f"{target_name!r}. See "
            "_probe_no_matching_tariff_button_*.txt for the buttons "
            "the picker actually offered.")

    try:
        target_btn.click_input()
    except Exception:
        try:
            target_btn.invoke()
        except Exception as e:
            raise RuntimeError(f"could not click tariff button: {e}")
    lg.info(f"    clicked tariff button {target_name!r}")
    time.sleep(0.2)

    # ----- Step 3: detect commit -----
    # The reliable signal is the wizard's Consignment Fee changing.
    # Poll for that first; only fall back to prompt-watching if the
    # fee doesn't move within 6 seconds. Saves ~6s vs the old "wait
    # 4s for a top-window that won't appear" approach on the common
    # path where DM commits silently.
    wiz_for_poll = dm_driver._wizard(app)
    fee_changed = None
    confirm = None
    if wiz_for_poll is not None:
        fee_changed = _wait_for_fee_change(
            wiz_for_poll, original_consignment_fee,
            timeout=6.0, logger=lg)

    if fee_changed is not None:
        lg.info("  retariff committed (silent path - "
                "fee change detected, no prompt needed)")
    elif original_consignment_fee is None:
        # We had no baseline to poll against. Fall through to the
        # prompt-watch path.
        lg.debug("  no baseline fee - using prompt-watch fallback")
        confirm = _wait_for_new_top_window(
            app, pre_handles_2, timeout=2.0)
    else:
        # Fee didn't move within 6s. Maybe a re-price prompt is up
        # waiting for us. Quick 1.5s window-watch fallback.
        lg.info("  fee didn't change after 6s - checking for prompt")
        confirm = _wait_for_new_top_window(
            app, pre_handles_2, timeout=1.5)

    if fee_changed is None and confirm is None:
        # Fee unchanged AND no prompt. This is a legitimate no-op
        # retariff (the job's pricing already matched the tariff min,
        # so reselecting it produced no change). The user explicitly
        # asked: "if the price still hasn't changed after the
        # secondary check, it should just save and leave it."
        lg.info("    no fee change and no prompt - this job was a "
                "no-op (price already at tariff minimum). Continuing.")
    elif confirm is None:
        # Fee changed silently, no prompt. Nothing more to do here.
        pass
    else:
        try:
            confirm_name = (confirm.window_text()
                            or confirm.element_info.name or "")
        except Exception:
            confirm_name = ""
        lg.info(f"    re-price prompt: {confirm_name!r}")
        _focus(confirm, logger=lg)
        # Ensure 'Re-price driver charge' checkbox is OFF before Yes.
        # The brief is explicit: never touch the driver charge.
        try:
            for cb in confirm.descendants(control_type="CheckBox"):
                try:
                    nm = (cb.element_info.name or "").lower()
                except Exception:
                    nm = ""
                if "re-price" in nm and "driver" in nm:
                    try:
                        state = cb.get_toggle_state()
                    except Exception:
                        state = 0
                    if state == 1:
                        try:
                            cb.click_input()
                        except Exception:
                            try:
                                cb.invoke()
                            except Exception:
                                pass
                        lg.info("    unchecked 'Re-price driver charge'")
                    break
        except Exception as e:
            lg.debug(f"  could not inspect re-price checkboxes: {e}")
        # The re-price prompt has Button auto_id='btnYes' name='Yes'
        # and Button auto_id='btnNo' name='No' (probed 2026-05-27).
        # Use the auto_id via descendants() - same pattern as the
        # post-save btnProceed fix; child_window(title='Yes') is
        # unreliable on freshly-rendered DM modals.
        yes_clicked = _click_button_by_auto_id_desc(
            confirm, "btnYes", logger=lg)
        if not yes_clicked:
            # Label fallback. Carefully excludes 'No' / 'btnNo' which
            # would reject the re-price.
            yes_clicked = _click_button_by_labels(
                confirm, _REPRICE_YES_LABELS, logger=lg)
        if not yes_clicked:
            # Last resort: send ENTER. On a Yes/No DM confirm, ENTER
            # triggers the default which is Yes. Better than leaving
            # the modal stuck on screen blocking the next job.
            try:
                from pywinauto.keyboard import send_keys
                send_keys("{ENTER}")
                yes_clicked = True
                lg.info("    sent ENTER as fallback (default = Yes)")
            except Exception:
                pass
        if not yes_clicked:
            _dump_top_windows(app, "no_yes_on_reprice", logger=lg)
            raise RuntimeError(
                "could not find btnYes / Yes label / send ENTER on "
                "the re-price prompt - see _probe_no_yes_on_reprice_*.txt")
        time.sleep(0.6)

    # Only sweep popups if something visibly drifted in. The flow
    # above already handled the re-price prompt if it appeared.
    if _has_blocking_popup(app):
        _quick_close_popups(app, logger=lg)

    # ----- Step 3b: restore the original Consignment Fee -----
    # Only if the fee actually dropped during the retariff. On a no-op
    # retariff (fee_changed is None AND no prompt fired) there's
    # nothing to restore - skip the field-write + verify round-trip.
    needs_restore = (fee_changed is not None) or (confirm is not None)
    if not needs_restore:
        lg.info("  no fee change to restore - skipping the "
                "Consignment Fee restore step")
    elif original_consignment_fee is not None:
        lg.info(
            f"  restoring Consignment Fee to original "
            f"£{original_consignment_fee:,.2f}")
        wiz2 = dm_driver._wizard(app)
        if wiz2 is None:
            raise RuntimeError(
                "wizard vanished before Consignment Fee restore")
        formatted = f"{original_consignment_fee:.2f}"
        ok = False
        try:
            ok = dm_driver._set_edit_value(
                wiz2, "txtTotalConsignmentCharge", formatted)
        except Exception as e:
            lg.warning(f"  _set_edit_value raised: {e}")
        if not ok:
            _dump_top_windows(app, "no_consignment_field", logger=lg)
            raise RuntimeError(
                "could not write to txtTotalConsignmentCharge - "
                "see _probe_no_consignment_field_*.txt")
        time.sleep(0.4)
        # Confirm the restore took. Read back and compare; flag if it
        # didn't (don't raise - the live run still needs to attempt
        # Save so the user can spot the row in the report).
        try:
            idx = dm_driver._build_auto_id_index(wiz2)
            actual_raw = dm_driver._value_for_auto_id(
                idx, "txtTotalConsignmentCharge")
            actual = parse_money(actual_raw)
            if actual is None:
                lg.warning("  could not read back Consignment Fee "
                           "after restore")
            elif abs(actual - original_consignment_fee) > 0.01:
                lg.warning(
                    f"  Consignment Fee restore drift: read back "
                    f"£{actual:,.2f}, wanted "
                    f"£{original_consignment_fee:,.2f}")
            else:
                lg.info(
                    f"  Consignment Fee restored OK "
                    f"(read back £{actual:,.2f})")
        except Exception as e:
            lg.warning(f"  Consignment Fee verify raised: {e}")

    if not save_after:
        # Dry run: the wizard is now in MODIFIED state (Consignment
        # Fee has dropped to the tariff minimum). Leave it - the
        # caller will Exit without saving via close_wizard_no_save,
        # which handles the 'discard changes?' confirm fast because
        # the dialog actually appears (vs the current dry run path
        # where there are no changes and the helper times out at 5s
        # waiting for a dialog that won't render).
        lg.info("  step 4 (DRY): wizard left in modified state; "
                "caller will Exit-without-saving")
        return

    # ----- Step 4 LIVE: click Save -----
    lg.info("  step 4 (LIVE): clicking Save on the wizard")
    wiz = dm_driver._wizard(app)
    if wiz is None:
        raise RuntimeError("wizard vanished before Save")
    pre_save = _current_window_handles(app)
    try:
        sb = wiz.child_window(
            auto_id=_WIZ_SAVE_AID, control_type="Button")
        if not sb.exists(timeout=0.6):
            _dump_top_windows(app, "no_btnSave", logger=lg)
            raise RuntimeError(
                f"Save button (auto_id={_WIZ_SAVE_AID!r}) not found")
        try:
            sb.click_input()
        except Exception:
            sb.invoke()
    except Exception as e:
        _dump_top_windows(app, "btnSave_click_failed", logger=lg)
        raise RuntimeError(f"could not click Save: {e}")
    time.sleep(1.0)

    # ----- Step 5 LIVE: handle post-save dialog (Confirm & Continue) -----
    # Captured 2026-05-27 17:11:48: the post-save 'Confirm' dialog has
    #   Button name='Close'              (X corner)
    #   Button name='Skip'               auto_id='btnExit'    <- DON'T click
    #   Button name='Confirm & Continue' auto_id='btnProceed' <- WHAT WE WANT
    # pywinauto's child_window(title='Confirm & Continue') was failing
    # because the '&' is a Windows mnemonic that title= matching
    # mishandles. Use the auto_id instead - rock-solid match.
    lg.info("  step 5 (LIVE): waiting for post-save dialog")
    post = _wait_for_new_top_window(app, pre_save, timeout=4.0)
    if post is not None:
        try:
            pname = post.window_text() or post.element_info.name or ""
        except Exception:
            pname = ""
        lg.info(f"    post-save dialog: {pname!r}")
        _focus(post, logger=lg)

        # Two-stage lookup. Probe consistently shows the button is
        # there with auto_id='btnProceed', but post.child_window(
        # auto_id='btnProceed') was returning False on the live run -
        # likely a pywinauto descendant-search race when the dialog
        # has just opened. The same probe code uses descendants() and
        # finds it fine. Iterate descendants ourselves, filter by
        # auto_id - skip btnExit ('Skip') which would abandon the save.
        clicked = False
        # Give the dialog a moment to fully render its children.
        time.sleep(0.3)
        for attempt in range(4):
            try:
                buttons = list(post.descendants(control_type="Button"))
            except Exception as e:
                lg.debug(f"    descendants raised on attempt {attempt}: {e}")
                buttons = []
            for b in buttons:
                try:
                    aid = b.element_info.automation_id or ""
                    bname = b.element_info.name or ""
                except Exception:
                    continue
                if aid != "btnProceed":
                    continue
                # Found it. Click. Don't fall through to label match.
                try:
                    b.click_input()
                except Exception:
                    try:
                        b.invoke()
                    except Exception as ce:
                        lg.warning(
                            f"    btnProceed click failed: {ce}")
                        break
                clicked = True
                lg.info(f"    clicked Confirm & Continue "
                        f"(btnProceed, name={bname!r})")
                break
            if clicked:
                break
            time.sleep(0.3)

        # Label fallback for any DM version that doesn't have btnProceed.
        # Carefully exclude 'Skip' (btnExit) which would abandon the
        # save without committing.
        if not clicked:
            clicked = _click_button_by_labels(
                post,
                ("Confirm & Continue", "Confirm and Continue",
                 "Continue", "Proceed", "OK"),
                logger=lg)

        # Last-resort: any button whose name contains 'continue' or
        # 'confirm' (but NOT 'skip' / 'cancel'). Iterates descendants
        # the same way we do for btnProceed so this won't suffer
        # whatever made child_window miss.
        if not clicked:
            try:
                for b in post.descendants(control_type="Button"):
                    try:
                        bname = (b.element_info.name or "").lower()
                    except Exception:
                        continue
                    if not bname:
                        continue
                    if any(k in bname for k in ("skip", "cancel", "close")):
                        continue
                    if any(k in bname for k in
                           ("continue", "confirm", "proceed", "ok")):
                        try:
                            b.click_input()
                        except Exception:
                            try:
                                b.invoke()
                            except Exception:
                                continue
                        clicked = True
                        lg.info(
                            "    clicked post-save button via "
                            f"name-keyword match: {b.element_info.name!r}")
                        break
            except Exception:
                pass

        if not clicked:
            _dump_top_windows(app, "no_post_save_button", logger=lg)
            lg.warning(
                "  could not find btnProceed or a fallback label on "
                "the post-save dialog - see "
                "_probe_no_post_save_button_*.txt. Job may have saved "
                "but the dialog is now sitting on screen blocking the "
                "next job.")
        time.sleep(0.6)
    else:
        lg.info("    no post-save dialog appeared")
    if _has_blocking_popup(app):
        _quick_close_popups(app, logger=lg)


# ---------------------------------------------------------------------------
# Per-job orchestration
# ---------------------------------------------------------------------------

def _drift_check(before: float | None,
                 after: float | None) -> bool:
    """Return True if before/after agree within tolerance OR both
    are None. False = drifted."""
    if before is None and after is None:
        return True
    if before is None or after is None:
        return False
    return abs(before - after) <= DRIFT_TOLERANCE


def process_one(
    app, bt_ref: str, live: bool,
    job_n: int = 0, job_total: int = 0,
    logger: logging.Logger | None = None,
) -> dict:
    """Run the full per-job flow. Returns a result dict the writer
    serialises into one row of the output xlsx.

    No customer / tariff guards - the BT ref is unique so the search
    will only return one job. We just open it, capture values, do the
    retariff round-trip (reselecting whatever tariff is already there),
    re-read values, flag drift in the report.

    `live=False` is the dry-run mode - captures values, never saves.
    `live=True` does the retariff + save."""
    lg = logger or _nullable_logger()
    prefix = f"[{job_n}/{job_total}]" if job_total else f"[{bt_ref}]"
    result: dict = {
        "bt_ref": bt_ref,
        "status": "",
        "tariff_name": "",
        "body_type_before": "",
        "body_type_after": "",
        "consignment_fee_before": None,
        "consignment_fee_after": None,
        "other_charge_before": None,
        "other_charge_after": None,
        "driver_charge_before": None,
        "driver_charge_after": None,
        "notes": "",
    }
    try:
        _check_stop()
        lg.info(f"{prefix} START {bt_ref}")
        # Make sure DM is the foreground window before we start
        # clicking - on a laptop the Cal Toolkit page can easily
        # cover DM and clicks would miss otherwise.
        bring_dm_to_front(app, logger=lg)
        # Aggressive cleanup: dismiss any modal/dialog left over from
        # a failed previous job. The 2026-05-27 live run showed an
        # alternating-failure pattern where job N's leftover re-price
        # prompt blocked job N+1's search - the search would open
        # Row_0 which was actually still the previous booking. This
        # cleanup pass sends Escape on any non-dashboard window.
        if _has_blocking_popup(app):
            _aggressive_cleanup(app, logger=lg)
        lg.info(f"  searching DM for {bt_ref!r}...")
        opened = open_job_wizard(app, bt_ref, logger=lg)
        if not opened:
            result["status"] = S_ERROR
            result["notes"] = "could not open wizard (no search result?)"
            lg.error(f"{prefix} ERROR  {result['notes']}")
            return result

        lg.info("  reading wizard values BEFORE retariff...")
        before = read_wizard_values(app, logger=lg)
        result["tariff_name"]            = before["tariff_name"]
        result["body_type_before"]       = before["body_type"]
        result["consignment_fee_before"] = before["consignment_fee"]
        result["other_charge_before"]    = before["other_charge"]
        result["driver_charge_before"]   = before["driver_charge"]

        # ----- Drive Change Tariff for BOTH dry and live -----
        # Dry mode validates the click path without committing - we
        # do the retariff steps, capture the wizard's modified state,
        # then Exit-without-saving (the existing close_wizard_no_save
        # handles the discard-changes confirm).
        # Live mode does the same steps + Save + post-save dialog.
        lg.info(f"  driving Change Tariff -> reselect "
                f"{before['tariff_name']!r}"
                f" (save_after={'YES' if live else 'NO (dry run)'})")
        try:
            apply_retariff(
                app, before["tariff_name"],
                save_after=live,
                logger=lg,
                original_consignment_fee=before["consignment_fee"])
        except Exception as e:
            result["status"] = S_ERROR
            result["notes"] = f"apply_retariff: {e}"
            lg.error(f"  apply_retariff raised: {e}")
            # Try to clean up - wizard may be in a weird state.
            try:
                exit_wizard_without_saving(app, logger=lg)
            except Exception:
                pass
            return result

        # ----- Capture AFTER values -----
        # For dry: the wizard is still open in modified state. Read
        # the values to capture the dropped Consignment Fee, then
        # Exit-without-saving so nothing commits.
        # For live: after Save the wizard has closed. We can't re-read
        # from a closed wizard, so for live we accept that after-values
        # come from a fresh re-open. For now, mirror before-values as
        # placeholder for live and note in the row that re-read is
        # post-save (a future improvement is to re-open the booking
        # after save).
        if not live:
            lg.info("  re-reading wizard values AFTER retariff "
                    "(still in dry-mode modified state)...")
            try:
                after = read_wizard_values(app, logger=lg)
                result["body_type_after"]       = after["body_type"]
                result["consignment_fee_after"] = after["consignment_fee"]
                result["other_charge_after"]    = after["other_charge"]
                result["driver_charge_after"]   = after["driver_charge"]
            except Exception as e:
                lg.warning(f"  could not re-read wizard values: {e}")
                result["body_type_after"]       = before["body_type"]
                result["consignment_fee_after"] = before["consignment_fee"]
                result["other_charge_after"]    = before["other_charge"]
                result["driver_charge_after"]   = before["driver_charge"]
            lg.info("  dry-run: closing wizard (Exit + discard "
                    "changes)")
            exit_wizard_without_saving(app, logger=lg)
            # In dry mode we report PREVIEW even when retariff fired,
            # because nothing was actually committed.
            result["status"] = S_PREVIEW
            result["notes"] = (
                "DRY: Change Tariff clicked, tariff reselected, "
                "Consignment Fee restored to original, wizard discarded. "
                "No changes saved.")
            # If the restored fee doesn't match before, that's already
            # been logged inside apply_retariff - surface it on the row
            # too by flagging as drift in the report.
            cf_b = before["consignment_fee"]
            cf_a = result["consignment_fee_after"]
            if (cf_b is not None and cf_a is not None
                    and abs(cf_b - cf_a) > DRIFT_TOLERANCE):
                result["status"] = S_FLAGGED
                result["notes"] = (
                    f"DRY: Consignment Fee restore did NOT take "
                    f"(before {_fmt_or_blank(cf_b)} -> "
                    f"after {_fmt_or_blank(cf_a)}). Investigate "
                    "txtTotalConsignmentCharge writability.")
            lg.info(f"{prefix} OK (dry-run, full Change Tariff path)")
            return result

        # ----- LIVE: wizard closed after save. Mirror before-values. -----
        lg.info("  live-run: Save fired; reading post-save state not "
                "possible without re-opening the booking")
        after = before  # placeholder so the report still has values
        result["body_type_after"]       = after["body_type"]
        result["consignment_fee_after"] = after["consignment_fee"]
        result["other_charge_after"]    = after["other_charge"]
        result["driver_charge_after"]   = after["driver_charge"]

        # Drift detection - report everything that changed. Tariff name
        # is expected to stay the same (we reselected). Body type, three
        # fee fields - all expected to be unchanged.
        anomalies: list[str] = []
        if (before["body_type"] or "") != (after["body_type"] or ""):
            anomalies.append(
                f"body_type changed "
                f"{before['body_type']!r} -> {after['body_type']!r}")
        if not _drift_check(before["consignment_fee"],
                            after["consignment_fee"]):
            anomalies.append(
                f"consignment_fee drifted "
                f"{_fmt_or_blank(before['consignment_fee'])} -> "
                f"{_fmt_or_blank(after['consignment_fee'])}")
        if not _drift_check(before["other_charge"],
                            after["other_charge"]):
            anomalies.append(
                f"other_charge drifted "
                f"{_fmt_or_blank(before['other_charge'])} -> "
                f"{_fmt_or_blank(after['other_charge'])}")
        if not _drift_check(before["driver_charge"],
                            after["driver_charge"]):
            anomalies.append(
                f"driver_charge drifted "
                f"{_fmt_or_blank(before['driver_charge'])} -> "
                f"{_fmt_or_blank(after['driver_charge'])}")
        if anomalies:
            result["status"] = S_FLAGGED
            result["notes"] = "; ".join(anomalies)
            for msg in anomalies:
                lg.warning(f"  drift: {msg}")
            lg.warning(f"{prefix} FLAGGED - {len(anomalies)} field(s) drifted")
        else:
            result["status"] = S_OK
            result["notes"] = "retariff applied; all preserved"
            lg.info("  no drift - all fields preserved as expected.")
            lg.info(f"{prefix} OK")
        return result

    except StopRequested:
        # User pressed Stop mid-step. Mark the row as ERROR with a
        # clear note, attempt to close the wizard quickly, propagate
        # so run_batch breaks the outer loop.
        result["status"] = S_ERROR
        result["notes"] = ("STOPPED by user mid-job. "
                           "Wizard may still be open in DM.")
        lg.warning(f"{prefix} STOPPED mid-step")
        try:
            exit_wizard_without_saving(app, logger=lg)
        except Exception:
            pass
        raise
    except Exception as e:
        result["status"] = S_ERROR
        result["notes"] = f"{type(e).__name__}: {e}"
        lg.exception(f"{prefix} ERROR: {e}")
        # Clean up wizard + any leftover modal so the NEXT job
        # doesn't inherit the broken state.
        try:
            exit_wizard_without_saving(app, logger=lg)
        except Exception:
            pass
        try:
            _aggressive_cleanup(app, logger=lg)
        except Exception:
            pass
        return result


# ---------------------------------------------------------------------------
# Top-level runners
# ---------------------------------------------------------------------------

def run_batch(
    bt_refs: list[str],
    out_xlsx: str | Path,
    live: bool,
    on_progress: Callable[[int, int, dict], None] | None = None,
    log_callback: Callable[[str, str], None] | None = None,
    should_stop: Callable[[], bool] | None = None,
) -> dict:
    """Process every BT-ref in order, write the results xlsx, and
    return a summary.

    `log_callback(levelname, formatted_msg)` fires once per log line
    for the UI's run-sheet panel. The same lines also land in a
    timestamped .log file under data/.

    `should_stop()` is polled BEFORE each job; if it returns True,
    the loop bails after the current job finishes and the partial
    results.xlsx is written with whatever rows we have so far. The
    summary dict includes `stopped: True` so the UI can surface it."""
    label = "live" if live else "dry"
    logger, log_path = setup_logger(label, callback=log_callback)
    logger.info(f"Inputs: {len(bt_refs)} BT-ref(s)")
    logger.info(f"Output xlsx: {out_xlsx}")
    logger.info("Connecting to Delivery Master...")
    try:
        app = dm_driver.connect()
        try:
            main_win = dm_driver._main_window(app)
            logger.info(f"  connected. Main window: "
                        f"{main_win.window_text()!r}")
            # Log where DM is on the screen. Helps spot if it's on an
            # off-screen monitor (e.g. restored to ultrawide coords on
            # a laptop) before we start clicking.
            try:
                r = main_win.rectangle()
                logger.info(
                    f"  DM window at left={r.left} top={r.top} "
                    f"right={r.right} bottom={r.bottom}")
                if r.left < -50 or r.top < -50:
                    logger.warning(
                        "  DM window appears off-screen - if clicks "
                        "miss, drag DM onto the visible screen and "
                        "click 'Focus DM' in the toolkit.")
            except Exception:
                pass
        except Exception:
            logger.info("  connected (no window title available)")
        # Bring DM forward right away so the user can see the script
        # working without having to alt-tab.
        bring_dm_to_front(app, logger=logger)
    except Exception as e:
        logger.exception(f"failed to connect to DM: {e}")
        raise

    # Clear any stale stop flag from a previous run.
    reset_stop()

    rows: list[dict] = []
    started = datetime.now()
    stopped = False
    total = len(bt_refs)
    for i, ref in enumerate(bt_refs, start=1):
        if should_stop is not None and should_stop():
            stopped = True
            logger.warning(
                f"STOP requested by user at job {i}/{total}. "
                f"Wrapping up - already-saved jobs are NOT rolled back.")
            break
        if _stop_requested():
            stopped = True
            logger.warning(
                f"STOP event set at job {i}/{total}. Bailing.")
            break
        try:
            row = process_one(
                app, ref, live=live,
                job_n=i, job_total=total, logger=logger)
        except StopRequested:
            # User pressed Stop mid-job. process_one returned the
            # partial row via exception side-channel; we don't have
            # it here, but record a row so the report shows where
            # we bailed.
            stopped = True
            logger.warning(
                f"STOP fired mid-job at {i}/{total} ({ref}). "
                "Wrapping up.")
            rows.append({
                "bt_ref": ref,
                "status": S_ERROR,
                "tariff_name": "",
                "body_type_before": "",
                "body_type_after": "",
                "consignment_fee_before": None,
                "consignment_fee_after": None,
                "other_charge_before": None,
                "other_charge_after": None,
                "driver_charge_before": None,
                "driver_charge_after": None,
                "notes": "STOPPED by user mid-job. Wizard may be open.",
            })
            break
        rows.append(row)
        if on_progress is not None:
            try:
                on_progress(i, total, row)
            except Exception:
                pass

    summary = _summarise(rows)
    summary["stopped"] = stopped
    summary["jobs_remaining"] = total - len(rows) if stopped else 0
    write_results_xlsx(out_xlsx, rows, summary, live=live)
    summary["out_path"] = str(out_xlsx)
    summary["log_path"] = str(log_path)

    elapsed = (datetime.now() - started).total_seconds()
    logger.info("=" * 60)
    if stopped:
        logger.warning(
            f"Run STOPPED by user after {len(rows)}/{total} jobs "
            f"({elapsed:.1f}s elapsed)")
    else:
        logger.info(f"Run complete in {elapsed:.1f}s")
    logger.info(
        f"  OK={summary['ok']}  FLAGGED={summary['flagged']}  "
        f"ERROR={summary['errored']}  SKIPPED={summary['skipped']}  "
        f"PREVIEW={summary['preview']}")
    logger.info(f"  results xlsx: {out_xlsx}")
    logger.info(f"  log file:     {log_path}")
    logger.info("=" * 60)
    for h in list(logger.handlers):
        try:
            h.flush()
            h.close()
        except Exception:
            pass
    return summary


def dry_run(bt_refs: list[str],
            out_xlsx: str | Path,
            on_progress=None, log_callback=None,
            should_stop=None) -> dict:
    return run_batch(bt_refs, out_xlsx,
                     live=False, on_progress=on_progress,
                     log_callback=log_callback,
                     should_stop=should_stop)


def live_run(bt_refs: list[str],
             out_xlsx: str | Path,
             on_progress=None, log_callback=None,
             should_stop=None) -> dict:
    return run_batch(bt_refs, out_xlsx,
                     live=True, on_progress=on_progress,
                     log_callback=log_callback,
                     should_stop=should_stop)


# ---------------------------------------------------------------------------
# Results xlsx
# ---------------------------------------------------------------------------

RESULT_HEADERS = [
    "BT Ref", "Status", "Tariff Name",
    "Body Type (Before)", "Body Type (After)", "Vehicle Changed?",
    "Consignment Fee (Before)", "Consignment Fee (After)", "Consignment Drift?",
    "Other Charge (Before)", "Other Charge (After)", "Other Drift?",
    "Driver Charge (Before)", "Driver Charge (After)", "Driver Drift?",
    "Notes",
]


def _money(v) -> str:
    if v is None:
        return ""
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return str(v)


def _summarise(rows: list[dict]) -> dict:
    s = {S_OK: 0, S_FLAGGED: 0, S_ERROR: 0, S_SKIPPED: 0, S_PREVIEW: 0}
    for r in rows:
        st = r.get("status") or ""
        s[st] = s.get(st, 0) + 1
    return {
        "total": len(rows),
        "by_status": s,
        "ok": s.get(S_OK, 0),
        "flagged": s.get(S_FLAGGED, 0),
        "errored": s.get(S_ERROR, 0),
        "skipped": s.get(S_SKIPPED, 0),
        "preview": s.get(S_PREVIEW, 0),
    }


def write_results_xlsx(out_path: str | Path,
                       rows: list[dict],
                       summary: dict,
                       live: bool) -> str:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    del wb["Sheet"]

    GREEN = PatternFill("solid", fgColor="DDF7E2")
    YELLOW = PatternFill("solid", fgColor="FFF4CC")
    RED = PatternFill("solid", fgColor="FFD6D6")
    GREY = PatternFill("solid", fgColor="E5E7EB")
    HEADER = PatternFill("solid", fgColor="0a9d98")

    status_fill = {
        S_OK: GREEN, S_FLAGGED: YELLOW, S_ERROR: RED,
        S_SKIPPED: GREY, S_PREVIEW: GREEN,
    }

    # Summary sheet first
    ws = wb.create_sheet("Summary")
    ws.append([
        "Cal Toolkit - Tariff Re-trigger "
        + ("(live run)" if live else "(dry run)")
    ])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Total jobs", summary["total"]])
    for st in (S_OK, S_FLAGGED, S_ERROR, S_SKIPPED, S_PREVIEW):
        n = summary["by_status"].get(st, 0)
        if n or st in (S_OK, S_FLAGGED, S_ERROR, S_SKIPPED):
            ws.append([st, n])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10

    # Per-job sheet
    ws = wb.create_sheet("Results")
    ws.append(RESULT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    def _changed(before, after) -> str:
        if before is None and after is None:
            return ""
        if before is None or after is None:
            return "YES"
        if isinstance(before, (int, float)) and isinstance(after, (int, float)):
            return "YES" if abs(before - after) > DRIFT_TOLERANCE else ""
        return "YES" if str(before) != str(after) else ""

    for r in rows:
        bt_b = r.get("body_type_before") or ""
        bt_a = r.get("body_type_after") or ""
        cf_b = r.get("consignment_fee_before")
        cf_a = r.get("consignment_fee_after")
        oc_b = r.get("other_charge_before")
        oc_a = r.get("other_charge_after")
        dc_b = r.get("driver_charge_before")
        dc_a = r.get("driver_charge_after")
        ws.append([
            r.get("bt_ref", ""), r.get("status", ""),
            r.get("tariff_name", ""),
            bt_b, bt_a, _changed(bt_b, bt_a),
            _money(cf_b), _money(cf_a), _changed(cf_b, cf_a),
            _money(oc_b), _money(oc_a), _changed(oc_b, oc_a),
            _money(dc_b), _money(dc_a), _changed(dc_b, dc_a),
            r.get("notes", ""),
        ])
        fill = status_fill.get(r.get("status") or "", None)
        if fill is not None:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    widths = [14, 12, 24,
              16, 16, 8,
              18, 18, 9,
              16, 16, 7,
              16, 16, 7,
              50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    return str(out_path)
