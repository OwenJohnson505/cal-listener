"""
Customer 360 Sync engine.

Pure read-only of the DM 'Customer List' xlsx export (16 columns,
roughly 1,400 rows per depot). Filters by Start Date or Last Trading
Date, bucks each row into one of:

    - new        - no Customer 360 profile matches the customer name
    - updated    - profile exists but email/address differs in the
                   export
    - unchanged  - profile exists and matches what's already on file

Writes back via customer_profile_store.merge_profile when the UI
calls save_new() / save_updates(). cloud_sync.upsert_rows is
triggered by merge_profile so colleagues see the new/updated rows on
their next refresh.

Public surface:

    read_customer_list(path)               -> list[dict]
    filter_by_date_range(rows, start, end,
                         column='start_date') -> list[dict]
    detect_depot_from_path(path)           -> 'North' | 'South' | ''
    categorize(rows, depot)                -> (new, updated, unchanged)
    save_new(rows, depot)                  -> (saved, errors)
    save_updates(rows)                     -> (saved, errors)
"""
from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path
from typing import Iterable


# ---------------------------------------------------------------------------
# Column map - matches DM's 'Customer List' export. If DM ever renames
# any of these, only this dict needs editing.
# ---------------------------------------------------------------------------

COL_CUSTOMER_NAME      = "Customer Name"
COL_ACCOUNT_CODE       = "Account Code"
COL_SALES_PERSON       = "Sales Person"
COL_START_DATE         = "Start Date"
COL_NOMINAL_CODE       = "Nominal Code"
COL_ADDRESS_1          = "Address Line 1"
COL_ADDRESS_2          = "Address Line 2"
COL_POSTCODE           = "Postcode"
COL_TOWN               = "Town"
COL_COUNTY             = "County"
COL_COUNTRY            = "Country"
COL_TEL                = "Tel No"
COL_MOBILE             = "Mobile No"
COL_EMAIL              = "Email"
COL_CONTACTS           = "Contacts"
COL_LAST_TRADING_DATE  = "Last Trading Date"

ALL_COLS = [
    COL_CUSTOMER_NAME, COL_ACCOUNT_CODE, COL_SALES_PERSON,
    COL_START_DATE, COL_NOMINAL_CODE, COL_ADDRESS_1, COL_ADDRESS_2,
    COL_POSTCODE, COL_TOWN, COL_COUNTY, COL_COUNTRY, COL_TEL,
    COL_MOBILE, COL_EMAIL, COL_CONTACTS, COL_LAST_TRADING_DATE,
]


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def find_customer_list(downloads: Path | None = None) -> Path | None:
    """Auto-discover the most recent DM customer-list xlsx in the
    user's Downloads folder. DM's export filename pattern is
    '<Depot> Customer List TMS-<hash>.xlsx' so we glob on
    '*Customer List TMS*.xlsx'. Returns None if nothing found."""
    if downloads is None:
        downloads = Path.home() / "Downloads"
    if not downloads.exists():
        return None
    candidates = list(downloads.glob("*Customer List TMS*.xlsx"))
    if not candidates:
        candidates = list(downloads.glob("*ustomer*ist*.xlsx"))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def detect_depot_from_path(path: str | Path) -> str:
    """DM's export filename leads with the depot name ('North
    Customer List TMS-...'). Return 'North' or 'South' if the
    filename starts with one of those, otherwise empty string so the
    UI prompts the user to pick."""
    stem = Path(str(path)).stem.lower()
    if stem.startswith("north"):
        return "North"
    if stem.startswith("south"):
        return "South"
    # Also tolerate the depot appearing anywhere in the name in case
    # the user has renamed the file.
    if "north" in stem:
        return "North"
    if "south" in stem:
        return "South"
    return ""


# ---------------------------------------------------------------------------
# Reader
# ---------------------------------------------------------------------------

def _coerce_date(v) -> datetime | None:
    """openpyxl returns datetime for date-typed cells; if the cell was
    typed as text we try parsing common UK formats. Returns None if
    the value can't be coerced."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y",
                "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def read_customer_list(path: str | Path) -> list[dict]:
    """Read every row of the DM customer-list xlsx and return a list
    of dicts keyed by the original column name. Dates are returned
    as datetime objects (Start Date, Last Trading Date) or None when
    the cell is blank.

    The reader is forgiving about column order - it looks each
    column up by header name, so DM can add new columns at the end
    without breaking us. Missing required columns raise ValueError
    with a helpful list of what was actually found."""
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else ""
                   for h in header_row]
        # Build name -> index lookup. Case-insensitive in case DM
        # capitalises differently between exports.
        idx_by_name: dict[str, int] = {}
        for i, h in enumerate(headers):
            if h:
                idx_by_name[h.lower()] = i
        required = [COL_CUSTOMER_NAME, COL_START_DATE]
        missing = [c for c in required if c.lower() not in idx_by_name]
        if missing:
            raise ValueError(
                f"Customer list xlsx is missing required column(s): "
                f"{missing}. Found columns: {headers}")

        def _get(row: tuple, col: str):
            i = idx_by_name.get(col.lower())
            if i is None:
                return None
            return row[i] if i < len(row) else None

        out: list[dict] = []
        for r in rows_iter:
            if r is None:
                continue
            name = _get(r, COL_CUSTOMER_NAME)
            if name is None or str(name).strip() == "":
                continue
            out.append({
                "customer_name":    str(name).strip(),
                "account_code":     _get(r, COL_ACCOUNT_CODE),
                "sales_person":     _get(r, COL_SALES_PERSON),
                "start_date":       _coerce_date(_get(r, COL_START_DATE)),
                "nominal_code":     _get(r, COL_NOMINAL_CODE),
                "address_1":        _get(r, COL_ADDRESS_1),
                "address_2":        _get(r, COL_ADDRESS_2),
                "postcode":         _get(r, COL_POSTCODE),
                "town":             _get(r, COL_TOWN),
                "county":           _get(r, COL_COUNTY),
                "country":          _get(r, COL_COUNTRY),
                "tel":              _get(r, COL_TEL),
                "mobile":           _get(r, COL_MOBILE),
                "email":            (str(_get(r, COL_EMAIL)).strip()
                                     if _get(r, COL_EMAIL) else ""),
                "contacts":         _get(r, COL_CONTACTS),
                "last_trading_date": _coerce_date(
                    _get(r, COL_LAST_TRADING_DATE)),
            })
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Date filter
# ---------------------------------------------------------------------------

DATE_FIELD_START   = "start_date"
DATE_FIELD_LAST    = "last_trading_date"


def filter_by_date_range(rows: Iterable[dict],
                          start: date | datetime,
                          end: date | datetime,
                          column: str = DATE_FIELD_START
                          ) -> list[dict]:
    """Return only those rows where the specified date column falls
    inclusively between start and end. Rows with a blank value in
    that column are dropped (they aren't 'in range' for any range).

    The column argument is one of:
        'start_date'         - customers created in the window (default)
        'last_trading_date'  - customers active in the window"""
    if isinstance(start, datetime):
        start_d = start.date()
    else:
        start_d = start
    if isinstance(end, datetime):
        end_d = end.date()
    else:
        end_d = end
    out: list[dict] = []
    for r in rows:
        v = r.get(column)
        if v is None:
            continue
        if isinstance(v, datetime):
            d = v.date()
        elif isinstance(v, date):
            d = v
        else:
            continue
        if d < start_d or d > end_d:
            continue
        out.append(r)
    return out


# ---------------------------------------------------------------------------
# Categorisation
# ---------------------------------------------------------------------------

def _import_profile_store():
    try:
        import customer_profile_store as cps  # type: ignore
        return cps
    except Exception as e:
        print(f"[customer_360_sync] cps import failed: {e}",
              file=sys.stderr, flush=True)
        return None


def _profile_email(prof: dict) -> str:
    """Read a profile's stored invoice email. Lives under
    tms_references.email (see customer_profile_store schema)."""
    if not isinstance(prof, dict):
        return ""
    refs = prof.get("tms_references") or {}
    if isinstance(refs, dict):
        return (refs.get("email") or "").strip()
    return ""


def categorize(rows: list[dict], depot: str
               ) -> tuple[list[dict], list[dict], list[dict]]:
    """Bucket rows by whether they exist in Customer 360 and whether
    the export's email differs from what's on file. Returns
    (new, updated, unchanged). Each row in 'updated' is augmented
    with diff_summary + existing_profile_primary_name so the UI can
    show 'old -> new' on the row."""
    cps = _import_profile_store()
    new: list[dict] = []
    updated: list[dict] = []
    unchanged: list[dict] = []
    if cps is None:
        # Without the profile store everything is treated as 'new'
        # so the UI still does something useful. The save step will
        # also fail loudly in that case.
        return list(rows), [], []
    for r in rows:
        name = (r.get("customer_name") or "").strip()
        if not name:
            continue
        prof = None
        try:
            prof = cps.find_profile_by_tms_name(name)
        except Exception:
            prof = None
        if prof is None:
            new.append(r)
            continue
        # Profile exists. Diff the export's email against the stored
        # email and emit 'updated' iff they differ.
        existing_email = _profile_email(prof)
        export_email = (r.get("email") or "").strip()
        diffs = []
        if export_email and export_email.lower() != existing_email.lower():
            diffs.append(f"email: {existing_email!r} -> {export_email!r}")
        existing_depot = (prof.get("depot") or "").strip()
        if depot and existing_depot and existing_depot != depot:
            diffs.append(f"depot: {existing_depot!r} -> {depot!r}")
        # Augment row with profile context.
        r_aug = dict(r)
        r_aug["existing_primary_name"] = (
            prof.get("primary_name") or name).strip()
        r_aug["existing_email"]        = existing_email
        r_aug["diff_summary"]          = "; ".join(diffs)
        if diffs:
            updated.append(r_aug)
        else:
            unchanged.append(r_aug)
    return new, updated, unchanged


# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------

def _row_to_profile_partial(r: dict, depot: str) -> dict:
    """Build the profile-store partial dict from an export row.
    primary_name and clearbooks_name stay in lockstep (the migration
    we ran earlier set this convention). Email/address land in
    tms_references."""
    name = (r.get("customer_name") or "").strip()
    partial: dict = {
        "primary_name":    name,
        "clearbooks_name": name,
        "depot":           depot or "",
        "tms_names":       [name],
    }
    refs: dict = {}
    if r.get("email"):
        refs["email"] = r["email"]
    addr_parts = [r.get("address_1"), r.get("address_2"),
                  r.get("town"), r.get("postcode")]
    addr = ", ".join(str(p).strip()
                     for p in addr_parts
                     if p and str(p).strip())
    if addr:
        refs["address"] = addr
    if r.get("tel"):
        refs["phone"] = str(r["tel"])
    if r.get("start_date") and isinstance(r["start_date"], datetime):
        refs["start_date"] = r["start_date"].date().isoformat()
    if refs:
        partial["tms_references"] = refs
    return partial


def save_new(rows: list[dict], depot: str
             ) -> tuple[int, list[str]]:
    """Create a Customer 360 profile for each row. Returns
    (saved_count, errors). Errors are per-row strings so the UI can
    surface them; the loop doesn't abort on failure."""
    cps = _import_profile_store()
    if cps is None:
        return 0, ["customer_profile_store import failed; cannot save"]
    saved = 0
    errors: list[str] = []
    for r in rows:
        name = (r.get("customer_name") or "").strip()
        if not name:
            continue
        try:
            cps.merge_profile(name, _row_to_profile_partial(r, depot))
            saved += 1
        except Exception as e:
            errors.append(f"{name}: {e}")
    return saved, errors


def save_updates(rows: list[dict]) -> tuple[int, list[str]]:
    """Apply only the diffed fields (currently email + depot) to the
    existing profiles. Doesn't touch other fields the user may have
    set in Customer 360. Returns (saved_count, errors)."""
    cps = _import_profile_store()
    if cps is None:
        return 0, ["customer_profile_store import failed; cannot save"]
    saved = 0
    errors: list[str] = []
    for r in rows:
        existing = (r.get("existing_primary_name") or "").strip()
        if not existing:
            continue
        partial: dict = {}
        new_email = (r.get("email") or "").strip()
        if new_email and new_email.lower() != (
                r.get("existing_email") or "").lower():
            partial["tms_references"] = {"email": new_email}
        # depot may have changed too if the user reassigned the row.
        # We don't currently surface a depot column in the update tab
        # but the engine respects an 'override_depot' key if present.
        if r.get("override_depot"):
            partial["depot"] = r["override_depot"]
        if not partial:
            continue
        try:
            cps.merge_profile(existing, partial)
            saved += 1
        except Exception as e:
            errors.append(f"{existing}: {e}")
    return saved, errors
